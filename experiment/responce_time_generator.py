#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Import modules
import random 

import numpy as np
from openpyxl import load_workbook
import pandas as pd

import template_experiment as tp

# NOTE: success seeds are slf: '5, 8, 17, 20'
#                         obs: '28'
SEED = 5
GAME_PAT = 5
BLOCK_NUM = 4
OBS_TRIAL_NUM = 9
MEAN = [30, 40, 50]
SD = 7
IMG_TIME_SD = 0.5
CONF_TIME_SD = 0.5

LEFT = 'left'
RIGHT = 'right'

ALPHA = 0.216
BETA = 0.141

IMG_MEAN_TIME = np.array([[1.367, 1.270, 1.153], 
                 [1.242, 1.235, 1.031],
                 [1.060, 1.022, 0.826]])
IMG_MIN_TIME = np.array([[0.431, 0.356, 0.421], 
                [0.512, 0.463, 0.367],
                [0.514, 0.409, 0.404]])
IMG_MAX_TIME = np.array([[2.856, 2.689, 3.121],
                [3.420, 3.033, 3.164],
                [3.020, 2.818, 2.476]])

CONF_MEAN_TIME = np.array([[0.550, 0.605, 0.482],
                  [0.476, 0.512, 0.608],
                  [0.451, 0.423, 0.372]])
CONF_MIN_TIME = np.array([[0.027, 0.065, 0.033],
                 [0.019, 0.042, 0.033],
                 [0.018, 0.022, 0.033]])
CONF_MAX_TIME = np.array([[1.233, 1.289, 1.456],
                 [1.806, 1.340, 1.866],
                 [1.616, 1.679, 1.117]])

# Fix seed
random.seed(SEED)
np.random.seed(seed=SEED)

def interpolate_time(times):
    time1 = times[0]
    time2 = (2/3)*times[0] + (1/3)*times[1]
    time3 = (1/3)*times[1] + (2/3)*times[2]
    time4 = times[2]
    new_times = np.array([time1, time2, time3, time4])

    return new_times

def responce_time_generator(
                        img_mean_time, img_min_time, img_max_time, 
                        conf_mean_time, conf_min_time, conf_max_time, 
                        pair_pats, trial_num):
    img_times = []
    conf_times = []

    img_time = -1
    conf_time = -1
    for block, pair_pats_ in enumerate(pair_pats):
        for pair_pat in pair_pats_:
            while True:
                img_time = np.random.normal(img_mean_time[block][pair_pat-1], IMG_TIME_SD)
                conf_time = np.random.normal(conf_mean_time[block][pair_pat-1], CONF_TIME_SD)
                if (img_min_time[block][pair_pat-1] <= img_time <= img_max_time[block][pair_pat-1]) and (conf_min_time[block][pair_pat-1] <= conf_time <= conf_max_time[block][pair_pat-1]):
                    break
            img_times.append(img_time)
            conf_times.append(conf_time)
    img_times = np.array(img_times)
    conf_times = np.array(conf_times)
    return img_times, conf_times

def load_data(game_pat):
    pair_pats = []
    order_pats = []
    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        df = (pd.read_excel(f'{condition_dir_path}game{game_pat}/obs{game_pat}{block+1}.xlsx', sheet_name='condition'))
        pair_pats.append(df['obs_pair_pat'].values)
        order_pats.append(df['obs_order_pat'].values)
    pair_pats = np.array(pair_pats)
    order_pats = np.array(order_pats)

    return pair_pats, order_pats


def write_excel(img_times, conf_times, game_pat):
    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        condition_file_path = f'{condition_dir_path}game{game_pat}/obs{game_pat}{block+1}.xlsx'
        wb = load_workbook(condition_file_path)
        ws = wb['condition']
        for trial in range(OBS_TRIAL_NUM):
            ws[f'J{trial+2}'] = img_times[BLOCK_NUM*block+trial]
            ws[f'K{trial+2}'] = conf_times[BLOCK_NUM*block+trial]
        wb.save(condition_file_path)


def main():
    game_pat = GAME_PAT
    pair_pats, order_pats = load_data(game_pat)
    img_mean_time = interpolate_time(IMG_MEAN_TIME)
    img_min_time = interpolate_time(IMG_MIN_TIME)
    img_max_time = interpolate_time(IMG_MAX_TIME)
    conf_mean_time = interpolate_time(CONF_MEAN_TIME)
    conf_min_time = interpolate_time(CONF_MIN_TIME)
    conf_max_time = interpolate_time(CONF_MAX_TIME)
    img_times, conf_times = responce_time_generator(
                                img_mean_time, img_min_time, img_max_time, 
                                conf_mean_time, conf_min_time, conf_max_time, 
                                pair_pats, OBS_TRIAL_NUM)
    print(img_times)
    print(conf_times)
    write_excel(img_times, conf_times, game_pat)


if __name__ == '__main__':
    main()


