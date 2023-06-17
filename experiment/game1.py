#!/usr/bin/env python
# -*- coding: utf-8 -*-

import psychopy
psychopy.useVersion('2022.2.5')


# --- Import packages ---
from psychopy import locale_setup
from psychopy import prefs
from psychopy import sound, gui, visual, core, data, event, logging, clock, colors, layout
from psychopy.constants import (NOT_STARTED, STARTED, PLAYING, PAUSED,
                                STOPPED, FINISHED, PRESSED, RELEASED, FOREVER)

import numpy as np  # whole numpy lib is available, prepend 'np.'
from numpy import (sin, cos, tan, log, log10, pi, average,
                   sqrt, std, deg2rad, rad2deg, linspace, asarray)
from numpy.random import random, randint, normal, shuffle, choice as randchoice
import os  # handy system and path functions
import sys  # to get file system encoding

import psychopy.iohub as io
from psychopy.hardware import keyboard

# Global variables
# Time
IMG_TIME = 5.0
CONF_TIME = 3.0
REWARD_TIME = 2.0
FEEDBACK_TIME = 1.0

# Size and position
IMG_SIZE = 200
CONF_SIZE = 200


# Ensure that relative paths start from the same directory as this script
_thisDir = os.path.dirname(os.path.abspath(__file__))
os.chdir(_thisDir)
# Store info about the experiment session
psychopyVersion = '2022.2.4'
expName = 'Pattern_1'  # from the Builder filename that created this script
expInfo = {
    'participant': '10_2_yokoi',
}
# --- Show participant info dialog --
dlg = gui.DlgFromDict(dictionary=expInfo, sortKeys=False, title=expName)
if dlg.OK == False:
    core.quit()  # user pressed cancel
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName
expInfo['psychopyVersion'] = psychopyVersion

# Data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
filename = _thisDir + os.sep + u'data/%s_%s_%s' % (expInfo['participant'], expName, expInfo['date'])

# An ExperimentHandler isn't essential but helps with data saving
thisExp = data.ExperimentHandler(name=expName, version='',
    extraInfo=expInfo, runtimeInfo=None,
    originPath='/Users/strix_uralensis/Documents/Experiment/original/Exp_Pattern1.py',
    savePickle=True, saveWideText=True,
    dataFileName=filename)
logging.console.setLevel(logging.WARNING)  # this outputs to the screen, not a file

endExpNow = False  # flag for 'escape' or other condition => quit the exp
frameTolerance = 0.001  # how close to onset before 'same' frame

# Start Code - component code to be run after the window creation

# --- Setup the Window ---
win = visual.Window(
    size=[1920, 1080], fullscr=True, screen=0, 
    winType='pyglet', allowStencil=False,
    monitor='screen1', color=[-1.000,-1.000,-1.000], colorSpace='rgb',
    blendMode='avg', useFBO=True, 
    units='pix')
win.mouseVisible = False
# store frame rate of monitor if we can measure it
expInfo['frameRate'] = win.getActualFrameRate()
if expInfo['frameRate'] != None:
    frameDur = 1.0 / round(expInfo['frameRate'])
else:
    frameDur = 1.0 / 60.0  # could not measure, so guess
# --- Setup input devices ---
ioConfig = {}

# Setup iohub keyboard
ioConfig['Keyboard'] = dict(use_keymap='psychopy')

ioSession = '1'
if 'session' in expInfo:
    ioSession = str(expInfo['session'])
ioServer = io.launchHubServer(window=win, **ioConfig)
eyetracker = None

# create a default keyboard (e.g. to check for escape)
defaultKeyboard = keyboard.Keyboard(backend='iohub')

"""=================================================================================================================="""
# --- Initialize components for Routine "Init" ---
# Run 'Begin Experiment' code from INIT_CODE
# Import modules
import datetime
import numpy as np
from openpyxl import Workbook, load_workbook
import pandas as pd
import scipy.io as io

# Initial settings
win.mouseVisible = False
choice_time_max = 5.0
conf_time_max   = 3.0
feedback_time   = 1.0
reward_time     = 2.0

#choice_time_max = .3
#conf_time_max   = .3
#feedback_time   = .3
#reward_time     = .3

left_key   = 'left'
center_key = 'down'
right_key  = 'right'

# Information of the experiment dialog 
participant = expInfo["participant"]

# Number of trials in each stage
slf_rept  = 36
obs_rept  = 36
test_rept = 90

# Serial number
game_serial_num = 1

slf_serial_num  = 0
obs_serial_num  = 0
test_serial_num = 0

# Block counter
block_num = 1

# Pattern of slf & obs are zero
slf_seq_pattern = 0
obs_seq_pattern = 0

# Trial color
slf_color  = "lime"
obs_color  = "cyan"
test_color = "orange"

# Determine the images based on the excel file
loc_list  = [30, 40, 50]

idx_list = [1, 2, 3]
# Set images
slf_img = ['slf11.bmp', 'slf12.bmp', 'slf13.bmp']
obs_img = ['obs11.bmp', 'obs12.bmp', 'obs13.bmp']
test_img = slf_img + obs_img

slf_dict = dict(zip(loc_list, slf_img))
obs_dict = dict(zip(loc_list, obs_img))

# Create files to store data
get_current_time = datetime.datetime.now() # Get time in the form of "yyyy-mm-dd hh:mm:ss"
out_name = f'subj_{participant}_1_{get_current_time:%y%m%d%H%M}' 
            # participant_yyyymmddというファイル名

# Define sheet names and data table
slf_sheet = "slf1"
obs_sheet = "obs1"
test_sheet = "test1"

slf_data_table =  [[0] * 24 for i in range(slf_rept)]
obs_data_table =  [[0] * 24 for i in range(obs_rept)]
test_data_table = [[0] * 24 for i in range(test_rept)]

# Create matlab file(dict)
out_dict = {'slf1':slf_data_table, 'obs1':obs_data_table, 'test1':test_data_table}

# Create excel file
out_xlsx = out_name + ".xlsx"
out_book = Workbook()
#out_book = load_workbook(filename=out_xlsx)
out_book.create_sheet(index=0, title=slf_sheet)
out_book.create_sheet(index=1, title=obs_sheet)
out_book.create_sheet(index=2, title=test_sheet)
out_book.save(out_xlsx)
STANDBY_TEXT = visual.TextStim(win=win, name='STANDBY_TEXT',
    text='Press ‘space’ to start',
    font='Arial',
    pos=(0, 0), height=48, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
STANDBY_RESP = keyboard.Keyboard()


"""=================================================================================================================="""
# --- Initialize components for Routine "SlfInstr" ---
SLF_INSTR = visual.TextStim(win=win, name='SLF_INSTR',
    text='Self',
    font='Arial',
    pos=(0, 0), height=48, wrapWidth=None, ori=0, 
    color=slf_color, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);


"""=================================================================================================================="""
# --- Initialize components for Routine "SlfChoice" ---
SLF_FIX1 = visual.TextStim(win=win, name='SLF_FIX1',
    text='+',
    font='Arial',
    pos=(0, 0), height=60, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
SLF_IMG1 = visual.ImageStim(
    win=win,
    name='SLF_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=[0,0], size=(200, 200),
    color='white', colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-2.0)
SLF_IMG2 = visual.ImageStim(
    win=win,
    name='SLF_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-3.0)
SLF_CHOICE_RESP = keyboard.Keyboard()


"""=================================================================================================================="""
# --- Initialize components for Routine "SlfChoiceFb" ---
SLF_FIX2 = visual.TextStim(win=win, name='SLF_FIX2',
    text='+',
    font='Arial',
    pos=(0, 0), height=60.0, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
SLF_FB1_FRAME = visual.Rect(
    win=win, name='SLF_FB1_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0, pos=[0,0], anchor='center',
    lineWidth=20,     colorSpace='rgb',  lineColor=slf_color, fillColor=slf_color,
    opacity=1, depth=-1.0, interpolate=True)
SLF_FB1_IMG1 = visual.ImageStim(
    win=win,
    name='SLF_FB1_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(-200, 0), size=(200,200),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-2.0)
SLF_FB1_IMG2 = visual.ImageStim(
    win=win,
    name='SLF_FB1_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(200, 0), size=(200, 200),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-3.0)


"""=================================================================================================================="""
# --- Initialize components for Routine "SlfConf" ---
SLF_LOW_BOX = visual.Rect(
    win=win, name='SLF_LOW_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(-250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-1.0, interpolate=True)
SLF_LOW_TEXT = visual.TextStim(win=win, name='SLF_LOW_TEXT',
    text='1',
    font='Arial',
    pos=(-250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-2.0);
SLF_MID_BOX = visual.Rect(
    win=win, name='SLF_MID_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(0, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-3.0, interpolate=True)
SLF_MID_TEXT = visual.TextStim(win=win, name='SLF_MID_TEXT',
    text='2',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);
SLF_HIGH_BOX = visual.Rect(
    win=win, name='SLF_HIGH_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-5.0, interpolate=True)
SLF_HIGH_TEXT = visual.TextStim(win=win, name='SLF_HIGH_TEXT',
    text='3',
    font='Arial',
    pos=(250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-6.0);
SLF_CONF_RESP = keyboard.Keyboard()


"""=================================================================================================================="""
# --- Initialize components for Routine "SlfConfFb" ---
SLF_FB2_FRAME = visual.Rect(
    win=win, name='SLF_FB2_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0.0, pos=[0,0], anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor=slf_color, fillColor=slf_color,
    opacity=1.0, depth=0.0, interpolate=True)
SLF_FB2_LOW_BOX = visual.Rect(
    win=win, name='SLF_FB2_LOW_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(-250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-1.0, interpolate=True)
SLF_FB2_LOW_TEXT = visual.TextStim(win=win, name='SLF_FB2_LOW_TEXT',
    text='1',
    font='Arial',
    pos=(-250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-2.0);
SLF_FB2_MID_BOX = visual.Rect(
    win=win, name='SLF_FB2_MID_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(0, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-3.0, interpolate=True)
SLF_FB2_MID_TEXT = visual.TextStim(win=win, name='SLF_FB2_MID_TEXT',
    text='2',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);
SLF_FB2_HIGH_BOX = visual.Rect(
    win=win, name='SLF_FB2_HIGH_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-5.0, interpolate=True)
SLF_FB2_HIGH_TEXT = visual.TextStim(win=win, name='SLF_FB2_HIGH_TEXT',
    text='3',
    font='Arial',
    pos=(250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-6.0);


"""=================================================================================================================="""
# --- Initialize components for Routine "SlfReward" ---
SLF_FB3_FRAME = visual.Rect(
    win=win, name='SLF_FB3_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0.0, pos=[0,0], anchor='center',
    lineWidth=20.0,     colorSpace='rgb',  lineColor=slf_color, fillColor=slf_color,
    opacity=1.0, depth=-1.0, interpolate=True)
SLF_FB3_IMG1 = visual.ImageStim(
    win=win,
    name='SLF_FB3_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0.0, pos=(-200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False,
    texRes=128.0, interpolate=True, depth=-2.0)
SLF_FB3_IMG2 = visual.ImageStim(
    win=win,
    name='SLF_FB3_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0.0, pos=(200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False,
    texRes=128.0, interpolate=True, depth=-3.0)
SLF_REWARD = visual.TextStim(win=win, name='SLF_REWARD',
    text='',
    font='Arial',
    units='pix', pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);


"""=================================================================================================================="""
# --- Initialize components for Routine "ObsInstr" ---
OBS_INSTR = visual.TextStim(win=win, name='OBS_INSTR',
    text='Observation',
    font='Arial',
    pos=(0, 0), height=48, wrapWidth=None, ori=0, 
    color=obs_color, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);


"""=================================================================================================================="""
# --- Initialize components for Routine "ObsChoice" ---
OBS_FIX = visual.TextStim(win=win, name='OBS_FIX',
    text='+',
    font='Arial',
    pos=(0, 0), height=60, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
OBS_FB1_FRAME = visual.Rect(
    win=win, name='OBS_FB1_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0, pos=[0,0], anchor='center',
    lineWidth=20,     colorSpace='rgb',  lineColor=obs_color, fillColor=obs_color,
    opacity=1, depth=-2.0, interpolate=True)
OBS_IMG1 = visual.ImageStim(
    win=win,
    name='OBS_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=[0,0], size=(200, 200),
    color='white', colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-3.0)
OBS_IMG2 = visual.ImageStim(
    win=win,
    name='OBS_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-4.0)


"""=================================================================================================================="""
# --- Initialize components for Routine "ObsConf" ---
OBS_FB2_FRAME = visual.Rect(
    win=win, name='OBS_FB2_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0.0, pos=[0,0], anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor=obs_color, fillColor=obs_color,
    opacity=1.0, depth=0.0, interpolate=True)
OBS_LOW_BOX = visual.Rect(
    win=win, name='OBS_LOW_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(-250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-1.0, interpolate=True)
OBS_LOW_TEXT = visual.TextStim(win=win, name='OBS_LOW_TEXT',
    text='1',
    font='Arial',
    pos=(-250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-2.0);
OBS_MID_BOX = visual.Rect(
    win=win, name='OBS_MID_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(0, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-3.0, interpolate=True)
OBS_MID_TEXT = visual.TextStim(win=win, name='OBS_MID_TEXT',
    text='2',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);
OBS_HIGH_BOX = visual.Rect(
    win=win, name='OBS_HIGH_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-5.0, interpolate=True)
OBS_HIGH_TEXT = visual.TextStim(win=win, name='OBS_HIGH_TEXT',
    text='3',
    font='Arial',
    pos=(250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-6.0);


"""=================================================================================================================="""
# --- Initialize components for Routine "ObsReward" ---
OBS_FB3_FRAME = visual.Rect(
    win=win, name='OBS_FB3_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0.0, pos=[0,0], anchor='center',
    lineWidth=20.0,     colorSpace='rgb',  lineColor=obs_color, fillColor=obs_color,
    opacity=1.0, depth=-1.0, interpolate=True)
OBS_FB3_IMG1 = visual.ImageStim(
    win=win,
    name='OBS_FB3_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0.0, pos=(-200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False,
    texRes=128.0, interpolate=True, depth=-2.0)
OBS_FB3_IMG2 = visual.ImageStim(
    win=win,
    name='OBS_FB3_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0.0, pos=(200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False,
    texRes=128.0, interpolate=True, depth=-3.0)
OBS_REWARD = visual.TextStim(win=win, name='OBS_REWARD',
    text='',
    font='Arial',
    units='pix', pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);

# --- Initialize components for Routine "TestInstr" ---
TEST_INSTR = visual.TextStim(win=win, name='TEST_INSTR',
    text='Test',
    font='Arial',
    pos=(0, 0), height=48, wrapWidth=None, ori=0, 
    color=test_color, colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=0.0);

# --- Initialize components for Routine "TestChoice" ---
TEST_FIX1 = visual.TextStim(win=win, name='TEST_FIX1',
    text='+',
    font='Arial',
    pos=(0, 0), height=60, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
TEST_IMG1 = visual.ImageStim(
    win=win,
    name='TEST_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(-200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-2.0)
TEST_IMG2 = visual.ImageStim(
    win=win,
    name='TEST_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-3.0)
TEST_CHOICE_RESP = keyboard.Keyboard()


"""=================================================================================================================="""
# --- Initialize components for Routine "TestChoiceFb" ---
TEST_FIX2 = visual.TextStim(win=win, name='TEST_FIX2',
    text='+',
    font='Arial',
    pos=(0, 0), height=60.0, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=None, 
    languageStyle='LTR',
    depth=0.0);
TEST_FB1_FRAME = visual.Rect(
    win=win, name='TEST_FB1_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0, pos=[0,0], anchor='center',
    lineWidth=20,     colorSpace='rgb',  lineColor=test_color, fillColor=test_color,
    opacity=1, depth=-1.0, interpolate=True)
TEST_FB1_IMG1 = visual.ImageStim(
    win=win,
    name='TEST_FB1_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(-200, 0), size=(200,200),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-2.0)
TEST_FB1_IMG2 = visual.ImageStim(
    win=win,
    name='TEST_FB1_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0, pos=(200, 0), size=(200, 200),
    color=[1,1,1], colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False,
    texRes=128, interpolate=True, depth=-3.0)


"""=================================================================================================================="""
# --- Initialize components for Routine "TestConf" ---
TEST_LOW_BOX = visual.Rect(
    win=win, name='TEST_LOW_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(-250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-1.0, interpolate=True)
TEST_LOW_TEXT = visual.TextStim(win=win, name='TEST_LOW_TEXT',
    text='1',
    font='Arial',
    pos=(-250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-2.0);
SLF_TEST_BOX = visual.Rect(
    win=win, name='SLF_TEST_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(0, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-3.0, interpolate=True)
TEST_MID_TEXT = visual.TextStim(win=win, name='TEST_MID_TEXT',
    text='2',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);
TEST_HIGH_BOX = visual.Rect(
    win=win, name='TEST_HIGH_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-5.0, interpolate=True)
TEST_HIGH_TEXT = visual.TextStim(win=win, name='TEST_HIGH_TEXT',
    text='3',
    font='Arial',
    pos=(250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-6.0);
TEST_CONF_RESP = keyboard.Keyboard()


"""=================================================================================================================="""
# --- Initialize components for Routine "TestConfFb" ---
TEST_FB2_FRAME = visual.Rect(
    win=win, name='TEST_FB2_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0.0, pos=[0,0], anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor=test_color, fillColor=test_color,
    opacity=1.0, depth=0.0, interpolate=True)
TEST_FB2_LOW_BOX = visual.Rect(
    win=win, name='TEST_FB2_LOW_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(-250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-1.0, interpolate=True)
TEST_FB2_LOW_TEXT = visual.TextStim(win=win, name='TEST_FB2_LOW_TEXT',
    text='1',
    font='Arial',
    pos=(-250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-2.0);
TEST_FB2_MID_BOX = visual.Rect(
    win=win, name='TEST_FB2_MID_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(0, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-3.0, interpolate=True)
TEST_FB2_MID_TEXT = visual.TextStim(win=win, name='TEST_FB2_MID_TEXT',
    text='2',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);
TEST_FB2_HIGH_BOX = visual.Rect(
    win=win, name='TEST_FB2_HIGH_BOX',
    width=(200, 200)[0], height=(200, 200)[1],
    ori=0.0, pos=(250, 0), anchor='center',
    lineWidth=1.0,     colorSpace='rgb',  lineColor='white', fillColor='white',
    opacity=1.0, depth=-5.0, interpolate=True)
TEST_FB2_HIGH_TEXT = visual.TextStim(win=win, name='TEST_FB2_HIGH_TEXT',
    text='3',
    font='Arial',
    pos=(250, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='black', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-6.0);


"""=================================================================================================================="""
# --- Initialize components for Routine "TestReward" ---
TEST_FB3_FRAME = visual.Rect(
    win=win, name='TEST_FB3_FRAME',
    width=(220, 220)[0], height=(220, 220)[1],
    ori=0.0, pos=[0,0], anchor='center',
    lineWidth=20.0,     colorSpace='rgb',  lineColor=test_color, fillColor=test_color,
    opacity=1.0, depth=-1.0, interpolate=True)
TEST_FB3_IMG1 = visual.ImageStim(
    win=win,
    name='TEST_FB3_IMG1', 
    image='sin', mask=None, anchor='center',
    ori=0.0, pos=(-200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False,
    texRes=128.0, interpolate=True, depth=-2.0)
TEST_FB3_IMG2 = visual.ImageStim(
    win=win,
    name='TEST_FB3_IMG2', 
    image='sin', mask=None, anchor='center',
    ori=0.0, pos=(200, 0), size=(200, 200),
    color='white', colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False,
    texRes=128.0, interpolate=True, depth=-3.0)
TEST_REWARD = visual.TextStim(win=win, name='TEST_REWARD',
    text='',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-4.0);


"""=================================================================================================================="""
# --- Initialize components for Routine "BlockRest" ---
BLOCK_REST_TEXT = visual.TextStim(win=win, name='BLOCK_REST_TEXT',
    text='',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=-1.0);
BLOCK_REST_SKIP = keyboard.Keyboard()


"""=================================================================================================================="""
# --- Initialize components for Routine "Appreciation" ---
APPRECIATION = visual.TextStim(win=win, name='APPRECIATION',
    text='Thank you for your cooperation in the experiment!',
    font='Arial',
    pos=(0, 0), height=48.0, wrapWidth=None, ori=0.0, 
    color='white', colorSpace='rgb', opacity=1.0, 
    languageStyle='LTR',
    depth=0.0);


"""=================================================================================================================="""
# Create some handy timers
globalClock = core.Clock()  # to track the time since experiment started
routineTimer = core.Clock()  # to track time remaining of each (possibly non-slip) routine 

# --- Prepare to start Routine "Init" ---
continueRoutine = True
routineForceEnded = False
# update component parameters for each repeat
STANDBY_RESP.keys = []
STANDBY_RESP.rt = []
_STANDBY_RESP_allKeys = []
# keep track of which components have finished
InitComponents = [STANDBY_TEXT, STANDBY_RESP]
for thisComponent in InitComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
frameN = -1

# --- Run Routine "Init" ---
while continueRoutine:
    # get current time
    t = routineTimer.getTime()
    tThisFlip = win.getFutureFlipTime(clock=routineTimer)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *STANDBY_TEXT* updates
    if STANDBY_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        STANDBY_TEXT.frameNStart = frameN  # exact frame index
        STANDBY_TEXT.tStart = t  # local t and not account for scr refresh
        STANDBY_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(STANDBY_TEXT, 'tStartRefresh')  # time at next scr refresh
        # add timestamp to datafile
        thisExp.timestampOnFlip(win, 'STANDBY_TEXT.started')
        STANDBY_TEXT.setAutoDraw(True)
    
    # *STANDBY_RESP* updates
    waitOnFlip = False
    if STANDBY_RESP.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        STANDBY_RESP.frameNStart = frameN  # exact frame index
        STANDBY_RESP.tStart = t  # local t and not account for scr refresh
        STANDBY_RESP.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(STANDBY_RESP, 'tStartRefresh')  # time at next scr refresh
        # add timestamp to datafile
        thisExp.timestampOnFlip(win, 'STANDBY_RESP.started')
        STANDBY_RESP.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(STANDBY_RESP.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(STANDBY_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if STANDBY_RESP.status == STARTED and not waitOnFlip:
        theseKeys = STANDBY_RESP.getKeys(keyList=['space'], waitRelease=False)
        _STANDBY_RESP_allKeys.extend(theseKeys)
        if len(_STANDBY_RESP_allKeys):
            STANDBY_RESP.keys = _STANDBY_RESP_allKeys[-1].name  # just the last key pressed
            STANDBY_RESP.rt = _STANDBY_RESP_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        routineForceEnded = True
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in InitComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# --- Ending Routine "Init" ---
for thisComponent in InitComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
# Run 'End Routine' code from INIT_CODE
# Initialize clock
onset = core.MonotonicClock()
core.wait(3)
# check responses
if STANDBY_RESP.keys in ['', [], None]:  # No response was made
    STANDBY_RESP.keys = None
thisExp.addData('STANDBY_RESP.keys',STANDBY_RESP.keys)
if STANDBY_RESP.keys != None:  # we had a response
    thisExp.addData('STANDBY_RESP.rt', STANDBY_RESP.rt)
thisExp.nextEntry()
# the Routine "Init" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
BlockLoop = data.TrialHandler(nReps=1.0, method='sequential', 
    extraInfo=expInfo, originPath=-1,
    trialList=data.importConditions('sequence/SequenceConditionPattern1.xlsx'),
    seed=None, name='BlockLoop')
thisExp.addLoop(BlockLoop)  # add the loop to the experiment
thisBlockLoop = BlockLoop.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisBlockLoop.rgb)
if thisBlockLoop != None:
    for paramName in thisBlockLoop:
        exec('{} = thisBlockLoop[paramName]'.format(paramName))

for thisBlockLoop in BlockLoop:
    currentLoop = BlockLoop
    # abbreviate parameter names if possible (e.g. rgb = thisBlockLoop.rgb)
    if thisBlockLoop != None:
        for paramName in thisBlockLoop:
            exec('{} = thisBlockLoop[paramName]'.format(paramName))
    
    # --- Prepare to start Routine "SlfInstr" ---
    continueRoutine = True
    routineForceEnded = False
    # update component parameters for each repeat
    # keep track of which components have finished
    SlfInstrComponents = [SLF_INSTR]
    for thisComponent in SlfInstrComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    """=================================================================================================================="""
    # --- Run Routine "SlfInstr" ---
    while continueRoutine and routineTimer.getTime() < 2.0:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *SLF_INSTR* updates
        if SLF_INSTR.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            SLF_INSTR.frameNStart = frameN  # exact frame index
            SLF_INSTR.tStart = t  # local t and not account for scr refresh
            SLF_INSTR.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(SLF_INSTR, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'SLF_INSTR.started')
            SLF_INSTR.setAutoDraw(True)
        if SLF_INSTR.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > SLF_INSTR.tStartRefresh + 2-frameTolerance:
                # keep track of stop time/frame for later
                SLF_INSTR.tStop = t  # not accounting for scr refresh
                SLF_INSTR.frameNStop = frameN  # exact frame index
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_INSTR.stopped')
                SLF_INSTR.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in SlfInstrComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "SlfInstr" ---
    for thisComponent in SlfInstrComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
    if routineForceEnded:
        routineTimer.reset()
    else:
        routineTimer.addTime(-2.000000)
    
    # set up handler to look after randomisation of conditions etc
    SlfSequence = data.TrialHandler(nReps=1.0, method='sequential', 
        extraInfo=expInfo, originPath=-1,
        trialList=data.importConditions('sequence/' + slf_file),
        seed=None, name='SlfSequence')
    thisExp.addLoop(SlfSequence)  # add the loop to the experiment
    thisSlfSequence = SlfSequence.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb = thisSlfSequence.rgb)
    if thisSlfSequence != None:
        for paramName in thisSlfSequence:
            exec('{} = thisSlfSequence[paramName]'.format(paramName))
    
    """=================================================================================================================="""
    for thisSlfSequence in SlfSequence:
        currentLoop = SlfSequence
        # abbreviate parameter names if possible (e.g. rgb = thisSlfSequence.rgb)
        if thisSlfSequence != None:
            for paramName in thisSlfSequence:
                exec('{} = thisSlfSequence[paramName]'.format(paramName))
        
        # --- Prepare to start Routine "SlfChoice" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        # Run 'Begin Routine' code from SLF_CHOICE_CODE
        slf_img1 = slf_dict[slf_loc1]
        slf_img2 = slf_dict[slf_loc2]
        
        # Get the presented time in slf-choice
        slf_choice_pres_time = onset.getTime() + 0.5 # Consider the fixation
        SLF_IMG1.setPos((-200, 0))
        SLF_IMG1.setImage('img/' + slf_img1)
        SLF_IMG2.setImage('img/' + slf_img2)
        SLF_CHOICE_RESP.keys = []
        SLF_CHOICE_RESP.rt = []
        _SLF_CHOICE_RESP_allKeys = []
        # keep track of which components have finished
        SlfChoiceComponents = [SLF_FIX1, SLF_IMG1, SLF_IMG2, SLF_CHOICE_RESP]
        for thisComponent in SlfChoiceComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "SlfChoice" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *SLF_FIX1* updates
            if SLF_FIX1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FIX1.frameNStart = frameN  # exact frame index
                SLF_FIX1.tStart = t  # local t and not account for scr refresh
                SLF_FIX1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FIX1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FIX1.started')
                SLF_FIX1.setAutoDraw(True)
            if SLF_FIX1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FIX1.tStartRefresh + choice_time_max+0.5-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FIX1.tStop = t  # not accounting for scr refresh
                    SLF_FIX1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FIX1.stopped')
                    SLF_FIX1.setAutoDraw(False)
            
            # *SLF_IMG1* updates
            if SLF_IMG1.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMG1.frameNStart = frameN  # exact frame index
                SLF_IMG1.tStart = t  # local t and not account for scr refresh
                SLF_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMG1.started')
                SLF_IMG1.setAutoDraw(True)
            if SLF_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMG1.tStartRefresh + choice_time_max-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMG1.tStop = t  # not accounting for scr refresh
                    SLF_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMG1.stopped')
                    SLF_IMG1.setAutoDraw(False)
            
            # *SLF_IMG2* updates
            if SLF_IMG2.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMG2.frameNStart = frameN  # exact frame index
                SLF_IMG2.tStart = t  # local t and not account for scr refresh
                SLF_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMG2.started')
                SLF_IMG2.setAutoDraw(True)
            if SLF_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMG2.tStartRefresh + choice_time_max-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMG2.tStop = t  # not accounting for scr refresh
                    SLF_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMG2.stopped')
                    SLF_IMG2.setAutoDraw(False)
            
            # *SLF_CHOICE_RESP* updates
            waitOnFlip = False
            if SLF_CHOICE_RESP.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                SLF_CHOICE_RESP.frameNStart = frameN  # exact frame index
                SLF_CHOICE_RESP.tStart = t  # local t and not account for scr refresh
                SLF_CHOICE_RESP.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_CHOICE_RESP, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_CHOICE_RESP.started')
                SLF_CHOICE_RESP.status = STARTED
                # keyboard checking is just starting
                waitOnFlip = True
                win.callOnFlip(SLF_CHOICE_RESP.clock.reset)  # t=0 on next screen flip
                win.callOnFlip(SLF_CHOICE_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
            if SLF_CHOICE_RESP.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_CHOICE_RESP.tStartRefresh + choice_time_max-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_CHOICE_RESP.tStop = t  # not accounting for scr refresh
                    SLF_CHOICE_RESP.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CHOICE_RESP.stopped')
                    SLF_CHOICE_RESP.status = FINISHED
            if SLF_CHOICE_RESP.status == STARTED and not waitOnFlip:
                theseKeys = SLF_CHOICE_RESP.getKeys(keyList=[left_key,right_key], waitRelease=False)
                _SLF_CHOICE_RESP_allKeys.extend(theseKeys)
                if len(_SLF_CHOICE_RESP_allKeys):
                    SLF_CHOICE_RESP.keys = _SLF_CHOICE_RESP_allKeys[-1].name  # just the last key pressed
                    SLF_CHOICE_RESP.rt = _SLF_CHOICE_RESP_allKeys[-1].rt
                    # was this correct?
                    if (SLF_CHOICE_RESP.keys == str('slfCorrect')) or (SLF_CHOICE_RESP.keys == 'slfCorrect'):
                        SLF_CHOICE_RESP.corr = 1
                    else:
                        SLF_CHOICE_RESP.corr = 0
                    # a response ends the routine
                    continueRoutine = False
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in SlfChoiceComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "SlfChoice" ---
        for thisComponent in SlfChoiceComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # Run 'End Routine' code from SLF_CHOICE_CODE
        # Get the responsed time in slf-choice
        slf_choice_resp_time = onset.getTime()
        
        # Initialize in case skipping slf-conf
        slf_conf_resp = 0
        slf_conf_pres_time = 0
        slf_conf_resp_time = 0
        is_slf_resp = 0
        
        # Receive choice responses and determine reward and feedback parameters
        if SLF_CHOICE_RESP.keys == left_key: 
            slf_choice_pos = (-200,0)
            slf_choice_resp = 1
            slf_choice_loc = slf_loc1
            slf_reward = slf_reward1
            slf_reward_text = str(np.round(slf_reward, 1))
            is_slf_resp = 1
        elif SLF_CHOICE_RESP.keys == right_key: 
            slf_choice_pos = (200,0)
            slf_choice_resp = 2
            slf_choice_loc = slf_loc2
            slf_reward = slf_reward2
            slf_reward_text = str(np.round(slf_reward, 1))
            is_slf_resp = 1
        else:
            slf_choice_pos = (100000,0)
            slf_choice_resp = 0
            slf_choice_loc = 0
            slf_reward = 0
            slf_reward_text = "F"
        
        # Add points according to the response
        slf_pt = 1 if slf_choice_resp == slf_correct else 0
        # check responses
        if SLF_CHOICE_RESP.keys in ['', [], None]:  # No response was made
            SLF_CHOICE_RESP.keys = None
            # was no response the correct answer?!
            if str('slfCorrect').lower() == 'none':
               SLF_CHOICE_RESP.corr = 1;  # correct non-response
            else:
               SLF_CHOICE_RESP.corr = 0;  # failed to respond (incorrectly)
        # store data for SlfSequence (TrialHandler)
        SlfSequence.addData('SLF_CHOICE_RESP.keys',SLF_CHOICE_RESP.keys)
        SlfSequence.addData('SLF_CHOICE_RESP.corr', SLF_CHOICE_RESP.corr)
        if SLF_CHOICE_RESP.keys != None:  # we had a response
            SlfSequence.addData('SLF_CHOICE_RESP.rt', SLF_CHOICE_RESP.rt)
        # the Routine "SlfChoice" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # --- Prepare to start Routine "SlfChoiceFb" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        SLF_FB1_FRAME.setPos(slf_choice_pos)
        SLF_FB1_IMG1.setImage('img/' + slf_img1)
        SLF_FB1_IMG2.setImage('img/' + slf_img2)
        # keep track of which components have finished
        SlfChoiceFbComponents = [SLF_FIX2, SLF_FB1_FRAME, SLF_FB1_IMG1, SLF_FB1_IMG2]
        for thisComponent in SlfChoiceFbComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        """=================================================================================================================="""
        # --- Run Routine "SlfChoiceFb" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *SLF_FIX2* updates
            if SLF_FIX2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FIX2.frameNStart = frameN  # exact frame index
                SLF_FIX2.tStart = t  # local t and not account for scr refresh
                SLF_FIX2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FIX2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FIX2.started')
                SLF_FIX2.setAutoDraw(True)
            if SLF_FIX2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FIX2.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FIX2.tStop = t  # not accounting for scr refresh
                    SLF_FIX2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FIX2.stopped')
                    SLF_FIX2.setAutoDraw(False)
            
            # *SLF_FB1_FRAME* updates
            if SLF_FB1_FRAME.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FB1_FRAME.frameNStart = frameN  # exact frame index
                SLF_FB1_FRAME.tStart = t  # local t and not account for scr refresh
                SLF_FB1_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FB1_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FB1_FRAME.started')
                SLF_FB1_FRAME.setAutoDraw(True)
            if SLF_FB1_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FB1_FRAME.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FB1_FRAME.tStop = t  # not accounting for scr refresh
                    SLF_FB1_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB1_FRAME.stopped')
                    SLF_FB1_FRAME.setAutoDraw(False)
            
            # *SLF_FB1_IMG1* updates
            if SLF_FB1_IMG1.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FB1_IMG1.frameNStart = frameN  # exact frame index
                SLF_FB1_IMG1.tStart = t  # local t and not account for scr refresh
                SLF_FB1_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FB1_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FB1_IMG1.started')
                SLF_FB1_IMG1.setAutoDraw(True)
            if SLF_FB1_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FB1_IMG1.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FB1_IMG1.tStop = t  # not accounting for scr refresh
                    SLF_FB1_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB1_IMG1.stopped')
                    SLF_FB1_IMG1.setAutoDraw(False)
            
            # *SLF_FB1_IMG2* updates
            if SLF_FB1_IMG2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FB1_IMG2.frameNStart = frameN  # exact frame index
                SLF_FB1_IMG2.tStart = t  # local t and not account for scr refresh
                SLF_FB1_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FB1_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FB1_IMG2.started')
                SLF_FB1_IMG2.setAutoDraw(True)
            if SLF_FB1_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FB1_IMG2.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FB1_IMG2.tStop = t  # not accounting for scr refresh
                    SLF_FB1_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB1_IMG2.stopped')
                    SLF_FB1_IMG2.setAutoDraw(False)
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in SlfChoiceFbComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "SlfChoiceFb" ---
        for thisComponent in SlfChoiceFbComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # the Routine "SlfChoiceFb" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # set up handler to look after randomisation of conditions etc
        SlfSkip = data.TrialHandler(nReps=is_slf_resp, method='random', 
            extraInfo=expInfo, originPath=-1,
            trialList=[None],
            seed=None, name='SlfSkip')
        thisExp.addLoop(SlfSkip)  # add the loop to the experiment
        thisSlfSkip = SlfSkip.trialList[0]  # so we can initialise stimuli with some values
        # abbreviate parameter names if possible (e.g. rgb = thisSlfSkip.rgb)
        if thisSlfSkip != None:
            for paramName in thisSlfSkip:
                exec('{} = thisSlfSkip[paramName]'.format(paramName))
        
        for thisSlfSkip in SlfSkip:
            currentLoop = SlfSkip
            # abbreviate parameter names if possible (e.g. rgb = thisSlfSkip.rgb)
            if thisSlfSkip != None:
                for paramName in thisSlfSkip:
                    exec('{} = thisSlfSkip[paramName]'.format(paramName))
            
            # --- Prepare to start Routine "SlfConf" ---
            continueRoutine = True
            routineForceEnded = False
            # update component parameters for each repeat
            # Run 'Begin Routine' code from SLF_CONF_CODE
            # Get the presented time in slf-confidence
            slf_conf_pres_time = onset.getTime()
            SLF_CONF_RESP.keys = []
            SLF_CONF_RESP.rt = []
            _SLF_CONF_RESP_allKeys = []
            # keep track of which components have finished
            SlfConfComponents = [SLF_LOW_BOX, SLF_LOW_TEXT, SLF_MID_BOX, SLF_MID_TEXT, SLF_HIGH_BOX, SLF_HIGH_TEXT, SLF_CONF_RESP]
            for thisComponent in SlfConfComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "SlfConf" ---
            while continueRoutine:
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *SLF_LOW_BOX* updates
                if SLF_LOW_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_LOW_BOX.frameNStart = frameN  # exact frame index
                    SLF_LOW_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_LOW_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_LOW_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_LOW_BOX.started')
                    SLF_LOW_BOX.setAutoDraw(True)
                if SLF_LOW_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_LOW_BOX.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_LOW_BOX.tStop = t  # not accounting for scr refresh
                        SLF_LOW_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_LOW_BOX.stopped')
                        SLF_LOW_BOX.setAutoDraw(False)
                
                # *SLF_LOW_TEXT* updates
                if SLF_LOW_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_LOW_TEXT.frameNStart = frameN  # exact frame index
                    SLF_LOW_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_LOW_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_LOW_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_LOW_TEXT.started')
                    SLF_LOW_TEXT.setAutoDraw(True)
                if SLF_LOW_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_LOW_TEXT.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_LOW_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_LOW_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_LOW_TEXT.stopped')
                        SLF_LOW_TEXT.setAutoDraw(False)
                
                # *SLF_MID_BOX* updates
                if SLF_MID_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_MID_BOX.frameNStart = frameN  # exact frame index
                    SLF_MID_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_MID_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_MID_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_MID_BOX.started')
                    SLF_MID_BOX.setAutoDraw(True)
                if SLF_MID_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_MID_BOX.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_MID_BOX.tStop = t  # not accounting for scr refresh
                        SLF_MID_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_MID_BOX.stopped')
                        SLF_MID_BOX.setAutoDraw(False)
                
                # *SLF_MID_TEXT* updates
                if SLF_MID_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_MID_TEXT.frameNStart = frameN  # exact frame index
                    SLF_MID_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_MID_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_MID_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_MID_TEXT.started')
                    SLF_MID_TEXT.setAutoDraw(True)
                if SLF_MID_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_MID_TEXT.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_MID_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_MID_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_MID_TEXT.stopped')
                        SLF_MID_TEXT.setAutoDraw(False)
                
                # *SLF_HIGH_BOX* updates
                if SLF_HIGH_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_HIGH_BOX.frameNStart = frameN  # exact frame index
                    SLF_HIGH_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_HIGH_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_HIGH_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_HIGH_BOX.started')
                    SLF_HIGH_BOX.setAutoDraw(True)
                if SLF_HIGH_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_HIGH_BOX.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_HIGH_BOX.tStop = t  # not accounting for scr refresh
                        SLF_HIGH_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_HIGH_BOX.stopped')
                        SLF_HIGH_BOX.setAutoDraw(False)
                
                # *SLF_HIGH_TEXT* updates
                if SLF_HIGH_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_HIGH_TEXT.frameNStart = frameN  # exact frame index
                    SLF_HIGH_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_HIGH_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_HIGH_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_HIGH_TEXT.started')
                    SLF_HIGH_TEXT.setAutoDraw(True)
                if SLF_HIGH_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_HIGH_TEXT.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_HIGH_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_HIGH_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_HIGH_TEXT.stopped')
                        SLF_HIGH_TEXT.setAutoDraw(False)
                
                # *SLF_CONF_RESP* updates
                waitOnFlip = False
                if SLF_CONF_RESP.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONF_RESP.frameNStart = frameN  # exact frame index
                    SLF_CONF_RESP.tStart = t  # local t and not account for scr refresh
                    SLF_CONF_RESP.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONF_RESP, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONF_RESP.started')
                    SLF_CONF_RESP.status = STARTED
                    # keyboard checking is just starting
                    waitOnFlip = True
                    win.callOnFlip(SLF_CONF_RESP.clock.reset)  # t=0 on next screen flip
                    win.callOnFlip(SLF_CONF_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
                if SLF_CONF_RESP.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONF_RESP.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONF_RESP.tStop = t  # not accounting for scr refresh
                        SLF_CONF_RESP.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONF_RESP.stopped')
                        SLF_CONF_RESP.status = FINISHED
                if SLF_CONF_RESP.status == STARTED and not waitOnFlip:
                    theseKeys = SLF_CONF_RESP.getKeys(keyList=[left_key,center_key,right_key], waitRelease=False)
                    _SLF_CONF_RESP_allKeys.extend(theseKeys)
                    if len(_SLF_CONF_RESP_allKeys):
                        SLF_CONF_RESP.keys = _SLF_CONF_RESP_allKeys[-1].name  # just the last key pressed
                        SLF_CONF_RESP.rt = _SLF_CONF_RESP_allKeys[-1].rt
                        # a response ends the routine
                        continueRoutine = False
                
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    routineForceEnded = True
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in SlfConfComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "SlfConf" ---
            for thisComponent in SlfConfComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # Run 'End Routine' code from SLF_CONF_CODE
            # Get the responsed time in slf-confidence
            slf_conf_resp_time = onset.getTime()
            
            # Receive confidence responses and determine feedback parameters
            if SLF_CONF_RESP.keys == left_key:
                slf_conf_pos = (-250,0)
                slf_conf_resp = 7
            elif SLF_CONF_RESP.keys == center_key: 
                slf_conf_pos = (0,0)
                slf_conf_resp = 8
            elif SLF_CONF_RESP.keys == right_key: 
                slf_conf_pos = (250,0)
                slf_conf_resp = 9
            else:
                slf_conf_pos = (100000,0)
                slf_conf_resp = 0
            # check responses
            if SLF_CONF_RESP.keys in ['', [], None]:  # No response was made
                SLF_CONF_RESP.keys = None
            SlfSkip.addData('SLF_CONF_RESP.keys',SLF_CONF_RESP.keys)
            if SLF_CONF_RESP.keys != None:  # we had a response
                SlfSkip.addData('SLF_CONF_RESP.rt', SLF_CONF_RESP.rt)
            # the Routine "SlfConf" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            
            # --- Prepare to start Routine "SlfConfFb" ---
            continueRoutine = True
            routineForceEnded = False
            # update component parameters for each repeat
            SLF_FB2_FRAME.setPos(slf_conf_pos)
            # keep track of which components have finished
            SlfConfFbComponents = [SLF_FB2_FRAME, SLF_FB2_LOW_BOX, SLF_FB2_LOW_TEXT, SLF_FB2_MID_BOX, SLF_FB2_MID_TEXT, SLF_FB2_HIGH_BOX, SLF_FB2_HIGH_TEXT]
            for thisComponent in SlfConfFbComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "SlfConfFb" ---
            while continueRoutine:
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *SLF_FB2_FRAME* updates
                if SLF_FB2_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_FB2_FRAME.frameNStart = frameN  # exact frame index
                    SLF_FB2_FRAME.tStart = t  # local t and not account for scr refresh
                    SLF_FB2_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_FB2_FRAME, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB2_FRAME.started')
                    SLF_FB2_FRAME.setAutoDraw(True)
                if SLF_FB2_FRAME.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_FB2_FRAME.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_FB2_FRAME.tStop = t  # not accounting for scr refresh
                        SLF_FB2_FRAME.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_FB2_FRAME.stopped')
                        SLF_FB2_FRAME.setAutoDraw(False)
                
                # *SLF_FB2_LOW_BOX* updates
                if SLF_FB2_LOW_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_FB2_LOW_BOX.frameNStart = frameN  # exact frame index
                    SLF_FB2_LOW_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_FB2_LOW_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_FB2_LOW_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB2_LOW_BOX.started')
                    SLF_FB2_LOW_BOX.setAutoDraw(True)
                if SLF_FB2_LOW_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_FB2_LOW_BOX.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_FB2_LOW_BOX.tStop = t  # not accounting for scr refresh
                        SLF_FB2_LOW_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_FB2_LOW_BOX.stopped')
                        SLF_FB2_LOW_BOX.setAutoDraw(False)
                
                # *SLF_FB2_LOW_TEXT* updates
                if SLF_FB2_LOW_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_FB2_LOW_TEXT.frameNStart = frameN  # exact frame index
                    SLF_FB2_LOW_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_FB2_LOW_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_FB2_LOW_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB2_LOW_TEXT.started')
                    SLF_FB2_LOW_TEXT.setAutoDraw(True)
                if SLF_FB2_LOW_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_FB2_LOW_TEXT.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_FB2_LOW_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_FB2_LOW_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_FB2_LOW_TEXT.stopped')
                        SLF_FB2_LOW_TEXT.setAutoDraw(False)
                
                # *SLF_FB2_MID_BOX* updates
                if SLF_FB2_MID_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_FB2_MID_BOX.frameNStart = frameN  # exact frame index
                    SLF_FB2_MID_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_FB2_MID_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_FB2_MID_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB2_MID_BOX.started')
                    SLF_FB2_MID_BOX.setAutoDraw(True)
                if SLF_FB2_MID_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_FB2_MID_BOX.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_FB2_MID_BOX.tStop = t  # not accounting for scr refresh
                        SLF_FB2_MID_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_FB2_MID_BOX.stopped')
                        SLF_FB2_MID_BOX.setAutoDraw(False)
                
                # *SLF_FB2_MID_TEXT* updates
                if SLF_FB2_MID_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_FB2_MID_TEXT.frameNStart = frameN  # exact frame index
                    SLF_FB2_MID_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_FB2_MID_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_FB2_MID_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB2_MID_TEXT.started')
                    SLF_FB2_MID_TEXT.setAutoDraw(True)
                if SLF_FB2_MID_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_FB2_MID_TEXT.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_FB2_MID_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_FB2_MID_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_FB2_MID_TEXT.stopped')
                        SLF_FB2_MID_TEXT.setAutoDraw(False)
                
                # *SLF_FB2_HIGH_BOX* updates
                if SLF_FB2_HIGH_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_FB2_HIGH_BOX.frameNStart = frameN  # exact frame index
                    SLF_FB2_HIGH_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_FB2_HIGH_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_FB2_HIGH_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB2_HIGH_BOX.started')
                    SLF_FB2_HIGH_BOX.setAutoDraw(True)
                if SLF_FB2_HIGH_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_FB2_HIGH_BOX.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_FB2_HIGH_BOX.tStop = t  # not accounting for scr refresh
                        SLF_FB2_HIGH_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_FB2_HIGH_BOX.stopped')
                        SLF_FB2_HIGH_BOX.setAutoDraw(False)
                
                # *SLF_FB2_HIGH_TEXT* updates
                if SLF_FB2_HIGH_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_FB2_HIGH_TEXT.frameNStart = frameN  # exact frame index
                    SLF_FB2_HIGH_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_FB2_HIGH_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_FB2_HIGH_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB2_HIGH_TEXT.started')
                    SLF_FB2_HIGH_TEXT.setAutoDraw(True)
                if SLF_FB2_HIGH_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_FB2_HIGH_TEXT.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_FB2_HIGH_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_FB2_HIGH_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_FB2_HIGH_TEXT.stopped')
                        SLF_FB2_HIGH_TEXT.setAutoDraw(False)
                
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    routineForceEnded = True
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in SlfConfFbComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "SlfConfFb" ---
            for thisComponent in SlfConfFbComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # the Routine "SlfConfFb" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            thisExp.nextEntry()
            
        # completed is_slf_resp repeats of 'SlfSkip'
        
        # get names of stimulus parameters
        if SlfSkip.trialList in ([], [None], None):
            params = []
        else:
            params = SlfSkip.trialList[0].keys()
        # save data for this loop
        SlfSkip.saveAsExcel(filename + '.xlsx', sheetName='SlfSkip',
            stimOut=params,
            dataOut=['n','all_mean','all_std', 'all_raw'])
        SlfSkip.saveAsText(filename + 'SlfSkip.csv', delim=',',
            stimOut=params,
            dataOut=['n','all_mean','all_std', 'all_raw'])
        
        # --- Prepare to start Routine "SlfReward" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        # Run 'Begin Routine' code from SLF_REWARD_CODE
        # Get the presented time in slf-reward
        slf_reward_pres_time = onset.getTime()
        SLF_FB3_FRAME.setPos(slf_choice_pos)
        SLF_FB3_IMG1.setImage('img/' + slf_img1)
        SLF_FB3_IMG2.setImage('img/' + slf_img2)
        SLF_REWARD.setText(slf_reward_text)
        # keep track of which components have finished
        SlfRewardComponents = [SLF_FB3_FRAME, SLF_FB3_IMG1, SLF_FB3_IMG2, SLF_REWARD]
        for thisComponent in SlfRewardComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "SlfReward" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *SLF_FB3_FRAME* updates
            if SLF_FB3_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FB3_FRAME.frameNStart = frameN  # exact frame index
                SLF_FB3_FRAME.tStart = t  # local t and not account for scr refresh
                SLF_FB3_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FB3_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FB3_FRAME.started')
                SLF_FB3_FRAME.setAutoDraw(True)
            if SLF_FB3_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FB3_FRAME.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FB3_FRAME.tStop = t  # not accounting for scr refresh
                    SLF_FB3_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB3_FRAME.stopped')
                    SLF_FB3_FRAME.setAutoDraw(False)
            
            # *SLF_FB3_IMG1* updates
            if SLF_FB3_IMG1.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FB3_IMG1.frameNStart = frameN  # exact frame index
                SLF_FB3_IMG1.tStart = t  # local t and not account for scr refresh
                SLF_FB3_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FB3_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FB3_IMG1.started')
                SLF_FB3_IMG1.setAutoDraw(True)
            if SLF_FB3_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FB3_IMG1.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FB3_IMG1.tStop = t  # not accounting for scr refresh
                    SLF_FB3_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB3_IMG1.stopped')
                    SLF_FB3_IMG1.setAutoDraw(False)
            
            # *SLF_FB3_IMG2* updates
            if SLF_FB3_IMG2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_FB3_IMG2.frameNStart = frameN  # exact frame index
                SLF_FB3_IMG2.tStart = t  # local t and not account for scr refresh
                SLF_FB3_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_FB3_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_FB3_IMG2.started')
                SLF_FB3_IMG2.setAutoDraw(True)
            if SLF_FB3_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_FB3_IMG2.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_FB3_IMG2.tStop = t  # not accounting for scr refresh
                    SLF_FB3_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_FB3_IMG2.stopped')
                    SLF_FB3_IMG2.setAutoDraw(False)
            
            # *SLF_REWARD* updates
            if SLF_REWARD.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                SLF_REWARD.frameNStart = frameN  # exact frame index
                SLF_REWARD.tStart = t  # local t and not account for scr refresh
                SLF_REWARD.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_REWARD, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_REWARD.started')
                SLF_REWARD.setAutoDraw(True)
            if SLF_REWARD.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_REWARD.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_REWARD.tStop = t  # not accounting for scr refresh
                    SLF_REWARD.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_REWARD.stopped')
                    SLF_REWARD.setAutoDraw(False)
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in SlfRewardComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "SlfReward" ---
        for thisComponent in SlfRewardComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # Run 'End Routine' code from SLF_REWARD_CODE
        # Get the responsed time in slf-reward
        slf_reward_resp_time = onset.getTime()
        
        slf_choice_time = slf_choice_resp_time - slf_choice_pres_time
        slf_conf_time = slf_conf_resp_time - slf_conf_pres_time
        
        # Output slf data
        slf_data_table[slf_serial_num] = [game_serial_num, slf_serial_num+1, block_num,
                                         slf_seq_pattern, slf_loc_pattern, 
                                         slf_loc1, slf_loc2, slf_scale, 
                                         slf_pt, slf_conf_resp, slf_reward,
                                         np.round(slf_choice_time, 3), np.round(slf_conf_time, 3), 
                                         slf_img.index(slf_img1)+1, slf_img.index(slf_img2)+1,
                                         slf_choice_resp, slf_correct, slf_choice_loc, 
                                         np.round(slf_choice_pres_time, 3), np.round(slf_choice_resp_time, 3),
                                         np.round(slf_conf_pres_time, 3), np.round(slf_conf_resp_time, 3),
                                         np.round(slf_reward_pres_time, 3), np.round(slf_reward_resp_time, 3)]
        
        # Save matlab file
        io.savemat(out_name + '.mat', {'out_mat':out_dict})
        
        # Save excel file
        out_book = load_workbook(filename=out_xlsx)
        slf_sheet_name = out_book.sheetnames[0]
        slf_sheet = out_book[slf_sheet_name]
        for i in range(1, len(slf_data_table[slf_serial_num])+1):
            slf_sheet.cell(slf_serial_num+1, i, value = slf_data_table[slf_serial_num][i-1])
        out_book.save(out_xlsx)
        
        # Update slf-count
        slf_serial_num += 1
        game_serial_num += 1
        # the Routine "SlfReward" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        thisExp.nextEntry()
        
    # completed 1.0 repeats of 'SlfSequence'
    
    # get names of stimulus parameters
    if SlfSequence.trialList in ([], [None], None):
        params = []
    else:
        params = SlfSequence.trialList[0].keys()
    # save data for this loop
    SlfSequence.saveAsExcel(filename + '.xlsx', sheetName='SlfSequence',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    SlfSequence.saveAsText(filename + 'SlfSequence.csv', delim=',',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    
    # --- Prepare to start Routine "ObsInstr" ---
    continueRoutine = True
    routineForceEnded = False
    # update component parameters for each repeat
    # keep track of which components have finished
    ObsInstrComponents = [OBS_INSTR]
    for thisComponent in ObsInstrComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "ObsInstr" ---
    while continueRoutine and routineTimer.getTime() < 2.0:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *OBS_INSTR* updates
        if OBS_INSTR.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            OBS_INSTR.frameNStart = frameN  # exact frame index
            OBS_INSTR.tStart = t  # local t and not account for scr refresh
            OBS_INSTR.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(OBS_INSTR, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'OBS_INSTR.started')
            OBS_INSTR.setAutoDraw(True)
        if OBS_INSTR.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > OBS_INSTR.tStartRefresh + 2-frameTolerance:
                # keep track of stop time/frame for later
                OBS_INSTR.tStop = t  # not accounting for scr refresh
                OBS_INSTR.frameNStop = frameN  # exact frame index
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_INSTR.stopped')
                OBS_INSTR.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in ObsInstrComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "ObsInstr" ---
    for thisComponent in ObsInstrComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
    if routineForceEnded:
        routineTimer.reset()
    else:
        routineTimer.addTime(-2.000000)
    
    # set up handler to look after randomisation of conditions etc
    ObsSequence = data.TrialHandler(nReps=1.0, method='sequential', 
        extraInfo=expInfo, originPath=-1,
        trialList=data.importConditions('sequence/' + obs_file),
        seed=None, name='ObsSequence')
    thisExp.addLoop(ObsSequence)  # add the loop to the experiment
    thisObsSequence = ObsSequence.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb = thisObsSequence.rgb)
    if thisObsSequence != None:
        for paramName in thisObsSequence:
            exec('{} = thisObsSequence[paramName]'.format(paramName))
    
    for thisObsSequence in ObsSequence:
        currentLoop = ObsSequence
        # abbreviate parameter names if possible (e.g. rgb = thisObsSequence.rgb)
        if thisObsSequence != None:
            for paramName in thisObsSequence:
                exec('{} = thisObsSequence[paramName]'.format(paramName))
        
        # --- Prepare to start Routine "ObsChoice" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        # Run 'Begin Routine' code from OBS_CHOICE_CODE
        # Determine the images and parameters based on the self record
        obs_img1 = obs_dict[obs_loc1]
        obs_img2 = obs_dict[obs_loc2]
        
        obs_choice_pos = (-200, 0) if obs_choice_resp == 1 else\
                         (200, 0) if obs_choice_resp == 2 else\
                         (100000, 0)
                         
        obs_conf_pos = (-250, 0) if obs_conf_resp == 7 else\
                       (0, 0) if obs_conf_resp == 8 else\
                       (250, 0) if obs_conf_resp == 9 else\
                       (100000, 0)
        OBS_FB1_FRAME.setPos(obs_choice_pos)
        OBS_IMG1.setPos((-200, 0))
        OBS_IMG1.setImage('img/' + obs_img1)
        OBS_IMG2.setImage('img/' + obs_img2)
        # keep track of which components have finished
        ObsChoiceComponents = [OBS_FIX, OBS_FB1_FRAME, OBS_IMG1, OBS_IMG2]
        for thisComponent in ObsChoiceComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "ObsChoice" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *OBS_FIX* updates
            if OBS_FIX.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_FIX.frameNStart = frameN  # exact frame index
                OBS_FIX.tStart = t  # local t and not account for scr refresh
                OBS_FIX.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_FIX, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_FIX.started')
                OBS_FIX.setAutoDraw(True)
            if OBS_FIX.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_FIX.tStartRefresh + obs_choice_time + feedback_time + 0.5-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_FIX.tStop = t  # not accounting for scr refresh
                    OBS_FIX.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_FIX.stopped')
                    OBS_FIX.setAutoDraw(False)
            
            # *OBS_FB1_FRAME* updates
            if OBS_FB1_FRAME.status == NOT_STARTED and tThisFlip >= obs_choice_time + 0.5-frameTolerance:
                # keep track of start time/frame for later
                OBS_FB1_FRAME.frameNStart = frameN  # exact frame index
                OBS_FB1_FRAME.tStart = t  # local t and not account for scr refresh
                OBS_FB1_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_FB1_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_FB1_FRAME.started')
                OBS_FB1_FRAME.setAutoDraw(True)
            if OBS_FB1_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_FB1_FRAME.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_FB1_FRAME.tStop = t  # not accounting for scr refresh
                    OBS_FB1_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_FB1_FRAME.stopped')
                    OBS_FB1_FRAME.setAutoDraw(False)
            
            # *OBS_IMG1* updates
            if OBS_IMG1.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                OBS_IMG1.frameNStart = frameN  # exact frame index
                OBS_IMG1.tStart = t  # local t and not account for scr refresh
                OBS_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_IMG1.started')
                OBS_IMG1.setAutoDraw(True)
            if OBS_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_IMG1.tStartRefresh + obs_choice_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_IMG1.tStop = t  # not accounting for scr refresh
                    OBS_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_IMG1.stopped')
                    OBS_IMG1.setAutoDraw(False)
            
            # *OBS_IMG2* updates
            if OBS_IMG2.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                OBS_IMG2.frameNStart = frameN  # exact frame index
                OBS_IMG2.tStart = t  # local t and not account for scr refresh
                OBS_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_IMG2.started')
                OBS_IMG2.setAutoDraw(True)
            if OBS_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_IMG2.tStartRefresh + obs_choice_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_IMG2.tStop = t  # not accounting for scr refresh
                    OBS_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_IMG2.stopped')
                    OBS_IMG2.setAutoDraw(False)
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in ObsChoiceComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "ObsChoice" ---
        for thisComponent in ObsChoiceComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # the Routine "ObsChoice" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # --- Prepare to start Routine "ObsConf" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        OBS_FB2_FRAME.setPos(obs_conf_pos)
        # keep track of which components have finished
        ObsConfComponents = [OBS_FB2_FRAME, OBS_LOW_BOX, OBS_LOW_TEXT, OBS_MID_BOX, OBS_MID_TEXT, OBS_HIGH_BOX, OBS_HIGH_TEXT]
        for thisComponent in ObsConfComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "ObsConf" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *OBS_FB2_FRAME* updates
            if OBS_FB2_FRAME.status == NOT_STARTED and tThisFlip >= obs_conf_time-frameTolerance:
                # keep track of start time/frame for later
                OBS_FB2_FRAME.frameNStart = frameN  # exact frame index
                OBS_FB2_FRAME.tStart = t  # local t and not account for scr refresh
                OBS_FB2_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_FB2_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_FB2_FRAME.started')
                OBS_FB2_FRAME.setAutoDraw(True)
            if OBS_FB2_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_FB2_FRAME.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_FB2_FRAME.tStop = t  # not accounting for scr refresh
                    OBS_FB2_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_FB2_FRAME.stopped')
                    OBS_FB2_FRAME.setAutoDraw(False)
            
            # *OBS_LOW_BOX* updates
            if OBS_LOW_BOX.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_LOW_BOX.frameNStart = frameN  # exact frame index
                OBS_LOW_BOX.tStart = t  # local t and not account for scr refresh
                OBS_LOW_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_LOW_BOX, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_LOW_BOX.started')
                OBS_LOW_BOX.setAutoDraw(True)
            if OBS_LOW_BOX.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_LOW_BOX.tStartRefresh + obs_conf_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_LOW_BOX.tStop = t  # not accounting for scr refresh
                    OBS_LOW_BOX.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_LOW_BOX.stopped')
                    OBS_LOW_BOX.setAutoDraw(False)
            
            # *OBS_LOW_TEXT* updates
            if OBS_LOW_TEXT.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_LOW_TEXT.frameNStart = frameN  # exact frame index
                OBS_LOW_TEXT.tStart = t  # local t and not account for scr refresh
                OBS_LOW_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_LOW_TEXT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_LOW_TEXT.started')
                OBS_LOW_TEXT.setAutoDraw(True)
            if OBS_LOW_TEXT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_LOW_TEXT.tStartRefresh + obs_conf_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_LOW_TEXT.tStop = t  # not accounting for scr refresh
                    OBS_LOW_TEXT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_LOW_TEXT.stopped')
                    OBS_LOW_TEXT.setAutoDraw(False)
            
            # *OBS_MID_BOX* updates
            if OBS_MID_BOX.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_MID_BOX.frameNStart = frameN  # exact frame index
                OBS_MID_BOX.tStart = t  # local t and not account for scr refresh
                OBS_MID_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_MID_BOX, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_MID_BOX.started')
                OBS_MID_BOX.setAutoDraw(True)
            if OBS_MID_BOX.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_MID_BOX.tStartRefresh + obs_conf_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_MID_BOX.tStop = t  # not accounting for scr refresh
                    OBS_MID_BOX.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_MID_BOX.stopped')
                    OBS_MID_BOX.setAutoDraw(False)
            
            # *OBS_MID_TEXT* updates
            if OBS_MID_TEXT.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_MID_TEXT.frameNStart = frameN  # exact frame index
                OBS_MID_TEXT.tStart = t  # local t and not account for scr refresh
                OBS_MID_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_MID_TEXT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_MID_TEXT.started')
                OBS_MID_TEXT.setAutoDraw(True)
            if OBS_MID_TEXT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_MID_TEXT.tStartRefresh + obs_conf_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_MID_TEXT.tStop = t  # not accounting for scr refresh
                    OBS_MID_TEXT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_MID_TEXT.stopped')
                    OBS_MID_TEXT.setAutoDraw(False)
            
            # *OBS_HIGH_BOX* updates
            if OBS_HIGH_BOX.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_HIGH_BOX.frameNStart = frameN  # exact frame index
                OBS_HIGH_BOX.tStart = t  # local t and not account for scr refresh
                OBS_HIGH_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_HIGH_BOX, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_HIGH_BOX.started')
                OBS_HIGH_BOX.setAutoDraw(True)
            if OBS_HIGH_BOX.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_HIGH_BOX.tStartRefresh + obs_conf_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_HIGH_BOX.tStop = t  # not accounting for scr refresh
                    OBS_HIGH_BOX.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_HIGH_BOX.stopped')
                    OBS_HIGH_BOX.setAutoDraw(False)
            
            # *OBS_HIGH_TEXT* updates
            if OBS_HIGH_TEXT.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_HIGH_TEXT.frameNStart = frameN  # exact frame index
                OBS_HIGH_TEXT.tStart = t  # local t and not account for scr refresh
                OBS_HIGH_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_HIGH_TEXT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_HIGH_TEXT.started')
                OBS_HIGH_TEXT.setAutoDraw(True)
            if OBS_HIGH_TEXT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_HIGH_TEXT.tStartRefresh + obs_conf_time + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_HIGH_TEXT.tStop = t  # not accounting for scr refresh
                    OBS_HIGH_TEXT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_HIGH_TEXT.stopped')
                    OBS_HIGH_TEXT.setAutoDraw(False)
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in ObsConfComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "ObsConf" ---
        for thisComponent in ObsConfComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # the Routine "ObsConf" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # --- Prepare to start Routine "ObsReward" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        OBS_FB3_FRAME.setPos(obs_choice_pos)
        OBS_FB3_IMG1.setImage('img/' + obs_img1)
        OBS_FB3_IMG2.setImage('img/' + obs_img2)
        OBS_REWARD.setText('?')
        # keep track of which components have finished
        ObsRewardComponents = [OBS_FB3_FRAME, OBS_FB3_IMG1, OBS_FB3_IMG2, OBS_REWARD]
        for thisComponent in ObsRewardComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "ObsReward" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *OBS_FB3_FRAME* updates
            if OBS_FB3_FRAME.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_FB3_FRAME.frameNStart = frameN  # exact frame index
                OBS_FB3_FRAME.tStart = t  # local t and not account for scr refresh
                OBS_FB3_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_FB3_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_FB3_FRAME.started')
                OBS_FB3_FRAME.setAutoDraw(True)
            if OBS_FB3_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_FB3_FRAME.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_FB3_FRAME.tStop = t  # not accounting for scr refresh
                    OBS_FB3_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_FB3_FRAME.stopped')
                    OBS_FB3_FRAME.setAutoDraw(False)
            
            # *OBS_FB3_IMG1* updates
            if OBS_FB3_IMG1.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_FB3_IMG1.frameNStart = frameN  # exact frame index
                OBS_FB3_IMG1.tStart = t  # local t and not account for scr refresh
                OBS_FB3_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_FB3_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_FB3_IMG1.started')
                OBS_FB3_IMG1.setAutoDraw(True)
            if OBS_FB3_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_FB3_IMG1.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_FB3_IMG1.tStop = t  # not accounting for scr refresh
                    OBS_FB3_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_FB3_IMG1.stopped')
                    OBS_FB3_IMG1.setAutoDraw(False)
            
            # *OBS_FB3_IMG2* updates
            if OBS_FB3_IMG2.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_FB3_IMG2.frameNStart = frameN  # exact frame index
                OBS_FB3_IMG2.tStart = t  # local t and not account for scr refresh
                OBS_FB3_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_FB3_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_FB3_IMG2.started')
                OBS_FB3_IMG2.setAutoDraw(True)
            if OBS_FB3_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_FB3_IMG2.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_FB3_IMG2.tStop = t  # not accounting for scr refresh
                    OBS_FB3_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_FB3_IMG2.stopped')
                    OBS_FB3_IMG2.setAutoDraw(False)
            
            # *OBS_REWARD* updates
            if OBS_REWARD.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                OBS_REWARD.frameNStart = frameN  # exact frame index
                OBS_REWARD.tStart = t  # local t and not account for scr refresh
                OBS_REWARD.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(OBS_REWARD, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'OBS_REWARD.started')
                OBS_REWARD.setAutoDraw(True)
            if OBS_REWARD.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > OBS_REWARD.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    OBS_REWARD.tStop = t  # not accounting for scr refresh
                    OBS_REWARD.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'OBS_REWARD.stopped')
                    OBS_REWARD.setAutoDraw(False)
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in ObsRewardComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "ObsReward" ---
        for thisComponent in ObsRewardComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # Run 'End Routine' code from OBS_REWARD_CODE
        # Output obs data
        obs_data_table[obs_serial_num] = [game_serial_num, obs_serial_num+1, block_num, 
                                         obs_seq_pattern, obs_loc_pattern, 
                                         obs_loc1, obs_loc2, obs_scale, 
                                         obs_pt, obs_conf_resp, obs_reward, 
                                         np.round(obs_choice_time, 3), np.round(obs_conf_time, 3),
                                         obs_img1_idx, obs_img2_idx,
                                         obs_choice_resp, obs_correct, obs_choice_loc, 
                                         np.round(obs_choice_pres_time, 3), np.round(obs_choice_resp_time, 3),
                                         np.round(obs_conf_pres_time, 3), np.round(obs_conf_resp_time, 3),
                                         np.round(obs_reward_pres_time, 3), np.round(obs_reward_resp_time, 3)]
        
        # Save matlab file
        io.savemat(out_name + '.mat', {'out_mat':out_dict})
        
        # Save excel file
        out_book = load_workbook(filename=out_xlsx)
        obs_sheet_name = out_book.sheetnames[1]
        obs_sheet = out_book[obs_sheet_name]
        for i in range(1, len(obs_data_table[obs_serial_num])+1):
            obs_sheet.cell(obs_serial_num+1, i, value = obs_data_table[obs_serial_num][i-1])
        out_book.save(out_xlsx)
        
        # Update obs-count
        obs_serial_num += 1
        game_serial_num += 1
        # the Routine "ObsReward" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        thisExp.nextEntry()
        
    # completed 1.0 repeats of 'ObsSequence'
    
    # get names of stimulus parameters
    if ObsSequence.trialList in ([], [None], None):
        params = []
    else:
        params = ObsSequence.trialList[0].keys()
    # save data for this loop
    ObsSequence.saveAsExcel(filename + '.xlsx', sheetName='ObsSequence',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    ObsSequence.saveAsText(filename + 'ObsSequence.csv', delim=',',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    
    # --- Prepare to start Routine "TestInstr" ---
    continueRoutine = True
    routineForceEnded = False
    # update component parameters for each repeat
    # keep track of which components have finished
    TestInstrComponents = [TEST_INSTR]
    for thisComponent in TestInstrComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "TestInstr" ---
    while continueRoutine and routineTimer.getTime() < 2.0:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *TEST_INSTR* updates
        if TEST_INSTR.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
            # keep track of start time/frame for later
            TEST_INSTR.frameNStart = frameN  # exact frame index
            TEST_INSTR.tStart = t  # local t and not account for scr refresh
            TEST_INSTR.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(TEST_INSTR, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'TEST_INSTR.started')
            TEST_INSTR.setAutoDraw(True)
        if TEST_INSTR.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > TEST_INSTR.tStartRefresh + 2-frameTolerance:
                # keep track of stop time/frame for later
                TEST_INSTR.tStop = t  # not accounting for scr refresh
                TEST_INSTR.frameNStop = frameN  # exact frame index
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_INSTR.stopped')
                TEST_INSTR.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in TestInstrComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "TestInstr" ---
    for thisComponent in TestInstrComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
    if routineForceEnded:
        routineTimer.reset()
    else:
        routineTimer.addTime(-2.000000)
    
    # set up handler to look after randomisation of conditions etc
    TestSequence = data.TrialHandler(nReps=1.0, method='sequential', 
        extraInfo=expInfo, originPath=-1,
        trialList=data.importConditions('sequence/' + test_file),
        seed=None, name='TestSequence')
    thisExp.addLoop(TestSequence)  # add the loop to the experiment
    thisTestSequence = TestSequence.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb = thisTestSequence.rgb)
    if thisTestSequence != None:
        for paramName in thisTestSequence:
            exec('{} = thisTestSequence[paramName]'.format(paramName))
    
    for thisTestSequence in TestSequence:
        currentLoop = TestSequence
        # abbreviate parameter names if possible (e.g. rgb = thisTestSequence.rgb)
        if thisTestSequence != None:
            for paramName in thisTestSequence:
                exec('{} = thisTestSequence[paramName]'.format(paramName))
        
        # --- Prepare to start Routine "TestChoice" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        # Run 'Begin Routine' code from TEST_CHOICE_CODE
        if test_seq_pattern == 1: # SS condition
            test_img1 = slf_dict[test_loc1]
            test_img2 = slf_dict[test_loc2]
            
        elif test_seq_pattern == 2: # OO condition
            test_img1 = obs_dict[test_loc1]
            test_img2 = obs_dict[test_loc2]
        
        elif test_seq_pattern == 3: # SO condition
            test_img1 = slf_dict[test_loc1]
            test_img2 = obs_dict[test_loc2]
        
        elif test_seq_pattern == 4: # OS condition
            test_img1 = obs_dict[test_loc1]
            test_img2 = slf_dict[test_loc2]
        
        #if test_loc1 == test_loc2:
        #    test_seq_pattern = 5
        
        # Get the presented time in test-choice
        test_choice_pres_time = onset.getTime() + 0.5 # Consider the fixation
        TEST_IMG1.setImage('img/' + test_img1)
        TEST_IMG2.setImage('img/' + test_img2)
        TEST_CHOICE_RESP.keys = []
        TEST_CHOICE_RESP.rt = []
        _TEST_CHOICE_RESP_allKeys = []
        # keep track of which components have finished
        TestChoiceComponents = [TEST_FIX1, TEST_IMG1, TEST_IMG2, TEST_CHOICE_RESP]
        for thisComponent in TestChoiceComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "TestChoice" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *TEST_FIX1* updates
            if TEST_FIX1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FIX1.frameNStart = frameN  # exact frame index
                TEST_FIX1.tStart = t  # local t and not account for scr refresh
                TEST_FIX1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FIX1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FIX1.started')
                TEST_FIX1.setAutoDraw(True)
            if TEST_FIX1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FIX1.tStartRefresh + choice_time_max+0.5-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FIX1.tStop = t  # not accounting for scr refresh
                    TEST_FIX1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FIX1.stopped')
                    TEST_FIX1.setAutoDraw(False)
            
            # *TEST_IMG1* updates
            if TEST_IMG1.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                TEST_IMG1.frameNStart = frameN  # exact frame index
                TEST_IMG1.tStart = t  # local t and not account for scr refresh
                TEST_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_IMG1.started')
                TEST_IMG1.setAutoDraw(True)
            if TEST_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_IMG1.tStartRefresh + choice_time_max-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_IMG1.tStop = t  # not accounting for scr refresh
                    TEST_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_IMG1.stopped')
                    TEST_IMG1.setAutoDraw(False)
            
            # *TEST_IMG2* updates
            if TEST_IMG2.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                TEST_IMG2.frameNStart = frameN  # exact frame index
                TEST_IMG2.tStart = t  # local t and not account for scr refresh
                TEST_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_IMG2.started')
                TEST_IMG2.setAutoDraw(True)
            if TEST_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_IMG2.tStartRefresh + choice_time_max-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_IMG2.tStop = t  # not accounting for scr refresh
                    TEST_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_IMG2.stopped')
                    TEST_IMG2.setAutoDraw(False)
            
            # *TEST_CHOICE_RESP* updates
            waitOnFlip = False
            if TEST_CHOICE_RESP.status == NOT_STARTED and tThisFlip >= 0.5-frameTolerance:
                # keep track of start time/frame for later
                TEST_CHOICE_RESP.frameNStart = frameN  # exact frame index
                TEST_CHOICE_RESP.tStart = t  # local t and not account for scr refresh
                TEST_CHOICE_RESP.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_CHOICE_RESP, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_CHOICE_RESP.started')
                TEST_CHOICE_RESP.status = STARTED
                # keyboard checking is just starting
                waitOnFlip = True
                win.callOnFlip(TEST_CHOICE_RESP.clock.reset)  # t=0 on next screen flip
                win.callOnFlip(TEST_CHOICE_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
            if TEST_CHOICE_RESP.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_CHOICE_RESP.tStartRefresh + choice_time_max-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_CHOICE_RESP.tStop = t  # not accounting for scr refresh
                    TEST_CHOICE_RESP.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_CHOICE_RESP.stopped')
                    TEST_CHOICE_RESP.status = FINISHED
            if TEST_CHOICE_RESP.status == STARTED and not waitOnFlip:
                theseKeys = TEST_CHOICE_RESP.getKeys(keyList=[left_key,right_key], waitRelease=False)
                _TEST_CHOICE_RESP_allKeys.extend(theseKeys)
                if len(_TEST_CHOICE_RESP_allKeys):
                    TEST_CHOICE_RESP.keys = _TEST_CHOICE_RESP_allKeys[-1].name  # just the last key pressed
                    TEST_CHOICE_RESP.rt = _TEST_CHOICE_RESP_allKeys[-1].rt
                    # was this correct?
                    if (TEST_CHOICE_RESP.keys == str('')) or (TEST_CHOICE_RESP.keys == ''):
                        TEST_CHOICE_RESP.corr = 1
                    else:
                        TEST_CHOICE_RESP.corr = 0
                    # a response ends the routine
                    continueRoutine = False
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in TestChoiceComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "TestChoice" ---
        for thisComponent in TestChoiceComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # Run 'End Routine' code from TEST_CHOICE_CODE
        # Get the responsed time in test-choice
        test_choice_resp_time = onset.getTime()
        
        # Initialize in case skipping slf-conf
        test_conf_resp = 0
        test_conf_pres_time = 0
        test_conf_resp_time = 0
        is_test_resp = 0
        
        # Receive choice responses and determine reward and feedback parameters
        if TEST_CHOICE_RESP.keys == left_key: 
            test_choice_pos = (-200,0)
            test_choice_resp = 1
            test_choice_loc = test_loc1
            test_reward = 0
            test_reward_text = "?"
            is_test_resp = 1
        elif TEST_CHOICE_RESP.keys == right_key: 
            test_choice_pos = (200,0)
            test_choice_resp = 2
            test_choice_loc = test_loc2
            test_reward = 0
            test_reward_text = "?"
            is_test_resp = 1
        else:
            test_choice_pos = (100000,0)
            test_choice_resp = 0
            test_choice_loc = 0
            test_reward = 0
            test_reward_text = "F"
        
        # Add points according to the response
        test_pt = 2 if test_correct == 3 else\
                  1 if test_choice_resp == test_correct else 0
        # check responses
        if TEST_CHOICE_RESP.keys in ['', [], None]:  # No response was made
            TEST_CHOICE_RESP.keys = None
            # was no response the correct answer?!
            if str('').lower() == 'none':
               TEST_CHOICE_RESP.corr = 1;  # correct non-response
            else:
               TEST_CHOICE_RESP.corr = 0;  # failed to respond (incorrectly)
        # store data for TestSequence (TrialHandler)
        TestSequence.addData('TEST_CHOICE_RESP.keys',TEST_CHOICE_RESP.keys)
        TestSequence.addData('TEST_CHOICE_RESP.corr', TEST_CHOICE_RESP.corr)
        if TEST_CHOICE_RESP.keys != None:  # we had a response
            TestSequence.addData('TEST_CHOICE_RESP.rt', TEST_CHOICE_RESP.rt)
        # the Routine "TestChoice" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # --- Prepare to start Routine "TestChoiceFb" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        TEST_FB1_FRAME.setPos(test_choice_pos)
        TEST_FB1_IMG1.setImage('img/' + test_img1)
        TEST_FB1_IMG2.setImage('img/' + test_img2)
        # keep track of which components have finished
        TestChoiceFbComponents = [TEST_FIX2, TEST_FB1_FRAME, TEST_FB1_IMG1, TEST_FB1_IMG2]
        for thisComponent in TestChoiceFbComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "TestChoiceFb" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *TEST_FIX2* updates
            if TEST_FIX2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FIX2.frameNStart = frameN  # exact frame index
                TEST_FIX2.tStart = t  # local t and not account for scr refresh
                TEST_FIX2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FIX2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FIX2.started')
                TEST_FIX2.setAutoDraw(True)
            if TEST_FIX2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FIX2.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FIX2.tStop = t  # not accounting for scr refresh
                    TEST_FIX2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FIX2.stopped')
                    TEST_FIX2.setAutoDraw(False)
            
            # *TEST_FB1_FRAME* updates
            if TEST_FB1_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FB1_FRAME.frameNStart = frameN  # exact frame index
                TEST_FB1_FRAME.tStart = t  # local t and not account for scr refresh
                TEST_FB1_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FB1_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FB1_FRAME.started')
                TEST_FB1_FRAME.setAutoDraw(True)
            if TEST_FB1_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FB1_FRAME.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FB1_FRAME.tStop = t  # not accounting for scr refresh
                    TEST_FB1_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB1_FRAME.stopped')
                    TEST_FB1_FRAME.setAutoDraw(False)
            
            # *TEST_FB1_IMG1* updates
            if TEST_FB1_IMG1.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FB1_IMG1.frameNStart = frameN  # exact frame index
                TEST_FB1_IMG1.tStart = t  # local t and not account for scr refresh
                TEST_FB1_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FB1_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FB1_IMG1.started')
                TEST_FB1_IMG1.setAutoDraw(True)
            if TEST_FB1_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FB1_IMG1.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FB1_IMG1.tStop = t  # not accounting for scr refresh
                    TEST_FB1_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB1_IMG1.stopped')
                    TEST_FB1_IMG1.setAutoDraw(False)
            
            # *TEST_FB1_IMG2* updates
            if TEST_FB1_IMG2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FB1_IMG2.frameNStart = frameN  # exact frame index
                TEST_FB1_IMG2.tStart = t  # local t and not account for scr refresh
                TEST_FB1_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FB1_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FB1_IMG2.started')
                TEST_FB1_IMG2.setAutoDraw(True)
            if TEST_FB1_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FB1_IMG2.tStartRefresh + feedback_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FB1_IMG2.tStop = t  # not accounting for scr refresh
                    TEST_FB1_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB1_IMG2.stopped')
                    TEST_FB1_IMG2.setAutoDraw(False)
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in TestChoiceFbComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "TestChoiceFb" ---
        for thisComponent in TestChoiceFbComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # the Routine "TestChoiceFb" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        # set up handler to look after randomisation of conditions etc
        TestSkip = data.TrialHandler(nReps=is_test_resp, method='random', 
            extraInfo=expInfo, originPath=-1,
            trialList=[None],
            seed=None, name='TestSkip')
        thisExp.addLoop(TestSkip)  # add the loop to the experiment
        thisTestSkip = TestSkip.trialList[0]  # so we can initialise stimuli with some values
        # abbreviate parameter names if possible (e.g. rgb = thisTestSkip.rgb)
        if thisTestSkip != None:
            for paramName in thisTestSkip:
                exec('{} = thisTestSkip[paramName]'.format(paramName))
        
        for thisTestSkip in TestSkip:
            currentLoop = TestSkip
            # abbreviate parameter names if possible (e.g. rgb = thisTestSkip.rgb)
            if thisTestSkip != None:
                for paramName in thisTestSkip:
                    exec('{} = thisTestSkip[paramName]'.format(paramName))
            
            # --- Prepare to start Routine "TestConf" ---
            continueRoutine = True
            routineForceEnded = False
            # update component parameters for each repeat
            # Run 'Begin Routine' code from TEST_CONF_CODE
            # Get the presented time in test-confidence
            test_conf_pres_time = onset.getTime()
            TEST_CONF_RESP.keys = []
            TEST_CONF_RESP.rt = []
            _TEST_CONF_RESP_allKeys = []
            # keep track of which components have finished
            TestConfComponents = [TEST_LOW_BOX, TEST_LOW_TEXT, SLF_TEST_BOX, TEST_MID_TEXT, TEST_HIGH_BOX, TEST_HIGH_TEXT, TEST_CONF_RESP]
            for thisComponent in TestConfComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "TestConf" ---
            while continueRoutine:
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *TEST_LOW_BOX* updates
                if TEST_LOW_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_LOW_BOX.frameNStart = frameN  # exact frame index
                    TEST_LOW_BOX.tStart = t  # local t and not account for scr refresh
                    TEST_LOW_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_LOW_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_LOW_BOX.started')
                    TEST_LOW_BOX.setAutoDraw(True)
                if TEST_LOW_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_LOW_BOX.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_LOW_BOX.tStop = t  # not accounting for scr refresh
                        TEST_LOW_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_LOW_BOX.stopped')
                        TEST_LOW_BOX.setAutoDraw(False)
                
                # *TEST_LOW_TEXT* updates
                if TEST_LOW_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_LOW_TEXT.frameNStart = frameN  # exact frame index
                    TEST_LOW_TEXT.tStart = t  # local t and not account for scr refresh
                    TEST_LOW_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_LOW_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_LOW_TEXT.started')
                    TEST_LOW_TEXT.setAutoDraw(True)
                if TEST_LOW_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_LOW_TEXT.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_LOW_TEXT.tStop = t  # not accounting for scr refresh
                        TEST_LOW_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_LOW_TEXT.stopped')
                        TEST_LOW_TEXT.setAutoDraw(False)
                
                # *SLF_TEST_BOX* updates
                if SLF_TEST_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_TEST_BOX.frameNStart = frameN  # exact frame index
                    SLF_TEST_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_TEST_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_TEST_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_TEST_BOX.started')
                    SLF_TEST_BOX.setAutoDraw(True)
                if SLF_TEST_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_TEST_BOX.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_TEST_BOX.tStop = t  # not accounting for scr refresh
                        SLF_TEST_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_TEST_BOX.stopped')
                        SLF_TEST_BOX.setAutoDraw(False)
                
                # *TEST_MID_TEXT* updates
                if TEST_MID_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_MID_TEXT.frameNStart = frameN  # exact frame index
                    TEST_MID_TEXT.tStart = t  # local t and not account for scr refresh
                    TEST_MID_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_MID_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_MID_TEXT.started')
                    TEST_MID_TEXT.setAutoDraw(True)
                if TEST_MID_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_MID_TEXT.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_MID_TEXT.tStop = t  # not accounting for scr refresh
                        TEST_MID_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_MID_TEXT.stopped')
                        TEST_MID_TEXT.setAutoDraw(False)
                
                # *TEST_HIGH_BOX* updates
                if TEST_HIGH_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_HIGH_BOX.frameNStart = frameN  # exact frame index
                    TEST_HIGH_BOX.tStart = t  # local t and not account for scr refresh
                    TEST_HIGH_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_HIGH_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_HIGH_BOX.started')
                    TEST_HIGH_BOX.setAutoDraw(True)
                if TEST_HIGH_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_HIGH_BOX.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_HIGH_BOX.tStop = t  # not accounting for scr refresh
                        TEST_HIGH_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_HIGH_BOX.stopped')
                        TEST_HIGH_BOX.setAutoDraw(False)
                
                # *TEST_HIGH_TEXT* updates
                if TEST_HIGH_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_HIGH_TEXT.frameNStart = frameN  # exact frame index
                    TEST_HIGH_TEXT.tStart = t  # local t and not account for scr refresh
                    TEST_HIGH_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_HIGH_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_HIGH_TEXT.started')
                    TEST_HIGH_TEXT.setAutoDraw(True)
                if TEST_HIGH_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_HIGH_TEXT.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_HIGH_TEXT.tStop = t  # not accounting for scr refresh
                        TEST_HIGH_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_HIGH_TEXT.stopped')
                        TEST_HIGH_TEXT.setAutoDraw(False)
                
                # *TEST_CONF_RESP* updates
                waitOnFlip = False
                if TEST_CONF_RESP.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_CONF_RESP.frameNStart = frameN  # exact frame index
                    TEST_CONF_RESP.tStart = t  # local t and not account for scr refresh
                    TEST_CONF_RESP.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_CONF_RESP, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_CONF_RESP.started')
                    TEST_CONF_RESP.status = STARTED
                    # keyboard checking is just starting
                    waitOnFlip = True
                    win.callOnFlip(TEST_CONF_RESP.clock.reset)  # t=0 on next screen flip
                    win.callOnFlip(TEST_CONF_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
                if TEST_CONF_RESP.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_CONF_RESP.tStartRefresh + conf_time_max-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_CONF_RESP.tStop = t  # not accounting for scr refresh
                        TEST_CONF_RESP.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_CONF_RESP.stopped')
                        TEST_CONF_RESP.status = FINISHED
                if TEST_CONF_RESP.status == STARTED and not waitOnFlip:
                    theseKeys = TEST_CONF_RESP.getKeys(keyList=[left_key,center_key,right_key], waitRelease=False)
                    _TEST_CONF_RESP_allKeys.extend(theseKeys)
                    if len(_TEST_CONF_RESP_allKeys):
                        TEST_CONF_RESP.keys = _TEST_CONF_RESP_allKeys[-1].name  # just the last key pressed
                        TEST_CONF_RESP.rt = _TEST_CONF_RESP_allKeys[-1].rt
                        # a response ends the routine
                        continueRoutine = False
                
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    routineForceEnded = True
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in TestConfComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "TestConf" ---
            for thisComponent in TestConfComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # Run 'End Routine' code from TEST_CONF_CODE
            # Get the responsed time in test-confidence
            test_conf_resp_time = onset.getTime()
            
            # Receive confidence responses in test and determine feedback parameters
            if TEST_CONF_RESP.keys == left_key: 
                test_conf_pos = (-250,0)
                test_conf_resp = 7
            elif TEST_CONF_RESP.keys == center_key: 
                test_conf_pos = (0,0)
                test_conf_resp = 8
            elif TEST_CONF_RESP.keys == right_key: 
                test_conf_pos = (250,0)
                test_conf_resp = 9
            else:
                test_conf_pos = (100000,0)
                test_conf_resp = 0
            # check responses
            if TEST_CONF_RESP.keys in ['', [], None]:  # No response was made
                TEST_CONF_RESP.keys = None
            TestSkip.addData('TEST_CONF_RESP.keys',TEST_CONF_RESP.keys)
            if TEST_CONF_RESP.keys != None:  # we had a response
                TestSkip.addData('TEST_CONF_RESP.rt', TEST_CONF_RESP.rt)
            # the Routine "TestConf" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            
            # --- Prepare to start Routine "TestConfFb" ---
            continueRoutine = True
            routineForceEnded = False
            # update component parameters for each repeat
            TEST_FB2_FRAME.setPos(test_conf_pos)
            # keep track of which components have finished
            TestConfFbComponents = [TEST_FB2_FRAME, TEST_FB2_LOW_BOX, TEST_FB2_LOW_TEXT, TEST_FB2_MID_BOX, TEST_FB2_MID_TEXT, TEST_FB2_HIGH_BOX, TEST_FB2_HIGH_TEXT]
            for thisComponent in TestConfFbComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            # --- Run Routine "TestConfFb" ---
            while continueRoutine:
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                # *TEST_FB2_FRAME* updates
                if TEST_FB2_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_FB2_FRAME.frameNStart = frameN  # exact frame index
                    TEST_FB2_FRAME.tStart = t  # local t and not account for scr refresh
                    TEST_FB2_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_FB2_FRAME, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB2_FRAME.started')
                    TEST_FB2_FRAME.setAutoDraw(True)
                if TEST_FB2_FRAME.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_FB2_FRAME.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_FB2_FRAME.tStop = t  # not accounting for scr refresh
                        TEST_FB2_FRAME.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_FB2_FRAME.stopped')
                        TEST_FB2_FRAME.setAutoDraw(False)
                
                # *TEST_FB2_LOW_BOX* updates
                if TEST_FB2_LOW_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_FB2_LOW_BOX.frameNStart = frameN  # exact frame index
                    TEST_FB2_LOW_BOX.tStart = t  # local t and not account for scr refresh
                    TEST_FB2_LOW_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_FB2_LOW_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB2_LOW_BOX.started')
                    TEST_FB2_LOW_BOX.setAutoDraw(True)
                if TEST_FB2_LOW_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_FB2_LOW_BOX.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_FB2_LOW_BOX.tStop = t  # not accounting for scr refresh
                        TEST_FB2_LOW_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_FB2_LOW_BOX.stopped')
                        TEST_FB2_LOW_BOX.setAutoDraw(False)
                
                # *TEST_FB2_LOW_TEXT* updates
                if TEST_FB2_LOW_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_FB2_LOW_TEXT.frameNStart = frameN  # exact frame index
                    TEST_FB2_LOW_TEXT.tStart = t  # local t and not account for scr refresh
                    TEST_FB2_LOW_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_FB2_LOW_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB2_LOW_TEXT.started')
                    TEST_FB2_LOW_TEXT.setAutoDraw(True)
                if TEST_FB2_LOW_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_FB2_LOW_TEXT.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_FB2_LOW_TEXT.tStop = t  # not accounting for scr refresh
                        TEST_FB2_LOW_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_FB2_LOW_TEXT.stopped')
                        TEST_FB2_LOW_TEXT.setAutoDraw(False)
                
                # *TEST_FB2_MID_BOX* updates
                if TEST_FB2_MID_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_FB2_MID_BOX.frameNStart = frameN  # exact frame index
                    TEST_FB2_MID_BOX.tStart = t  # local t and not account for scr refresh
                    TEST_FB2_MID_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_FB2_MID_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB2_MID_BOX.started')
                    TEST_FB2_MID_BOX.setAutoDraw(True)
                if TEST_FB2_MID_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_FB2_MID_BOX.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_FB2_MID_BOX.tStop = t  # not accounting for scr refresh
                        TEST_FB2_MID_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_FB2_MID_BOX.stopped')
                        TEST_FB2_MID_BOX.setAutoDraw(False)
                
                # *TEST_FB2_MID_TEXT* updates
                if TEST_FB2_MID_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_FB2_MID_TEXT.frameNStart = frameN  # exact frame index
                    TEST_FB2_MID_TEXT.tStart = t  # local t and not account for scr refresh
                    TEST_FB2_MID_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_FB2_MID_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB2_MID_TEXT.started')
                    TEST_FB2_MID_TEXT.setAutoDraw(True)
                if TEST_FB2_MID_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_FB2_MID_TEXT.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_FB2_MID_TEXT.tStop = t  # not accounting for scr refresh
                        TEST_FB2_MID_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_FB2_MID_TEXT.stopped')
                        TEST_FB2_MID_TEXT.setAutoDraw(False)
                
                # *TEST_FB2_HIGH_BOX* updates
                if TEST_FB2_HIGH_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_FB2_HIGH_BOX.frameNStart = frameN  # exact frame index
                    TEST_FB2_HIGH_BOX.tStart = t  # local t and not account for scr refresh
                    TEST_FB2_HIGH_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_FB2_HIGH_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB2_HIGH_BOX.started')
                    TEST_FB2_HIGH_BOX.setAutoDraw(True)
                if TEST_FB2_HIGH_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_FB2_HIGH_BOX.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_FB2_HIGH_BOX.tStop = t  # not accounting for scr refresh
                        TEST_FB2_HIGH_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_FB2_HIGH_BOX.stopped')
                        TEST_FB2_HIGH_BOX.setAutoDraw(False)
                
                # *TEST_FB2_HIGH_TEXT* updates
                if TEST_FB2_HIGH_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    TEST_FB2_HIGH_TEXT.frameNStart = frameN  # exact frame index
                    TEST_FB2_HIGH_TEXT.tStart = t  # local t and not account for scr refresh
                    TEST_FB2_HIGH_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(TEST_FB2_HIGH_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB2_HIGH_TEXT.started')
                    TEST_FB2_HIGH_TEXT.setAutoDraw(True)
                if TEST_FB2_HIGH_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > TEST_FB2_HIGH_TEXT.tStartRefresh + feedback_time-frameTolerance:
                        # keep track of stop time/frame for later
                        TEST_FB2_HIGH_TEXT.tStop = t  # not accounting for scr refresh
                        TEST_FB2_HIGH_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'TEST_FB2_HIGH_TEXT.stopped')
                        TEST_FB2_HIGH_TEXT.setAutoDraw(False)
                
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    routineForceEnded = True
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in TestConfFbComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            # --- Ending Routine "TestConfFb" ---
            for thisComponent in TestConfFbComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # the Routine "TestConfFb" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            thisExp.nextEntry()
            
        # completed is_test_resp repeats of 'TestSkip'
        
        # get names of stimulus parameters
        if TestSkip.trialList in ([], [None], None):
            params = []
        else:
            params = TestSkip.trialList[0].keys()
        # save data for this loop
        TestSkip.saveAsExcel(filename + '.xlsx', sheetName='TestSkip',
            stimOut=params,
            dataOut=['n','all_mean','all_std', 'all_raw'])
        TestSkip.saveAsText(filename + 'TestSkip.csv', delim=',',
            stimOut=params,
            dataOut=['n','all_mean','all_std', 'all_raw'])
        
        # --- Prepare to start Routine "TestReward" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        # Run 'Begin Routine' code from TEST_REWARD_CODE
        # Get the presented time in test-reward
        test_reward_pres_time = onset.getTime()
        TEST_FB3_FRAME.setPos(test_choice_pos)
        TEST_FB3_IMG1.setImage('img/' + test_img1)
        TEST_FB3_IMG2.setImage('img/' + test_img2)
        TEST_REWARD.setText(test_reward_text)
        # keep track of which components have finished
        TestRewardComponents = [TEST_FB3_FRAME, TEST_FB3_IMG1, TEST_FB3_IMG2, TEST_REWARD]
        for thisComponent in TestRewardComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        # --- Run Routine "TestReward" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            # *TEST_FB3_FRAME* updates
            if TEST_FB3_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FB3_FRAME.frameNStart = frameN  # exact frame index
                TEST_FB3_FRAME.tStart = t  # local t and not account for scr refresh
                TEST_FB3_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FB3_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FB3_FRAME.started')
                TEST_FB3_FRAME.setAutoDraw(True)
            if TEST_FB3_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FB3_FRAME.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FB3_FRAME.tStop = t  # not accounting for scr refresh
                    TEST_FB3_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB3_FRAME.stopped')
                    TEST_FB3_FRAME.setAutoDraw(False)
            
            # *TEST_FB3_IMG1* updates
            if TEST_FB3_IMG1.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FB3_IMG1.frameNStart = frameN  # exact frame index
                TEST_FB3_IMG1.tStart = t  # local t and not account for scr refresh
                TEST_FB3_IMG1.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FB3_IMG1, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FB3_IMG1.started')
                TEST_FB3_IMG1.setAutoDraw(True)
            if TEST_FB3_IMG1.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FB3_IMG1.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FB3_IMG1.tStop = t  # not accounting for scr refresh
                    TEST_FB3_IMG1.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB3_IMG1.stopped')
                    TEST_FB3_IMG1.setAutoDraw(False)
            
            # *TEST_FB3_IMG2* updates
            if TEST_FB3_IMG2.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                TEST_FB3_IMG2.frameNStart = frameN  # exact frame index
                TEST_FB3_IMG2.tStart = t  # local t and not account for scr refresh
                TEST_FB3_IMG2.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_FB3_IMG2, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_FB3_IMG2.started')
                TEST_FB3_IMG2.setAutoDraw(True)
            if TEST_FB3_IMG2.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_FB3_IMG2.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_FB3_IMG2.tStop = t  # not accounting for scr refresh
                    TEST_FB3_IMG2.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_FB3_IMG2.stopped')
                    TEST_FB3_IMG2.setAutoDraw(False)
            
            # *TEST_REWARD* updates
            if TEST_REWARD.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                TEST_REWARD.frameNStart = frameN  # exact frame index
                TEST_REWARD.tStart = t  # local t and not account for scr refresh
                TEST_REWARD.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(TEST_REWARD, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'TEST_REWARD.started')
                TEST_REWARD.setAutoDraw(True)
            if TEST_REWARD.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > TEST_REWARD.tStartRefresh + reward_time-frameTolerance:
                    # keep track of stop time/frame for later
                    TEST_REWARD.tStop = t  # not accounting for scr refresh
                    TEST_REWARD.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'TEST_REWARD.stopped')
                    TEST_REWARD.setAutoDraw(False)
            
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in TestRewardComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        # --- Ending Routine "TestReward" ---
        for thisComponent in TestRewardComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # Run 'End Routine' code from TEST_REWARD_CODE
        # Get the responsed time in test-reward
        test_reward_resp_time = onset.getTime()
        
        test_choice_time = test_choice_resp_time - test_choice_pres_time
        test_conf_time = test_conf_resp_time - test_conf_pres_time
        
        # Output test data
        test_data_table[test_serial_num] = [game_serial_num, test_serial_num+1, block_num, 
                                           test_seq_pattern, test_loc_pattern, 
                                           test_loc1, test_loc2, test_scale, 
                                           test_pt, test_conf_resp, test_reward,
                                           np.round(test_choice_time,3), np.round(test_conf_time,3),
                                           test_img.index(test_img1)+1, test_img.index(test_img2)+1, 
                                           test_choice_resp, test_correct, test_choice_loc, 
                                           np.round(test_choice_pres_time,3), np.round(test_choice_resp_time,3),
                                           np.round(test_conf_pres_time, 3), np.round(test_conf_resp_time, 3), 
                                           np.round(test_reward_pres_time, 3), np.round(test_reward_resp_time, 3)]
        
        # Save matlab file
        io.savemat(out_name + '.mat', {'out_mat':out_dict})
        
        # Save excel file
        out_book = load_workbook(filename=out_xlsx)
        test_sheet_name = out_book.sheetnames[2]
        test_sheet = out_book[test_sheet_name]
        for i in range(1, len(test_data_table[test_serial_num])+1):
            test_sheet.cell(test_serial_num+1, i, value = test_data_table[test_serial_num][i-1])
        out_book.save(out_xlsx)
        
        # Update test-count
        test_serial_num += 1
        game_serial_num += 1
        # the Routine "TestReward" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        thisExp.nextEntry()
        
    # completed 1.0 repeats of 'TestSequence'
    
    # get names of stimulus parameters
    if TestSequence.trialList in ([], [None], None):
        params = []
    else:
        params = TestSequence.trialList[0].keys()
    # save data for this loop
    TestSequence.saveAsExcel(filename + '.xlsx', sheetName='TestSequence',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    TestSequence.saveAsText(filename + 'TestSequence.csv', delim=',',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    
    # --- Prepare to start Routine "BlockRest" ---
    continueRoutine = True
    routineForceEnded = False
    # update component parameters for each repeat
    # Run 'Begin Routine' code from BLOCK_REST_CODE
    message = '1 Block is over.\nTake a break.\nPress ‘space’ to start next Block.' if block_num == 1 else\
              '2 Block is over.\nTake a break.\nPress ‘space’ to start next Block.' if block_num == 2 else\
              'Game is over.\nPress ‘space’ to exit.'
    BLOCK_REST_TEXT.setText(message)
    BLOCK_REST_SKIP.keys = []
    BLOCK_REST_SKIP.rt = []
    _BLOCK_REST_SKIP_allKeys = []
    # keep track of which components have finished
    BlockRestComponents = [BLOCK_REST_TEXT, BLOCK_REST_SKIP]
    for thisComponent in BlockRestComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    # --- Run Routine "BlockRest" ---
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *BLOCK_REST_TEXT* updates
        if BLOCK_REST_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            BLOCK_REST_TEXT.frameNStart = frameN  # exact frame index
            BLOCK_REST_TEXT.tStart = t  # local t and not account for scr refresh
            BLOCK_REST_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(BLOCK_REST_TEXT, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'BLOCK_REST_TEXT.started')
            BLOCK_REST_TEXT.setAutoDraw(True)
        
        # *BLOCK_REST_SKIP* updates
        waitOnFlip = False
        if BLOCK_REST_SKIP.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            BLOCK_REST_SKIP.frameNStart = frameN  # exact frame index
            BLOCK_REST_SKIP.tStart = t  # local t and not account for scr refresh
            BLOCK_REST_SKIP.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(BLOCK_REST_SKIP, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'BLOCK_REST_SKIP.started')
            BLOCK_REST_SKIP.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(BLOCK_REST_SKIP.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(BLOCK_REST_SKIP.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if BLOCK_REST_SKIP.status == STARTED and not waitOnFlip:
            theseKeys = BLOCK_REST_SKIP.getKeys(keyList=['space'], waitRelease=False)
            _BLOCK_REST_SKIP_allKeys.extend(theseKeys)
            if len(_BLOCK_REST_SKIP_allKeys):
                BLOCK_REST_SKIP.keys = _BLOCK_REST_SKIP_allKeys[-1].name  # just the last key pressed
                BLOCK_REST_SKIP.rt = _BLOCK_REST_SKIP_allKeys[-1].rt
                # a response ends the routine
                continueRoutine = False
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in BlockRestComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    # --- Ending Routine "BlockRest" ---
    for thisComponent in BlockRestComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # Run 'End Routine' code from BLOCK_REST_CODE
    block_num += 1
    core.wait(3)
    # check responses
    if BLOCK_REST_SKIP.keys in ['', [], None]:  # No response was made
        BLOCK_REST_SKIP.keys = None
    BlockLoop.addData('BLOCK_REST_SKIP.keys',BLOCK_REST_SKIP.keys)
    if BLOCK_REST_SKIP.keys != None:  # we had a response
        BlockLoop.addData('BLOCK_REST_SKIP.rt', BLOCK_REST_SKIP.rt)
    # the Routine "BlockRest" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    thisExp.nextEntry()
    
# completed 1.0 repeats of 'BlockLoop'

# get names of stimulus parameters
if BlockLoop.trialList in ([], [None], None):
    params = []
else:
    params = BlockLoop.trialList[0].keys()
# save data for this loop
BlockLoop.saveAsExcel(filename + '.xlsx', sheetName='BlockLoop',
    stimOut=params,
    dataOut=['n','all_mean','all_std', 'all_raw'])
BlockLoop.saveAsText(filename + 'BlockLoop.csv', delim=',',
    stimOut=params,
    dataOut=['n','all_mean','all_std', 'all_raw'])

# --- Prepare to start Routine "Appreciation" ---
continueRoutine = True
routineForceEnded = False
# update component parameters for each repeat
# keep track of which components have finished
AppreciationComponents = [APPRECIATION]
for thisComponent in AppreciationComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
frameN = -1

# --- Run Routine "Appreciation" ---
while continueRoutine and routineTimer.getTime() < 5.0:
    # get current time
    t = routineTimer.getTime()
    tThisFlip = win.getFutureFlipTime(clock=routineTimer)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *APPRECIATION* updates
    if APPRECIATION.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        APPRECIATION.frameNStart = frameN  # exact frame index
        APPRECIATION.tStart = t  # local t and not account for scr refresh
        APPRECIATION.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(APPRECIATION, 'tStartRefresh')  # time at next scr refresh
        # add timestamp to datafile
        thisExp.timestampOnFlip(win, 'APPRECIATION.started')
        APPRECIATION.setAutoDraw(True)
    if APPRECIATION.status == STARTED:
        # is it time to stop? (based on global clock, using actual start)
        if tThisFlipGlobal > APPRECIATION.tStartRefresh + 5-frameTolerance:
            # keep track of stop time/frame for later
            APPRECIATION.tStop = t  # not accounting for scr refresh
            APPRECIATION.frameNStop = frameN  # exact frame index
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'APPRECIATION.stopped')
            APPRECIATION.setAutoDraw(False)
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        routineForceEnded = True
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in AppreciationComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()

# --- Ending Routine "Appreciation" ---
for thisComponent in AppreciationComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
# using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
if routineForceEnded:
    routineTimer.reset()
else:
    routineTimer.addTime(-5.000000)

# --- End experiment ---
# Flip one final time so any remaining win.callOnFlip() 
# and win.timeOnFlip() tasks get executed before quitting
win.flip()

# these shouldn't be strictly necessary (should auto-save)
thisExp.saveAsWideText(filename+'.csv', delim='auto')
thisExp.saveAsPickle(filename)
# make sure everything is closed down
if eyetracker:
    eyetracker.setConnectionState(False)
thisExp.abort()  # or data files will save again on exit
win.close()
core.quit()
