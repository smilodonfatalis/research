#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Import modules
import random 

import numpy as np
from openpyxl import load_workbook
import pandas as pd

import template as tp

# NOTE: success seeds are '5, 8, 17, 20'
SEED = 5
GAME_PAT = 1
BLOCK_NUM = 4
SLF_TRIAL_NUM = 9
TEST_TRIAL_NUM = 30
MEAN = [30, 40, 50]
SD = 7

ALPHA = 0.216
BETA = 0.141

SIM_NUM = 2**10

# Fix seed
random.seed(SEED)
np.random.seed(seed=SEED)


def conf_pats_generator(trial_num):
    conf_pats = [1]*(BLOCK_NUM*trial_num//2) + [2]*(BLOCK_NUM*trial_num//2)
    random.shuffle(conf_pats)
    conf_pats = np.array(conf_pats)
    return conf_pats


def slf_write_excel(conf_pats, game_pat):

    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        condition_file_path = f'{condition_dir_path}game{game_pat}/slf{game_pat}{block+1}.xlsx'
        wb = load_workbook(condition_file_path)
        ws = wb['condition']
        for trial in range(SLF_TRIAL_NUM):
            ws[f'D{trial+2}'] = conf_pats[BLOCK_NUM*block+trial]
        wb.save(condition_file_path)

def test_write_excel(conf_pats, game_pat):

    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        condition_file_path = f'{condition_dir_path}game{game_pat}/test{game_pat}{block+1}.xlsx'
        wb = load_workbook(condition_file_path)
        ws = wb['condition']
        for trial in range(TEST_TRIAL_NUM):
            ws[f'D{trial+2}'] = conf_pats[BLOCK_NUM*block+trial]
        wb.save(condition_file_path)


def main():
    game_pat = GAME_PAT
    slf_conf_pats = conf_pats_generator(SLF_TRIAL_NUM)
    test_conf_pats = conf_pats_generator(TEST_TRIAL_NUM)
    slf_write_excel(slf_conf_pats, game_pat)
    test_write_excel(test_conf_pats, game_pat)


if __name__ == '__main__':
    main()
