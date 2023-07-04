#!/usr/bin/env python
# -*- coding: utf-8 -*-

import psychopy
psychopy.useVersion('2022.2.4')


# --- Import packages ---
from psychopy import locale_setup
from psychopy import prefs
from psychopy import sound, gui, visual, core, data, event, logging, clock, colors, layout
from psychopy.constants import (NOT_STARTED, STARTED, PLAYING, PAUSED,
                                STOPPED, FINISHED, PRESSED, RELEASED, FOREVER)
import psychopy.iohub as io
from psychopy.hardware import keyboard
print('===========================================')

import datetime
import os  # handy system and path functions
import sys  # to get file system encoding

import numpy as np  # whole numpy lib is available, prepend 'np.'
from numpy import (sin, cos, tan, log, log10, pi, average,
                   sqrt, std, deg2rad, rad2deg, linspace, asarray)
from numpy.random import random, randint, normal, shuffle, choice as randchoice
from openpyxl import Workbook, load_workbook
import pandas as pd



# Global variables
GAME_NUM = 3
WINDOW_SIZE = [960, 540]
isFullScr = True
# Time
TASK_TEST_TIME = 0.3
IMG_TIME = TASK_TEST_TIME
CONF_TIME = TASK_TEST_TIME
REWARD_TIME = TASK_TEST_TIME
FEEDBACK_TIME = 1.0
INSTR_TIME = TASK_TEST_TIME
FIXATION_TIME = 0.5
STANDBY_TIME = 3.0
APPRECIATION_TIME = 5.0

IMG_TIME = 5.0
CONF_TIME = 3.0
REWARD_TIME = 2.0
FEEDBACK_TIME = 1.0
INSTR_TIME = 2.0
FIXATION_TIME = 0.5
STANDBY_TIME = 3.0
APPRECIATION_TIME = 5.0

# Size and Position
IMG_SIZE = (200, 200)
IMG_POS_LEFT = (-200, 0)
IMG_POS_RIGHT = (200, 0)
IMG_POS_OUT = (10000, 0)
IMG_COLOR = 'white'

CONF_SIZE = (200, 200)
CONF_POS_LEFT = (-200, 0)
CONF_POS_RIGHT = (200, 0)
CONF_POS_OUT = (10000, 0)
CONF_BOX_COLOR = 'white'
CONF_TEXT_LOW = 'L'
CONF_TEXT_HIGH = 'H'
CONF_TEXT_COLOR = 'black'

FRAME_SIZE = (220, 220)

# Color
SLF_COLOR = 'lime'
OBS_COLOR = 'cyan'
TEST_COLOR = 'orange'

# Number
BLOCK_NUM = 4
SLF_TRIAL_NUM = 9
OBS_TRIAL_NUM = 9
TEST_TRIAL_NUM = 30

# Instruction
TEXT_FONT = 'Arial'
TEXT_POS = (0, 0)
TEXT_SIZE = 48
TEXT_COLOR = 'white'

# FIXATION = '+'
# FIXATION_SIZE = 60
FIXATION_SIZE = (30, 30)
FIXATION_POS = (0, 0)
FIXATION_COLOR = 'white'
FIXATION_WIDTH = 0.1

NOT_DISCLOSED = '?'
FAILURE = 'F'

UNITS = 'pix'
APPRECIATION = 'Thank you for your cooperation in the experiment!'

MEAN_LIST = [30, 40, 50]
SD = 10
SLF_SEQ_TYPE = 'slf'
OBS_SEQ_TYPE = 'obs'
TEST_SEQ_TYPE = 'test'

POS_LEFT = 'left'
POS_RIGHT = 'right'
POS_SAME = 'same'
POS_OUT = 'none'

CONF_LOW = 'low'
CONF_HIGH = 'high'
CONF_NONE = 'none'

def get_idx(mean):
    if mean==30:
        idx = 0
    elif mean==40:
        idx = 1
    elif mean==50:
        idx = 2
    return idx


# Ensure that relative paths start from the same directory as this script
_thisDir = os.path.dirname(os.path.abspath(__file__))
os.chdir(_thisDir)
# Store info about the experiment session
psychopyVersion = '2022.2.4'
expName = 'Pattern_3'  # from the Builder filename that created this script
expInfo = {
    'participant': 'test',
}
# --- Show participant info dialog --
dlg = gui.DlgFromDict(dictionary=expInfo, sortKeys=False, title=expName)
if dlg.OK == False:
    core.quit()  # user pressed cancel
expInfo['date'] = data.getDateStr()  # add a simple timestamp
expInfo['expName'] = expName
expInfo['psychopyVersion'] = psychopyVersion

# Data file name stem = absolute path + name; later add .psyexp, .csv, .log, etc
filename = _thisDir + os.sep + u'data/%s_%s_%s' % (expInfo['participant'], expName, expInfo['date'])

# An ExperimentHandler isn't essential but helps with data saving
thisExp = data.ExperimentHandler(name=expName, version='',
    extraInfo=expInfo, runtimeInfo=None,
    originPath=f'/Users/strix_uralensis/Documents/Experiment/research/game{GAME_NUM}.py',
    savePickle=True, saveWideText=True,
    dataFileName=filename)
logging.console.setLevel(logging.WARNING)  # this outputs to the screen, not a file

endExpNow = False  # flag for 'escape' or other condition => quit the exp
frameTolerance = 0.001  # how close to onset before 'same' frame

# Start Code - component code to be run after the window creation

# --- Setup the Window ---
win = visual.Window(
    size=WINDOW_SIZE, fullscr=isFullScr, screen=0, 
    winType='pyglet', allowStencil=False,
    monitor='screen1', color=[-1.000,-1.000,-1.000], colorSpace='rgb',
    blendMode='avg', useFBO=True, 
    units=UNITS)
win.mouseVisible = False
# store frame rate of monitor if we can measure it
expInfo['frameRate'] = win.getActualFrameRate()
if expInfo['frameRate'] != None:
    frameDur = 1.0 / round(expInfo['frameRate'])
else:
    frameDur = 1.0 / 60.0  # could not measure, so guess
# --- Setup input devices ---
ioConfig = {}

# Setup iohub keyboard
ioConfig['Keyboard'] = dict(use_keymap='psychopy')

ioSession = '1'
if 'session' in expInfo:
    ioSession = str(expInfo['session'])
ioServer = io.launchHubServer(window=win, **ioConfig)
eyetracker = None

# create a default keyboard (e.g. to check for escape)
defaultKeyboard = keyboard.Keyboard(backend='iohub')

"""============================================================="""
# --- Initialize components for Routine "Init" ---
# Run 'Begin Experiment' code from INIT_CODE


# Initial settings
win.mouseVisible = False

#IMG_TIME = .3
#CONF_TIME   = .3
#FEEDBACK_TIME   = .3
#REWARD_TIME     = .3

left_key   = 'left'
center_key = 'down'
right_key  = 'right'

# Information of the experiment dialog 
participant = expInfo["participant"]

# Number of trials in each stage
slf_trial_num = BLOCK_NUM * SLF_TRIAL_NUM
obs_trial_num = BLOCK_NUM * OBS_TRIAL_NUM
test_trial_num = BLOCK_NUM * TEST_TRIAL_NUM

# Counters
block_cnt = 1
through_game_cnt = 1
slf_trial_cnt = 1
obs_trial_cnt = 1
test_trial_cnt = 1

# Conditions
slf_condition = 0
obs_condition = 0

# Standard Deviation
slf_sd = SD
obs_sd = SD
test_sd = SD

# Images
slf_img = ['images/slf31.png', 'images/slf32.png', 'images/slf33.png']
obs_img = ['images/obs31.png', 'images/obs32.png', 'images/obs33.png']
test_img = [slf_img, obs_img]
class TestImg:
    def __init__(self, slf_img, obs_img):
        self.slf = slf_img
        self.obs = obs_img
test_img = TestImg(slf_img, obs_img)

# Create files to store data
get_current_time = datetime.datetime.now() # Get time in the form of "yyyy-mm-dd hh:mm:ss"
out_name = f'subj_{participant}_game{GAME_NUM}_{get_current_time:%y%m%d%H%M}' 
            # participant_yyyymmddというファイル名

# Define sheet names and data table
slf_sheet = f'slf{GAME_NUM}'
obs_sheet = f'obs{GAME_NUM}'
test_sheet = f'test{GAME_NUM}'

slf_data_table =  []
obs_data_table =  []
test_data_table = []

# Create matlab file(dict)
out_dict = {'slf1':slf_data_table, 'obs1':obs_data_table, 'test1':test_data_table}


# Create excel file
cols = ['block_num', 'trial_num', 'seq_type', 'seq_trial_num', 
        'condition', 'pair_pat', 'order_pat', 
        'mean_left', 'mean_right', 'sd', 
        'idx_left', 'idx_right',
        'pos_correct', 'mean_correct', 'idx_correct', 
        'pos_chosen', 'mean_chosen', 'idx_chosen',  
        'conf_pat', 'conf', 'acc', 'reward',
        'img_time', 'conf_time', 
        'img_pres_time', 'img_resp_time',
        'conf_pres_time', 'conf_resp_time',
        'reward_pres_time', 'reward_resp_time']

out_xlsx = out_name + ".xlsx"
out_book = Workbook()
#out_book = load_workbook(filename=out_xlsx)
out_book.create_sheet(index=0, title=slf_sheet)
out_book.create_sheet(index=1, title=obs_sheet)
out_book.create_sheet(index=2, title=test_sheet)

for i in range(3):
    sheet_name = out_book.sheetnames[i]
    sheet = out_book[sheet_name]
    for j in range(1, len(cols)+1):
        sheet.cell(1, j, value=cols[j-1])
out_book.save(out_xlsx)


"""============================================================="""
STANDBY_TEXT = visual.TextStim(win=win, name='STANDBY_TEXT',
    text='Press ‘space’ to start',
    font=TEXT_FONT,
    pos=(0, 0), height=48, wrapWidth=None, ori=0, 
    color='white', colorSpace='rgb', opacity=1, 
    languageStyle='LTR',
    depth=-1.0);
STANDBY_RESP = keyboard.Keyboard()


"""============================================================="""
# --- Initialize components for Routine "SlfInstr" ---
SLF_INSTR = visual.TextStim(
    win=win, 
    name='SLF_INSTR',
    text='Self',
    font=TEXT_FONT,
    pos=TEXT_POS, 
    height=TEXT_SIZE, 
    color=SLF_COLOR, 
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1, languageStyle='LTR',
    depth=0.0)


"""============================================================="""
# --- Initialize components for Routine "SlfChoice" ---
# SLF_IMG_FIX = visual.TextStim(
#     win=win, 
#     name='SLF_IMG_FIX',
#     text=FIXATION,
#     font=TEXT_FONT,
#     pos=TEXT_POS, 
#     height=FIXATION_SIZE, 
#     color=TEXT_COLOR, 
#     wrapWidth=None, ori=0, colorSpace='rgb', opacity=1, languageStyle='LTR',
#     depth=-1.0)
SLF_IMG_FIX = visual.ShapeStim(
    win=win, 
    name='SLF_IMG_FIX',
    vertices='cross',
    size=FIXATION_SIZE, 
    pos=FIXATION_POS,
    lineColor=FIXATION_COLOR, 
    fillColor=FIXATION_COLOR,
    anchor='center', ori=0, lineWidth=FIXATION_WIDTH, colorSpace='rgb', opacity=1, interpolate=True,
    depth=0.0)

SLF_IMG_LEFT = visual.ImageStim(
    win=win,
    name='SLF_IMG_LEFT', 
    pos=IMG_POS_LEFT, 
    size=IMG_SIZE,
    color=IMG_COLOR, 
    image='sin', 
    mask=None, anchor='center', ori=0, colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False, texRes=128, interpolate=True, 
    depth=-2.0)

SLF_IMG_RIGHT = visual.ImageStim(
    win=win,
    name='SLF_IMG_RIGHT', 
    pos=IMG_POS_RIGHT, 
    size=IMG_SIZE,
    color=IMG_COLOR,
    image='sin', 
    mask=None, anchor='center', ori=0,  colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False, texRes=128, interpolate=True, 
    depth=-3.0)

SLF_IMG_RESP = keyboard.Keyboard()


"""============================================================="""
# --- Initialize components for Routine "SlfChoiceFb" ---
# SLF_IMGFB_FIX = visual.TextStim(
#     win=win, 
#     name='SLF_IMGFB_FIX',
#     text=FIXATION,
#     font=TEXT_FONT,
#     pos=TEXT_POS, 
#     height=FIXATION_SIZE, 
#     color=TEXT_COLOR, 
#     wrapWidth=None, ori=0, colorSpace='rgb', opacity=None, languageStyle='LTR',
#     depth=0.0)
SLF_IMGFB_FIX = visual.ShapeStim(
    win=win, 
    name='SLF_IMGFB_FIX',
    vertices='cross',
    size=FIXATION_SIZE, 
    pos=FIXATION_POS,
    lineColor=FIXATION_COLOR, 
    fillColor=FIXATION_COLOR,
    anchor='center', ori=0, lineWidth=FIXATION_WIDTH, colorSpace='rgb', opacity=1, interpolate=True,
    depth=0.0)

SLF_IMGFB_FRAME = visual.Rect(
    win=win, 
    name='SLF_IMGFB_FRAME',
    width=FRAME_SIZE[0], 
    height=FRAME_SIZE[1],
    pos=[0,0],
    lineColor=SLF_COLOR, 
    fillColor=SLF_COLOR,
    anchor='center', ori=0, lineWidth=20, colorSpace='rgb', opacity=1, interpolate=True,
    depth=-1.0)

SLF_IMGFB_IMG_LEFT = visual.ImageStim(
    win=win,
    name='SLF_IMGFB_IMG_LEFT', 
    pos=IMG_POS_LEFT, 
    size=IMG_SIZE,
    color=IMG_COLOR, 
    image='sin', 
    mask=None, anchor='center', ori=0, colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False, texRes=128, interpolate=True, 
    depth=-2.0)

SLF_IMGFB_IMG_RIGHT = visual.ImageStim(
    win=win,
    name='SLF_IMGFB_IMG_RIGHT', 
    pos=IMG_POS_RIGHT, 
    size=IMG_SIZE,
    color=IMG_COLOR,
    image='sin', 
    mask=None, anchor='center', ori=0,  colorSpace='rgb', opacity=1,
    flipHoriz=False, flipVert=False, texRes=128, interpolate=True, 
    depth=-3.0)


"""============================================================="""
# --- Initialize components for Routine "SlfConf" ---
SLF_CONF_FIX = visual.ShapeStim(
    win=win, 
    name='SLF_CONF_FIX',
    vertices='cross',
    size=FIXATION_SIZE, 
    pos=FIXATION_POS,
    lineColor=FIXATION_COLOR, 
    fillColor=FIXATION_COLOR,
    anchor='center', ori=0, lineWidth=FIXATION_WIDTH, colorSpace='rgb', opacity=1, interpolate=True,
    depth=0.0)

SLF_CONF_LOW_BOX = visual.Rect(
    win=win, 
    name='SLF_CONF_LOW_BOX',
    width=CONF_SIZE[0], 
    height=CONF_SIZE[1],
    pos=CONF_POS_LEFT, 
    lineColor=CONF_BOX_COLOR, 
    fillColor=CONF_BOX_COLOR,
    anchor='center', ori=0, lineWidth=1.0, colorSpace='rgb', opacity=1.0, interpolate=True,
    depth=-1.0) 

SLF_CONF_LOW_TEXT = visual.TextStim(
    win=win, 
    name='SLF_CONF_LOW_TEXT',
    text=CONF_TEXT_LOW,
    font=TEXT_FONT,
    pos=CONF_POS_LEFT, 
    height=TEXT_SIZE, 
    color=CONF_TEXT_COLOR, 
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1.0, languageStyle='LTR',
    depth=-2.0)

SLF_CONF_HIGH_BOX = visual.Rect(
    win=win, 
    name='SLF_CONF_HIGH_BOX',
    width=CONF_SIZE[0], 
    height=CONF_SIZE[1],
    pos=CONF_POS_RIGHT, 
    lineColor=CONF_BOX_COLOR, 
    fillColor=CONF_BOX_COLOR,
    anchor='center', ori=0, lineWidth=1.0, colorSpace='rgb', opacity=1.0, interpolate=True,
    depth=-3.0)

SLF_CONF_HIGH_TEXT = visual.TextStim(
    win=win, 
    name='SLF_CONF_HIGH_TEXT',
    text=CONF_TEXT_HIGH,
    font=TEXT_FONT,
    pos=CONF_POS_RIGHT, 
    height=TEXT_SIZE, 
    color=CONF_TEXT_COLOR, 
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1.0, languageStyle='LTR',
    depth=-4.0)

SLF_CONF_RESP = keyboard.Keyboard()


"""============================================================="""
# --- Initialize components for Routine "SlfConfFb" ---
SLF_CONFFB_FIX = visual.ShapeStim(
    win=win, 
    name='SLF_CONFFB_FIX',
    vertices='cross',
    size=FIXATION_SIZE, 
    pos=FIXATION_POS,
    lineColor=FIXATION_COLOR, 
    fillColor=FIXATION_COLOR,
    anchor='center', ori=0, lineWidth=FIXATION_WIDTH, colorSpace='rgb', opacity=1, interpolate=True,
    depth=0.0)

SLF_CONFFB_FRAME = visual.Rect(
    win=win, 
    name='SLF_CONFFB_FRAME',
    width=FRAME_SIZE[0], 
    height=FRAME_SIZE[1],
    pos=[0,0], 
    lineColor=SLF_COLOR, 
    fillColor=SLF_COLOR, 
    anchor='center', ori=0, lineWidth=20, colorSpace='rgb', opacity=1, interpolate=True,
    depth=0.0)

SLF_CONFFB_LOW_BOX = visual.Rect(
    win=win, 
    name='SLF_CONFFB_LOW_BOX',
    width=CONF_SIZE[0], 
    height=CONF_SIZE[1],
    pos=CONF_POS_LEFT, 
    lineColor=CONF_BOX_COLOR, 
    fillColor=CONF_BOX_COLOR,
    anchor='center', ori=0, lineWidth=1.0, colorSpace='rgb', opacity=1.0, interpolate=True,
    depth=-1.0)

SLF_CONFFB_LOW_TEXT = visual.TextStim(
    win=win, name='SLF_CONFFB_LOW_TEXT',
    text=CONF_TEXT_LOW,
    font=TEXT_FONT,
    pos=CONF_POS_LEFT, 
    height=TEXT_SIZE, 
    color=CONF_TEXT_COLOR, 
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1.0, languageStyle='LTR',
    depth=-2.0)

SLF_CONFFB_HIGH_BOX = visual.Rect(
    win=win, 
    name='SLF_CONFFB_HIGH_BOX',
    width=CONF_SIZE[0], 
    height=CONF_SIZE[1],
    pos=CONF_POS_RIGHT, 
    lineColor=CONF_BOX_COLOR, 
    fillColor=CONF_BOX_COLOR,
    anchor='center', ori=0, lineWidth=1.0, colorSpace='rgb', opacity=1.0, interpolate=True,
    depth=-3.0)

SLF_CONFFB_HIGH_TEXT = visual.TextStim(
    win=win, name='SLF_CONFFB_HIGH_TEXT',
    text=CONF_TEXT_HIGH,
    font=TEXT_FONT,
    pos=CONF_POS_RIGHT, 
    height=TEXT_SIZE, 
    color=CONF_TEXT_COLOR, 
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1.0, languageStyle='LTR',
    depth=-4.0)


"""============================================================="""
# --- Initialize components for Routine "SlfReward" ---
SLF_REWARD_FRAME = visual.Rect(
    win=win, 
    name='SLF_REWARD_FRAME',
    width=FRAME_SIZE[0], 
    height=FRAME_SIZE[1],
    pos=[0,0],
    lineColor=SLF_COLOR, 
    fillColor=SLF_COLOR,
    anchor='center', ori=0, lineWidth=20.0, colorSpace='rgb', opacity=1.0, interpolate=True,
    depth=-1.0)

SLF_REWARD_IMG_LEFT = visual.ImageStim(
    win=win,
    name='SLF_REWARD_IMG_LEFT', 
    pos=IMG_POS_LEFT, 
    size=IMG_SIZE,
    color=IMG_COLOR, 
    image='sin', 
    mask=None, anchor='center', ori=0, colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False, texRes=128.0, interpolate=True, 
    depth=-2.0)

SLF_REWARD_IMG_RIGHT = visual.ImageStim(
    win=win,
    name='SLF_REWARD_IMG_RIGHT', 
    pos=IMG_POS_RIGHT,
    size=IMG_SIZE,
    color=IMG_COLOR,
    image='sin', 
    mask=None, anchor='center', ori=0, colorSpace='rgb', opacity=1.0,
    flipHoriz=False, flipVert=False, texRes=128.0, interpolate=True, 
    depth=-3.0)

SLF_REWARD_TEXT = visual.TextStim(
    win=win, 
    name='SLF_REWARD_TEXT',
    text='',
    font=TEXT_FONT,
    units=UNITS, 
    pos=TEXT_POS, 
    height=TEXT_SIZE, 
    color=TEXT_COLOR, 
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1.0, languageStyle='LTR',
    depth=-4.0)



"""============================================================="""
# --- Initialize components for Routine "BlockRest" ---
BLOCK_REST_TEXT = visual.TextStim(
    win=win, 
    name='BLOCK_REST_TEXT',
    text='',
    font=TEXT_FONT,
    pos=TEXT_POS, 
    height=TEXT_SIZE, 
    color=TEXT_COLOR, 
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1.0, languageStyle='LTR',
    depth=-1.0)

BLOCK_REST_SKIP = keyboard.Keyboard()


"""============================================================="""
# --- Initialize components for Routine "Appreciation" ---
APPRECIATION = visual.TextStim(
    win=win, 
    name='APPRECIATION',
    text=APPRECIATION,
    font=TEXT_FONT,
    pos=TEXT_POS,
    height=TEXT_SIZE,
    color=TEXT_COLOR,  
    wrapWidth=None, ori=0, colorSpace='rgb', opacity=1.0, languageStyle='LTR',
    depth=0.0)





















"""============================================================="""
# Create some handy timers
globalClock = core.Clock()  # to track the time since experiment started
routineTimer = core.Clock()  # to track time remaining of each (possibly non-slip) routine 


"""============================================================="""
# --- Prepare to start Routine "Init" ---
continueRoutine = True
routineForceEnded = False
# update component parameters for each repeat
STANDBY_RESP.keys = []
STANDBY_RESP.rt = []
_STANDBY_RESP_allKeys = []
# keep track of which components have finished
InitComponents = [STANDBY_TEXT, STANDBY_RESP]
for thisComponent in InitComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
frameN = -1


"""============================================================="""
# --- Run Routine "Init" ---
while continueRoutine:
    # get current time
    t = routineTimer.getTime()
    tThisFlip = win.getFutureFlipTime(clock=routineTimer)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    # *STANDBY_TEXT* updates
    if STANDBY_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        STANDBY_TEXT.frameNStart = frameN  # exact frame index
        STANDBY_TEXT.tStart = t  # local t and not account for scr refresh
        STANDBY_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(STANDBY_TEXT, 'tStartRefresh')  # time at next scr refresh
        # add timestamp to datafile
        thisExp.timestampOnFlip(win, 'STANDBY_TEXT.started')
        STANDBY_TEXT.setAutoDraw(True)
    
    # *STANDBY_RESP* updates
    waitOnFlip = False
    if STANDBY_RESP.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        STANDBY_RESP.frameNStart = frameN  # exact frame index
        STANDBY_RESP.tStart = t  # local t and not account for scr refresh
        STANDBY_RESP.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(STANDBY_RESP, 'tStartRefresh')  # time at next scr refresh
        # add timestamp to datafile
        thisExp.timestampOnFlip(win, 'STANDBY_RESP.started')
        STANDBY_RESP.status = STARTED
        # keyboard checking is just starting
        waitOnFlip = True
        win.callOnFlip(STANDBY_RESP.clock.reset)  # t=0 on next screen flip
        win.callOnFlip(STANDBY_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
    if STANDBY_RESP.status == STARTED and not waitOnFlip:
        theseKeys = STANDBY_RESP.getKeys(keyList=['space'], waitRelease=False)
        _STANDBY_RESP_allKeys.extend(theseKeys)
        if len(_STANDBY_RESP_allKeys):
            STANDBY_RESP.keys = _STANDBY_RESP_allKeys[-1].name  # just the last key pressed
            STANDBY_RESP.rt = _STANDBY_RESP_allKeys[-1].rt
            # a response ends the routine
            continueRoutine = False
    
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
    # if endExpNow:
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        routineForceEnded = True
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in InitComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()


"""============================================================="""
# --- Ending Routine "Init" ---
for thisComponent in InitComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
# Run 'End Routine' code from INIT_CODE
# Initialize clock
onset = core.MonotonicClock()
core.wait(STANDBY_TIME)
# check responses
if STANDBY_RESP.keys in ['', [], None]:  # No response was made
    STANDBY_RESP.keys = None
thisExp.addData('STANDBY_RESP.keys',STANDBY_RESP.keys)
if STANDBY_RESP.keys != None:  # we had a response
    thisExp.addData('STANDBY_RESP.rt', STANDBY_RESP.rt)
thisExp.nextEntry()
# the Routine "Init" was not non-slip safe, so reset the non-slip timer
routineTimer.reset()

# set up handler to look after randomisation of conditions etc
BlockLoop = data.TrialHandler(
    nReps=1.0, method='sequential', 
    extraInfo=expInfo, originPath=-1,
    trialList=data.importConditions(f'sequences/game{GAME_NUM}/condition3.xlsx'),
    seed=None, name='BlockLoop')
thisExp.addLoop(BlockLoop)  # add the loop to the experiment
thisBlockLoop = BlockLoop.trialList[0]  # so we can initialise stimuli with some values
# abbreviate parameter names if possible (e.g. rgb = thisBlockLoop.rgb)
if thisBlockLoop != None:
    for paramName in thisBlockLoop:
        exec('{} = thisBlockLoop[paramName]'.format(paramName))

for thisBlockLoop in BlockLoop:
    currentLoop = BlockLoop
    # abbreviate parameter names if possible (e.g. rgb = thisBlockLoop.rgb)
    if thisBlockLoop != None:
        for paramName in thisBlockLoop:
            exec('{} = thisBlockLoop[paramName]'.format(paramName))
    































    """============================================================="""
    # --- Prepare to start Routine "SlfInstr" ---
    continueRoutine = True
    routineForceEnded = False
    # update component parameters for each repeat
    # keep track of which components have finished
    SlfInstrComponents = [SLF_INSTR]
    for thisComponent in SlfInstrComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    """============================================================="""
    # --- Run Routine "SlfInstr" ---
    while continueRoutine and routineTimer.getTime() < INSTR_TIME:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        # *SLF_INSTR* updates
        if SLF_INSTR.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            SLF_INSTR.frameNStart = frameN  # exact frame index
            SLF_INSTR.tStart = t  # local t and not account for scr refresh
            SLF_INSTR.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(SLF_INSTR, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'SLF_INSTR.started')
            SLF_INSTR.setAutoDraw(True)
        if SLF_INSTR.status == STARTED:
            # is it time to stop? (based on global clock, using actual start)
            if tThisFlipGlobal > SLF_INSTR.tStartRefresh + INSTR_TIME-frameTolerance:
                # keep track of stop time/frame for later
                SLF_INSTR.tStop = t  # not accounting for scr refresh
                SLF_INSTR.frameNStop = frameN  # exact frame index
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_INSTR.stopped')
                SLF_INSTR.setAutoDraw(False)
        
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        # if endExpNow:
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in SlfInstrComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    
    """============================================================="""
    # --- Ending Routine "SlfInstr" ---
    for thisComponent in SlfInstrComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
    # using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
    if routineForceEnded:
        routineTimer.reset()
    else:
        routineTimer.addTime(-INSTR_TIME)
    
    """============================================================="""
    # set up handler to look after randomisation of conditions etc
    SlfSequence = data.TrialHandler(
        nReps=1.0, method='sequential', 
        extraInfo=expInfo, originPath=-1,
        trialList=data.importConditions(f'sequences/game{GAME_NUM}/' + slf_file),
        seed=None, name='SlfSequence')
    thisExp.addLoop(SlfSequence)  # add the loop to the experiment
    thisSlfSequence = SlfSequence.trialList[0]  # so we can initialise stimuli with some values
    # abbreviate parameter names if possible (e.g. rgb = thisSlfSequence.rgb)
    if thisSlfSequence != None:
        for paramName in thisSlfSequence:
            exec('{} = thisSlfSequence[paramName]'.format(paramName))
    
    """============================================================="""
    for thisSlfSequence in SlfSequence:
        currentLoop = SlfSequence
        # abbreviate parameter names if possible (e.g. rgb = thisSlfSequence.rgb)
        if thisSlfSequence != None:
            for paramName in thisSlfSequence:
                exec('{} = thisSlfSequence[paramName]'.format(paramName))
        
        
        """============================================================="""
        # --- Prepare to start Routine "SlfChoice" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        # Run 'Begin Routine' code from SLF_IMG_CODE
        slf_seq_type = SLF_SEQ_TYPE
        
        if slf_pair_pat==1:
            if slf_order_pat==1:
                slf_mean_left = MEAN_LIST[0]
                slf_mean_right = MEAN_LIST[1]
            else:
                slf_mean_left = MEAN_LIST[1]
                slf_mean_right = MEAN_LIST[0]
        if slf_pair_pat==2:
            if slf_order_pat==1:
                slf_mean_left = MEAN_LIST[1]
                slf_mean_right = MEAN_LIST[2]
            else:
                slf_mean_left = MEAN_LIST[2]
                slf_mean_right = MEAN_LIST[1]
        if slf_pair_pat==3:
            if slf_order_pat==1:
                slf_mean_left = MEAN_LIST[2]
                slf_mean_right = MEAN_LIST[0]
            else:
                slf_mean_left = MEAN_LIST[0]
                slf_mean_right = MEAN_LIST[2]
        
        if slf_mean_left > slf_mean_right:
            slf_pos_correct = POS_LEFT
            slf_mean_correct = slf_mean_left
            slf_idx_correct = get_idx(slf_mean_left)
        elif slf_mean_left < slf_mean_right:
            slf_pos_correct = POS_RIGHT
            slf_mean_correct = slf_mean_right
            slf_idx_correct = get_idx(slf_mean_right)

        slf_idx_left = get_idx(slf_mean_left)
        slf_idx_right = get_idx(slf_mean_right)
        slf_img_left = slf_img[slf_idx_left]
        slf_img_right = slf_img[slf_idx_right]
        
        # Get the presented time in slf-choice
        t_img_pres = onset.getTime() + FIXATION_TIME # Consider the fixation
        # SLF_IMG_LEFT.setPos(IMG_POS_LEFT)
        SLF_IMG_LEFT.setImage(slf_img_left)
        SLF_IMG_RIGHT.setImage(slf_img_right)
        SLF_IMG_RESP.keys = []
        SLF_IMG_RESP.rt = []
        _SLF_IMG_RESP_allKeys = []
        # keep track of which components have finished
        SlfChoiceComponents = [SLF_IMG_FIX, 
                               SLF_IMG_LEFT, 
                               SLF_IMG_RIGHT, 
                               SLF_IMG_RESP]
        
        for thisComponent in SlfChoiceComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        
        """============================================================="""
        # --- Run Routine "SlfChoice" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            """============================================================="""
            # *SLF_IMG_FIX* updates
            if SLF_IMG_FIX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMG_FIX.frameNStart = frameN  # exact frame index
                SLF_IMG_FIX.tStart = t  # local t and not account for scr refresh
                SLF_IMG_FIX.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMG_FIX, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMG_FIX.started')
                SLF_IMG_FIX.setAutoDraw(True)
            if SLF_IMG_FIX.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMG_FIX.tStartRefresh + IMG_TIME+FIXATION_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMG_FIX.tStop = t  # not accounting for scr refresh
                    SLF_IMG_FIX.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMG_FIX.stopped')
                    SLF_IMG_FIX.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_IMG_LEFT* updates
            if SLF_IMG_LEFT.status == NOT_STARTED and tThisFlip >= FIXATION_TIME-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMG_LEFT.frameNStart = frameN  # exact frame index
                SLF_IMG_LEFT.tStart = t  # local t and not account for scr refresh
                SLF_IMG_LEFT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMG_LEFT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMG_LEFT.started')
                SLF_IMG_LEFT.setAutoDraw(True)
            if SLF_IMG_LEFT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMG_LEFT.tStartRefresh + IMG_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMG_LEFT.tStop = t  # not accounting for scr refresh
                    SLF_IMG_LEFT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMG_LEFT.stopped')
                    SLF_IMG_LEFT.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_IMG_RIGHT* updates
            if SLF_IMG_RIGHT.status == NOT_STARTED and tThisFlip >= FIXATION_TIME-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMG_RIGHT.frameNStart = frameN  # exact frame index
                SLF_IMG_RIGHT.tStart = t  # local t and not account for scr refresh
                SLF_IMG_RIGHT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMG_RIGHT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMG_RIGHT.started')
                SLF_IMG_RIGHT.setAutoDraw(True)
            if SLF_IMG_RIGHT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMG_RIGHT.tStartRefresh + IMG_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMG_RIGHT.tStop = t  # not accounting for scr refresh
                    SLF_IMG_RIGHT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMG_RIGHT.stopped')
                    SLF_IMG_RIGHT.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_IMG_RESP* updates
            waitOnFlip = False
            if SLF_IMG_RESP.status == NOT_STARTED and tThisFlip >= FIXATION_TIME-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMG_RESP.frameNStart = frameN  # exact frame index
                SLF_IMG_RESP.tStart = t  # local t and not account for scr refresh
                SLF_IMG_RESP.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMG_RESP, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMG_RESP.started')
                SLF_IMG_RESP.status = STARTED
                # keyboard checking is just starting
                waitOnFlip = True
                win.callOnFlip(SLF_IMG_RESP.clock.reset)  # t=0 on next screen flip
                win.callOnFlip(SLF_IMG_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
            if SLF_IMG_RESP.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMG_RESP.tStartRefresh + IMG_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMG_RESP.tStop = t  # not accounting for scr refresh
                    SLF_IMG_RESP.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMG_RESP.stopped')
                    SLF_IMG_RESP.status = FINISHED
            if SLF_IMG_RESP.status == STARTED and not waitOnFlip:
                theseKeys = SLF_IMG_RESP.getKeys(keyList=[left_key,right_key], waitRelease=False)
                _SLF_IMG_RESP_allKeys.extend(theseKeys)
                if len(_SLF_IMG_RESP_allKeys):
                    SLF_IMG_RESP.keys = _SLF_IMG_RESP_allKeys[-1].name  # just the last key pressed
                    SLF_IMG_RESP.rt = _SLF_IMG_RESP_allKeys[-1].rt
                    # was this correct?
                    if (SLF_IMG_RESP.keys == str('slfCorrect')) or (SLF_IMG_RESP.keys == 'slfCorrect'):
                        SLF_IMG_RESP.corr = 1
                    else:
                        SLF_IMG_RESP.corr = 0
                    # a response ends the routine
                    continueRoutine = False
            
            """============================================================="""
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            # if endExpNow:
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in SlfChoiceComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        
        """============================================================="""
        # --- Ending Routine "SlfChoice" ---
        for thisComponent in SlfChoiceComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # Run 'End Routine' code from SLF_IMG_CODE
        # Get the responsed time in slf-choice
        t_img_resp = onset.getTime()
        
        """============================================================="""
        # Initialize in case skipping slf-conf
        slf_conf = 0
        t_conf_pres = 0
        t_conf_resp = 0
        is_slf_resp = 0
        
        # Receive choice responses and determine reward and feedback parameters
        if SLF_IMG_RESP.keys == left_key: 
            slf_img_fb_pos = IMG_POS_LEFT
            slf_pos_chosen = POS_LEFT
            slf_mean_chosen = slf_mean_left
            slf_idx_chosen = get_idx(slf_mean_chosen)
            slf_reward = slf_reward_left
            slf_reward_text = str(slf_reward)
            is_slf_resp = 1
        elif SLF_IMG_RESP.keys == right_key: 
            slf_img_fb_pos = IMG_POS_RIGHT
            slf_pos_chosen = POS_RIGHT
            slf_mean_chosen = slf_mean_right
            slf_idx_chosen = get_idx(slf_mean_chosen)
            slf_reward = slf_reward_right
            slf_reward_text = str(slf_reward)
            is_slf_resp = 1
        else:
            slf_img_fb_pos = IMG_POS_OUT
            slf_pos_chosen = POS_OUT
            slf_mean_chosen = 0
            slf_idx_chosen = np.nan
            slf_reward = 0
            slf_reward_text = FAILURE
            is_slf_resp = 0
        
        # Add points according to the response
        if slf_pos_chosen == slf_pos_correct:
            slf_acc = 1
        else:
            slf_acc = 0
        
        """============================================================="""
        # check responses
        if SLF_IMG_RESP.keys in ['', [], None]:  # No response was made
            SLF_IMG_RESP.keys = None
            # was no response the correct answer?!
            if str('slfCorrect').lower() == 'none':
               SLF_IMG_RESP.corr = 1;  # correct non-response
            else:
               SLF_IMG_RESP.corr = 0;  # failed to respond (incorrectly)
        # store data for SlfSequence (TrialHandler)
        SlfSequence.addData('SLF_IMG_RESP.keys',SLF_IMG_RESP.keys)
        SlfSequence.addData('SLF_IMG_RESP.corr', SLF_IMG_RESP.corr)
        if SLF_IMG_RESP.keys != None:  # we had a response
            SlfSequence.addData('SLF_IMG_RESP.rt', SLF_IMG_RESP.rt)
        # the Routine "SlfChoice" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        
        """============================================================="""
        # --- Prepare to start Routine "SlfChoiceFb" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        SLF_IMGFB_FRAME.setPos(slf_img_fb_pos)
        SLF_IMGFB_IMG_LEFT.setImage(slf_img_left)
        SLF_IMGFB_IMG_RIGHT.setImage(slf_img_right)
        # keep track of which components have finished
        SlfChoiceFbComponents = [SLF_IMGFB_FIX, 
                                 SLF_IMGFB_FRAME, 
                                 SLF_IMGFB_IMG_LEFT, 
                                 SLF_IMGFB_IMG_RIGHT]
        
        for thisComponent in SlfChoiceFbComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        
        """============================================================="""
        # --- Run Routine "SlfChoiceFb" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            """============================================================="""
            # *SLF_IMGFB_FIX* updates
            if SLF_IMGFB_FIX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMGFB_FIX.frameNStart = frameN  # exact frame index
                SLF_IMGFB_FIX.tStart = t  # local t and not account for scr refresh
                SLF_IMGFB_FIX.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMGFB_FIX, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMGFB_FIX.started')
                SLF_IMGFB_FIX.setAutoDraw(True)
            if SLF_IMGFB_FIX.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMGFB_FIX.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMGFB_FIX.tStop = t  # not accounting for scr refresh
                    SLF_IMGFB_FIX.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMGFB_FIX.stopped')
                    SLF_IMGFB_FIX.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_IMGFB_FRAME* updates
            if SLF_IMGFB_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMGFB_FRAME.frameNStart = frameN  # exact frame index
                SLF_IMGFB_FRAME.tStart = t  # local t and not account for scr refresh
                SLF_IMGFB_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMGFB_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMGFB_FRAME.started')
                SLF_IMGFB_FRAME.setAutoDraw(True)
            if SLF_IMGFB_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMGFB_FRAME.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMGFB_FRAME.tStop = t  # not accounting for scr refresh
                    SLF_IMGFB_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMGFB_FRAME.stopped')
                    SLF_IMGFB_FRAME.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_IMGFB_IMG_LEFT* updates
            if SLF_IMGFB_IMG_LEFT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMGFB_IMG_LEFT.frameNStart = frameN  # exact frame index
                SLF_IMGFB_IMG_LEFT.tStart = t  # local t and not account for scr refresh
                SLF_IMGFB_IMG_LEFT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMGFB_IMG_LEFT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMGFB_IMG_LEFT.started')
                SLF_IMGFB_IMG_LEFT.setAutoDraw(True)
            if SLF_IMGFB_IMG_LEFT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMGFB_IMG_LEFT.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMGFB_IMG_LEFT.tStop = t  # not accounting for scr refresh
                    SLF_IMGFB_IMG_LEFT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMGFB_IMG_LEFT.stopped')
                    SLF_IMGFB_IMG_LEFT.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_IMGFB_IMG_RIGHT* updates
            if SLF_IMGFB_IMG_RIGHT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_IMGFB_IMG_RIGHT.frameNStart = frameN  # exact frame index
                SLF_IMGFB_IMG_RIGHT.tStart = t  # local t and not account for scr refresh
                SLF_IMGFB_IMG_RIGHT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_IMGFB_IMG_RIGHT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_IMGFB_IMG_RIGHT.started')
                SLF_IMGFB_IMG_RIGHT.setAutoDraw(True)
            if SLF_IMGFB_IMG_RIGHT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_IMGFB_IMG_RIGHT.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_IMGFB_IMG_RIGHT.tStop = t  # not accounting for scr refresh
                    SLF_IMGFB_IMG_RIGHT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_IMGFB_IMG_RIGHT.stopped')
                    SLF_IMGFB_IMG_RIGHT.setAutoDraw(False)
            
            """============================================================="""
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            # if endExpNow:
                core.quit()

            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in SlfChoiceFbComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        
        """============================================================="""
        # --- Ending Routine "SlfChoiceFb" ---
        for thisComponent in SlfChoiceFbComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
        # the Routine "SlfChoiceFb" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        
        """============================================================="""
        # set up handler to look after randomisation of conditions etc
        SlfSkip = data.TrialHandler(
            nReps=is_slf_resp, method='random', 
            extraInfo=expInfo, originPath=-1,
            trialList=[None],
            seed=None, name='SlfSkip')
        thisExp.addLoop(SlfSkip)  # add the loop to the experiment
        thisSlfSkip = SlfSkip.trialList[0]  # so we can initialise stimuli with some values
        # abbreviate parameter names if possible (e.g. rgb = thisSlfSkip.rgb)
        if thisSlfSkip != None:
            for paramName in thisSlfSkip:
                exec('{} = thisSlfSkip[paramName]'.format(paramName))
        
        for thisSlfSkip in SlfSkip:
            currentLoop = SlfSkip
            # abbreviate parameter names if possible (e.g. rgb = thisSlfSkip.rgb)
            if thisSlfSkip != None:
                for paramName in thisSlfSkip:
                    exec('{} = thisSlfSkip[paramName]'.format(paramName))
            
            
            """============================================================="""
            # --- Prepare to start Routine "SlfConf" ---
            continueRoutine = True
            routineForceEnded = False
            # update component parameters for each repeat
            if slf_conf_pat == 1:
                slf_conf_left = CONF_LOW
                slf_conf_right = CONF_HIGH
                slf_conf_low_pos = CONF_POS_LEFT
                slf_conf_high_pos = CONF_POS_RIGHT
            else:
                slf_conf_left = CONF_HIGH
                slf_conf_right = CONF_LOW
                slf_conf_low_pos = CONF_POS_RIGHT
                slf_conf_high_pos = CONF_POS_LEFT

            SLF_CONF_LOW_BOX.setPos(slf_conf_low_pos)
            SLF_CONF_LOW_TEXT.setPos(slf_conf_low_pos)
            SLF_CONF_HIGH_BOX.setPos(slf_conf_high_pos)
            SLF_CONF_HIGH_TEXT.setPos(slf_conf_high_pos)
            # Run 'Begin Routine' code from SLF_CONF_CODE
            # Get the presented time in slf-confidence
            t_conf_pres = onset.getTime()
            SLF_CONF_RESP.keys = []
            SLF_CONF_RESP.rt = []
            _SLF_CONF_RESP_allKeys = []
            # keep track of which components have finished
            SlfConfComponents = [SLF_CONF_FIX,
                                 SLF_CONF_LOW_BOX, 
                                 SLF_CONF_LOW_TEXT, 
                                 SLF_CONF_HIGH_BOX, 
                                 SLF_CONF_HIGH_TEXT, 
                                 SLF_CONF_RESP]
            
            for thisComponent in SlfConfComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            
            """============================================================="""
            # --- Run Routine "SlfConf" ---
            while continueRoutine:
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                """============================================================="""
                # *SLF_CONF_FIX* updates
                if SLF_CONF_FIX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONF_FIX.frameNStart = frameN  # exact frame index
                    SLF_CONF_FIX.tStart = t  # local t and not account for scr refresh
                    SLF_CONF_FIX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONF_FIX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONF_FIX.started')
                    SLF_CONF_FIX.setAutoDraw(True)
                if SLF_CONF_FIX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONF_FIX.tStartRefresh + CONF_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONF_FIX.tStop = t  # not accounting for scr refresh
                        SLF_CONF_FIX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONF_FIX.stopped')
                        SLF_CONF_FIX.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONF_LOW_BOX* updates
                if SLF_CONF_LOW_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONF_LOW_BOX.frameNStart = frameN  # exact frame index
                    SLF_CONF_LOW_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_CONF_LOW_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONF_LOW_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONF_LOW_BOX.started')
                    SLF_CONF_LOW_BOX.setAutoDraw(True)
                if SLF_CONF_LOW_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONF_LOW_BOX.tStartRefresh + CONF_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONF_LOW_BOX.tStop = t  # not accounting for scr refresh
                        SLF_CONF_LOW_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONF_LOW_BOX.stopped')
                        SLF_CONF_LOW_BOX.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONF_LOW_TEXT* updates
                if SLF_CONF_LOW_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONF_LOW_TEXT.frameNStart = frameN  # exact frame index
                    SLF_CONF_LOW_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_CONF_LOW_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONF_LOW_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONF_LOW_TEXT.started')
                    SLF_CONF_LOW_TEXT.setAutoDraw(True)
                if SLF_CONF_LOW_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONF_LOW_TEXT.tStartRefresh + CONF_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONF_LOW_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_CONF_LOW_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONF_LOW_TEXT.stopped')
                        SLF_CONF_LOW_TEXT.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONF_HIGH_BOX* updates
                if SLF_CONF_HIGH_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONF_HIGH_BOX.frameNStart = frameN  # exact frame index
                    SLF_CONF_HIGH_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_CONF_HIGH_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONF_HIGH_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONF_HIGH_BOX.started')
                    SLF_CONF_HIGH_BOX.setAutoDraw(True)
                if SLF_CONF_HIGH_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONF_HIGH_BOX.tStartRefresh + CONF_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONF_HIGH_BOX.tStop = t  # not accounting for scr refresh
                        SLF_CONF_HIGH_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONF_HIGH_BOX.stopped')
                        SLF_CONF_HIGH_BOX.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONF_HIGH_TEXT* updates
                if SLF_CONF_HIGH_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONF_HIGH_TEXT.frameNStart = frameN  # exact frame index
                    SLF_CONF_HIGH_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_CONF_HIGH_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONF_HIGH_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONF_HIGH_TEXT.started')
                    SLF_CONF_HIGH_TEXT.setAutoDraw(True)
                if SLF_CONF_HIGH_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONF_HIGH_TEXT.tStartRefresh + CONF_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONF_HIGH_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_CONF_HIGH_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONF_HIGH_TEXT.stopped')
                        SLF_CONF_HIGH_TEXT.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONF_RESP* updates
                waitOnFlip = False
                if SLF_CONF_RESP.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONF_RESP.frameNStart = frameN  # exact frame index
                    SLF_CONF_RESP.tStart = t  # local t and not account for scr refresh
                    SLF_CONF_RESP.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONF_RESP, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONF_RESP.started')
                    SLF_CONF_RESP.status = STARTED
                    # keyboard checking is just starting
                    waitOnFlip = True
                    win.callOnFlip(SLF_CONF_RESP.clock.reset)  # t=0 on next screen flip
                    win.callOnFlip(SLF_CONF_RESP.clearEvents, eventType='keyboard')  # clear events on next screen flip
                if SLF_CONF_RESP.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONF_RESP.tStartRefresh + CONF_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONF_RESP.tStop = t  # not accounting for scr refresh
                        SLF_CONF_RESP.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONF_RESP.stopped')
                        SLF_CONF_RESP.status = FINISHED
                if SLF_CONF_RESP.status == STARTED and not waitOnFlip:
                    theseKeys = SLF_CONF_RESP.getKeys(keyList=[left_key,right_key], waitRelease=False)
                    _SLF_CONF_RESP_allKeys.extend(theseKeys)
                    if len(_SLF_CONF_RESP_allKeys):
                        SLF_CONF_RESP.keys = _SLF_CONF_RESP_allKeys[-1].name  # just the last key pressed
                        SLF_CONF_RESP.rt = _SLF_CONF_RESP_allKeys[-1].rt
                        # a response ends the routine
                        continueRoutine = False
                
                """============================================================="""
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                # if endExpNow:
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    routineForceEnded = True
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in SlfConfComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            
            """============================================================="""
            # --- Ending Routine "SlfConf" ---
            for thisComponent in SlfConfComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            
            """============================================================="""
            # Run 'End Routine' code from SLF_CONF_CODE
            # Get the responsed time in slf-confidence
            t_conf_resp = onset.getTime()
            
            # Receive confidence responses and determine feedback parameters
            if SLF_CONF_RESP.keys == left_key:
                slf_conf_fb_pos = CONF_POS_LEFT
                slf_conf = slf_conf_left
            elif SLF_CONF_RESP.keys == right_key: 
                slf_conf_fb_pos = CONF_POS_RIGHT
                slf_conf = slf_conf_right
            else:
                slf_conf_fb_pos = CONF_POS_OUT
                slf_conf = CONF_NONE
            # check responses
            if SLF_CONF_RESP.keys in ['', [], None]:  # No response was made
                SLF_CONF_RESP.keys = None
            SlfSkip.addData('SLF_CONF_RESP.keys',SLF_CONF_RESP.keys)
            if SLF_CONF_RESP.keys != None:  # we had a response
                SlfSkip.addData('SLF_CONF_RESP.rt', SLF_CONF_RESP.rt)
            # the Routine "SlfConf" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            
            
            """============================================================="""
            # --- Prepare to start Routine "SlfConfFb" ---
            continueRoutine = True
            routineForceEnded = False
            # update component parameters for each repeat
            SLF_CONFFB_FRAME.setPos(slf_conf_fb_pos)
            SLF_CONFFB_LOW_BOX.setPos(slf_conf_low_pos)
            SLF_CONFFB_LOW_TEXT.setPos(slf_conf_low_pos)
            SLF_CONFFB_HIGH_BOX.setPos(slf_conf_high_pos)
            SLF_CONFFB_HIGH_TEXT.setPos(slf_conf_high_pos)
            
            # keep track of which components have finished
            SlfConfFbComponents = [SLF_CONFFB_FIX,
                                   SLF_CONFFB_FRAME, 
                                   SLF_CONFFB_LOW_BOX, 
                                   SLF_CONFFB_LOW_TEXT, 
                                   SLF_CONFFB_HIGH_BOX, 
                                   SLF_CONFFB_HIGH_TEXT]
            
            for thisComponent in SlfConfFbComponents:
                thisComponent.tStart = None
                thisComponent.tStop = None
                thisComponent.tStartRefresh = None
                thisComponent.tStopRefresh = None
                if hasattr(thisComponent, 'status'):
                    thisComponent.status = NOT_STARTED
            # reset timers
            t = 0
            _timeToFirstFrame = win.getFutureFlipTime(clock="now")
            frameN = -1
            
            
            """============================================================="""
            # --- Run Routine "SlfConfFb" ---
            while continueRoutine:
                # get current time
                t = routineTimer.getTime()
                tThisFlip = win.getFutureFlipTime(clock=routineTimer)
                tThisFlipGlobal = win.getFutureFlipTime(clock=None)
                frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
                # update/draw components on each frame
                
                """============================================================="""
                # *SLF_CONFFB_FIX* updates
                if SLF_CONFFB_FIX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONFFB_FIX.frameNStart = frameN  # exact frame index
                    SLF_CONFFB_FIX.tStart = t  # local t and not account for scr refresh
                    SLF_CONFFB_FIX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONFFB_FIX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONFFB_FIX.started')
                    SLF_CONFFB_FIX.setAutoDraw(True)
                if SLF_CONFFB_FIX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONFFB_FIX.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONFFB_FIX.tStop = t  # not accounting for scr refresh
                        SLF_CONFFB_FIX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONFFB_FIX.stopped')
                        SLF_CONFFB_FIX.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONFFB_FRAME* updates
                if SLF_CONFFB_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONFFB_FRAME.frameNStart = frameN  # exact frame index
                    SLF_CONFFB_FRAME.tStart = t  # local t and not account for scr refresh
                    SLF_CONFFB_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONFFB_FRAME, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONFFB_FRAME.started')
                    SLF_CONFFB_FRAME.setAutoDraw(True)
                if SLF_CONFFB_FRAME.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONFFB_FRAME.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONFFB_FRAME.tStop = t  # not accounting for scr refresh
                        SLF_CONFFB_FRAME.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONFFB_FRAME.stopped')
                        SLF_CONFFB_FRAME.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONFFB_LOW_BOX* updates
                if SLF_CONFFB_LOW_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONFFB_LOW_BOX.frameNStart = frameN  # exact frame index
                    SLF_CONFFB_LOW_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_CONFFB_LOW_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONFFB_LOW_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONFFB_LOW_BOX.started')
                    SLF_CONFFB_LOW_BOX.setAutoDraw(True)
                if SLF_CONFFB_LOW_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONFFB_LOW_BOX.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONFFB_LOW_BOX.tStop = t  # not accounting for scr refresh
                        SLF_CONFFB_LOW_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONFFB_LOW_BOX.stopped')
                        SLF_CONFFB_LOW_BOX.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONFFB_LOW_TEXT* updates
                if SLF_CONFFB_LOW_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONFFB_LOW_TEXT.frameNStart = frameN  # exact frame index
                    SLF_CONFFB_LOW_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_CONFFB_LOW_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONFFB_LOW_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONFFB_LOW_TEXT.started')
                    SLF_CONFFB_LOW_TEXT.setAutoDraw(True)
                if SLF_CONFFB_LOW_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONFFB_LOW_TEXT.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONFFB_LOW_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_CONFFB_LOW_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONFFB_LOW_TEXT.stopped')
                        SLF_CONFFB_LOW_TEXT.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONFFB_HIGH_BOX* updates
                if SLF_CONFFB_HIGH_BOX.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONFFB_HIGH_BOX.frameNStart = frameN  # exact frame index
                    SLF_CONFFB_HIGH_BOX.tStart = t  # local t and not account for scr refresh
                    SLF_CONFFB_HIGH_BOX.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONFFB_HIGH_BOX, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONFFB_HIGH_BOX.started')
                    SLF_CONFFB_HIGH_BOX.setAutoDraw(True)
                if SLF_CONFFB_HIGH_BOX.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONFFB_HIGH_BOX.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONFFB_HIGH_BOX.tStop = t  # not accounting for scr refresh
                        SLF_CONFFB_HIGH_BOX.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONFFB_HIGH_BOX.stopped')
                        SLF_CONFFB_HIGH_BOX.setAutoDraw(False)
                
                """============================================================="""
                # *SLF_CONFFB_HIGH_TEXT* updates
                if SLF_CONFFB_HIGH_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                    # keep track of start time/frame for later
                    SLF_CONFFB_HIGH_TEXT.frameNStart = frameN  # exact frame index
                    SLF_CONFFB_HIGH_TEXT.tStart = t  # local t and not account for scr refresh
                    SLF_CONFFB_HIGH_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                    win.timeOnFlip(SLF_CONFFB_HIGH_TEXT, 'tStartRefresh')  # time at next scr refresh
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_CONFFB_HIGH_TEXT.started')
                    SLF_CONFFB_HIGH_TEXT.setAutoDraw(True)
                if SLF_CONFFB_HIGH_TEXT.status == STARTED:
                    # is it time to stop? (based on global clock, using actual start)
                    if tThisFlipGlobal > SLF_CONFFB_HIGH_TEXT.tStartRefresh + FEEDBACK_TIME-frameTolerance:
                        # keep track of stop time/frame for later
                        SLF_CONFFB_HIGH_TEXT.tStop = t  # not accounting for scr refresh
                        SLF_CONFFB_HIGH_TEXT.frameNStop = frameN  # exact frame index
                        # add timestamp to datafile
                        thisExp.timestampOnFlip(win, 'SLF_CONFFB_HIGH_TEXT.stopped')
                        SLF_CONFFB_HIGH_TEXT.setAutoDraw(False)
                
                """============================================================="""
                # check for quit (typically the Esc key)
                if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
                # if endExpNow:
                    core.quit()
                
                # check if all components have finished
                if not continueRoutine:  # a component has requested a forced-end of Routine
                    routineForceEnded = True
                    break
                continueRoutine = False  # will revert to True if at least one component still running
                for thisComponent in SlfConfFbComponents:
                    if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                        continueRoutine = True
                        break  # at least one component has not yet finished
                
                # refresh the screen
                if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                    win.flip()
            
            
            """============================================================="""
            # --- Ending Routine "SlfConfFb" ---
            for thisComponent in SlfConfFbComponents:
                if hasattr(thisComponent, "setAutoDraw"):
                    thisComponent.setAutoDraw(False)
            # the Routine "SlfConfFb" was not non-slip safe, so reset the non-slip timer
            routineTimer.reset()
            thisExp.nextEntry()
            
        # completed is_slf_resp repeats of 'SlfSkip'
        
        # get names of stimulus parameters
        if SlfSkip.trialList in ([], [None], None):
            params = []
        else:
            params = SlfSkip.trialList[0].keys()
        # save data for this loop
        SlfSkip.saveAsExcel(filename + '.xlsx', sheetName='SlfSkip',
            stimOut=params,
            dataOut=['n','all_mean','all_std', 'all_raw'])
        SlfSkip.saveAsText(filename + 'SlfSkip.csv', delim=',',
            stimOut=params,
            dataOut=['n','all_mean','all_std', 'all_raw'])
        
        
        """============================================================="""
        # --- Prepare to start Routine "SlfReward" ---
        continueRoutine = True
        routineForceEnded = False
        # update component parameters for each repeat
        # Run 'Begin Routine' code from SLF_REWARD_TEXT_CODE
        # Get the presented time in slf-reward
        t_reward_pres = onset.getTime()
        SLF_REWARD_FRAME.setPos(slf_img_fb_pos)
        SLF_REWARD_IMG_LEFT.setImage(slf_img_left)
        SLF_REWARD_IMG_RIGHT.setImage(slf_img_right)
        SLF_REWARD_TEXT.setText(slf_reward_text)
        # keep track of which components have finished
        SlfRewardComponents = [SLF_REWARD_FRAME, 
                               SLF_REWARD_IMG_LEFT, 
                               SLF_REWARD_IMG_RIGHT, 
                               SLF_REWARD_TEXT]
        for thisComponent in SlfRewardComponents:
            thisComponent.tStart = None
            thisComponent.tStop = None
            thisComponent.tStartRefresh = None
            thisComponent.tStopRefresh = None
            if hasattr(thisComponent, 'status'):
                thisComponent.status = NOT_STARTED
        # reset timers
        t = 0
        _timeToFirstFrame = win.getFutureFlipTime(clock="now")
        frameN = -1
        
        
        """============================================================="""
        # --- Run Routine "SlfReward" ---
        while continueRoutine:
            # get current time
            t = routineTimer.getTime()
            tThisFlip = win.getFutureFlipTime(clock=routineTimer)
            tThisFlipGlobal = win.getFutureFlipTime(clock=None)
            frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
            # update/draw components on each frame
            
            """============================================================="""
            # *SLF_REWARD_FRAME* updates
            if SLF_REWARD_FRAME.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_REWARD_FRAME.frameNStart = frameN  # exact frame index
                SLF_REWARD_FRAME.tStart = t  # local t and not account for scr refresh
                SLF_REWARD_FRAME.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_REWARD_FRAME, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_REWARD_FRAME.started')
                SLF_REWARD_FRAME.setAutoDraw(True)
            if SLF_REWARD_FRAME.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_REWARD_FRAME.tStartRefresh + REWARD_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_REWARD_FRAME.tStop = t  # not accounting for scr refresh
                    SLF_REWARD_FRAME.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_REWARD_FRAME.stopped')
                    SLF_REWARD_FRAME.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_REWARD_IMG_LEFT* updates
            if SLF_REWARD_IMG_LEFT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_REWARD_IMG_LEFT.frameNStart = frameN  # exact frame index
                SLF_REWARD_IMG_LEFT.tStart = t  # local t and not account for scr refresh
                SLF_REWARD_IMG_LEFT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_REWARD_IMG_LEFT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_REWARD_IMG_LEFT.started')
                SLF_REWARD_IMG_LEFT.setAutoDraw(True)
            if SLF_REWARD_IMG_LEFT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_REWARD_IMG_LEFT.tStartRefresh + REWARD_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_REWARD_IMG_LEFT.tStop = t  # not accounting for scr refresh
                    SLF_REWARD_IMG_LEFT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_REWARD_IMG_LEFT.stopped')
                    SLF_REWARD_IMG_LEFT.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_REWARD_IMG_RIGHT* updates
            if SLF_REWARD_IMG_RIGHT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
                # keep track of start time/frame for later
                SLF_REWARD_IMG_RIGHT.frameNStart = frameN  # exact frame index
                SLF_REWARD_IMG_RIGHT.tStart = t  # local t and not account for scr refresh
                SLF_REWARD_IMG_RIGHT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_REWARD_IMG_RIGHT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_REWARD_IMG_RIGHT.started')
                SLF_REWARD_IMG_RIGHT.setAutoDraw(True)
            if SLF_REWARD_IMG_RIGHT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_REWARD_IMG_RIGHT.tStartRefresh + REWARD_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_REWARD_IMG_RIGHT.tStop = t  # not accounting for scr refresh
                    SLF_REWARD_IMG_RIGHT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_REWARD_IMG_RIGHT.stopped')
                    SLF_REWARD_IMG_RIGHT.setAutoDraw(False)
            
            """============================================================="""
            # *SLF_REWARD_TEXT* updates
            if SLF_REWARD_TEXT.status == NOT_STARTED and tThisFlip >= 0-frameTolerance:
                # keep track of start time/frame for later
                SLF_REWARD_TEXT.frameNStart = frameN  # exact frame index
                SLF_REWARD_TEXT.tStart = t  # local t and not account for scr refresh
                SLF_REWARD_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
                win.timeOnFlip(SLF_REWARD_TEXT, 'tStartRefresh')  # time at next scr refresh
                # add timestamp to datafile
                thisExp.timestampOnFlip(win, 'SLF_REWARD_TEXT.started')
                SLF_REWARD_TEXT.setAutoDraw(True)
            if SLF_REWARD_TEXT.status == STARTED:
                # is it time to stop? (based on global clock, using actual start)
                if tThisFlipGlobal > SLF_REWARD_TEXT.tStartRefresh + REWARD_TIME-frameTolerance:
                    # keep track of stop time/frame for later
                    SLF_REWARD_TEXT.tStop = t  # not accounting for scr refresh
                    SLF_REWARD_TEXT.frameNStop = frameN  # exact frame index
                    # add timestamp to datafile
                    thisExp.timestampOnFlip(win, 'SLF_REWARD_TEXT.stopped')
                    SLF_REWARD_TEXT.setAutoDraw(False)
            
            """============================================================="""
            # check for quit (typically the Esc key)
            if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
            # if endExpNow:
                core.quit()
            
            # check if all components have finished
            if not continueRoutine:  # a component has requested a forced-end of Routine
                routineForceEnded = True
                break
            continueRoutine = False  # will revert to True if at least one component still running
            for thisComponent in SlfRewardComponents:
                if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                    continueRoutine = True
                    break  # at least one component has not yet finished
            
            # refresh the screen
            if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
                win.flip()
        
        
        """============================================================="""
        # --- Ending Routine "SlfReward" ---
        for thisComponent in SlfRewardComponents:
            if hasattr(thisComponent, "setAutoDraw"):
                thisComponent.setAutoDraw(False)
                
        """============================================================="""
        # Run 'End Routine' code from SLF_REWARD_TEXT_CODE
        # Get the responsed time in slf-reward
        t_reward_resp = onset.getTime()
        
        slf_img_time = t_img_resp - t_img_pres
        slf_conf_time = t_conf_resp - t_conf_pres
        
        # Output slf data
        slf_data_table.append([block_cnt, through_game_cnt, slf_seq_type, slf_trial_cnt, 
                               slf_condition, slf_pair_pat, slf_order_pat, # from file
                               slf_mean_left, slf_mean_right, slf_sd, 
                               slf_idx_left, slf_idx_right,
                               slf_pos_correct, slf_mean_correct, slf_idx_correct, 
                               slf_pos_chosen, slf_mean_chosen, slf_idx_chosen,  
                               slf_conf_pat, slf_conf, slf_acc, slf_reward,
                               np.round(slf_img_time, 3), np.round(slf_conf_time, 3), 
                               np.round(t_img_pres, 3), np.round(t_img_resp, 3),
                               np.round(t_conf_pres, 3), np.round(t_conf_resp, 3),
                               np.round(t_reward_pres, 3), np.round(t_reward_resp, 3)])
        
        # Save excel file
        out_book = load_workbook(filename=out_xlsx)
        slf_sheet_name = out_book.sheetnames[0]
        slf_sheet = out_book[slf_sheet_name]
        for i in range(1, len(slf_data_table[slf_trial_cnt-1])+1):
            slf_sheet.cell(slf_trial_cnt+1, i, value = slf_data_table[slf_trial_cnt-1][i-1])
        out_book.save(out_xlsx)
        
        # Update slf-count
        slf_trial_cnt += 1
        through_game_cnt += 1
        # the Routine "SlfReward" was not non-slip safe, so reset the non-slip timer
        routineTimer.reset()
        thisExp.nextEntry()
        
    # completed 1.0 repeats of 'SlfSequence'
    
    """============================================================="""
    # get names of stimulus parameters
    if SlfSequence.trialList in ([], [None], None):
        params = []
    else:
        params = SlfSequence.trialList[0].keys()
    # save data for this loop
    SlfSequence.saveAsExcel(filename + '.xlsx', sheetName='SlfSequence',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    SlfSequence.saveAsText(filename + 'SlfSequence.csv', delim=',',
        stimOut=params,
        dataOut=['n','all_mean','all_std', 'all_raw'])
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    """============================================================="""
    # --- Prepare to start Routine "BlockRest" ---
    continueRoutine = True
    routineForceEnded = False
    # update component parameters for each repeat
    # Run 'Begin Routine' code from BLOCK_REST_CODE
    if block_cnt < BLOCK_NUM:
        message = f'{block_cnt} Block is over.\nTake a break.\nPress "space" to start next Block.'
    else:
        message = 'Game is over.\nPress "space" to exit.'
        
    BLOCK_REST_TEXT.setText(message)
    BLOCK_REST_SKIP.keys = []
    BLOCK_REST_SKIP.rt = []
    _BLOCK_REST_SKIP_allKeys = []
    # keep track of which components have finished
    BlockRestComponents = [BLOCK_REST_TEXT, 
                           BLOCK_REST_SKIP]
    
    for thisComponent in BlockRestComponents:
        thisComponent.tStart = None
        thisComponent.tStop = None
        thisComponent.tStartRefresh = None
        thisComponent.tStopRefresh = None
        if hasattr(thisComponent, 'status'):
            thisComponent.status = NOT_STARTED
    # reset timers
    t = 0
    _timeToFirstFrame = win.getFutureFlipTime(clock="now")
    frameN = -1
    
    
    """============================================================="""
    # --- Run Routine "BlockRest" ---
    while continueRoutine:
        # get current time
        t = routineTimer.getTime()
        tThisFlip = win.getFutureFlipTime(clock=routineTimer)
        tThisFlipGlobal = win.getFutureFlipTime(clock=None)
        frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
        # update/draw components on each frame
        
        """============================================================="""
        # *BLOCK_REST_TEXT* updates
        if BLOCK_REST_TEXT.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            BLOCK_REST_TEXT.frameNStart = frameN  # exact frame index
            BLOCK_REST_TEXT.tStart = t  # local t and not account for scr refresh
            BLOCK_REST_TEXT.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(BLOCK_REST_TEXT, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'BLOCK_REST_TEXT.started')
            BLOCK_REST_TEXT.setAutoDraw(True)
        
        """============================================================="""
        # *BLOCK_REST_SKIP* updates
        waitOnFlip = False
        if BLOCK_REST_SKIP.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
            # keep track of start time/frame for later
            BLOCK_REST_SKIP.frameNStart = frameN  # exact frame index
            BLOCK_REST_SKIP.tStart = t  # local t and not account for scr refresh
            BLOCK_REST_SKIP.tStartRefresh = tThisFlipGlobal  # on global time
            win.timeOnFlip(BLOCK_REST_SKIP, 'tStartRefresh')  # time at next scr refresh
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'BLOCK_REST_SKIP.started')
            BLOCK_REST_SKIP.status = STARTED
            # keyboard checking is just starting
            waitOnFlip = True
            win.callOnFlip(BLOCK_REST_SKIP.clock.reset)  # t=0 on next screen flip
            win.callOnFlip(BLOCK_REST_SKIP.clearEvents, eventType='keyboard')  # clear events on next screen flip
        if BLOCK_REST_SKIP.status == STARTED and not waitOnFlip:
            theseKeys = BLOCK_REST_SKIP.getKeys(keyList=['space'], waitRelease=False)
            _BLOCK_REST_SKIP_allKeys.extend(theseKeys)
            if len(_BLOCK_REST_SKIP_allKeys):
                BLOCK_REST_SKIP.keys = _BLOCK_REST_SKIP_allKeys[-1].name  # just the last key pressed
                BLOCK_REST_SKIP.rt = _BLOCK_REST_SKIP_allKeys[-1].rt
                # a response ends the routine
                continueRoutine = False
        
        """============================================================="""
        # check for quit (typically the Esc key)
        if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
        # if endExpNow:
            core.quit()
        
        # check if all components have finished
        if not continueRoutine:  # a component has requested a forced-end of Routine
            routineForceEnded = True
            break
        continueRoutine = False  # will revert to True if at least one component still running
        for thisComponent in BlockRestComponents:
            if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
                continueRoutine = True
                break  # at least one component has not yet finished
        
        # refresh the screen
        if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
            win.flip()
    
    
    """============================================================="""
    # --- Ending Routine "BlockRest" ---
    for thisComponent in BlockRestComponents:
        if hasattr(thisComponent, "setAutoDraw"):
            thisComponent.setAutoDraw(False)
            
    """============================================================="""
    # Run 'End Routine' code from BLOCK_REST_CODE
    block_cnt += 1
    core.wait(STANDBY_TIME)
    
    """============================================================="""
    # check responses
    if BLOCK_REST_SKIP.keys in ['', [], None]:  # No response was made
        BLOCK_REST_SKIP.keys = None
    BlockLoop.addData('BLOCK_REST_SKIP.keys',BLOCK_REST_SKIP.keys)
    if BLOCK_REST_SKIP.keys != None:  # we had a response
        BlockLoop.addData('BLOCK_REST_SKIP.rt', BLOCK_REST_SKIP.rt)
    # the Routine "BlockRest" was not non-slip safe, so reset the non-slip timer
    routineTimer.reset()
    thisExp.nextEntry()
    
# completed 1.0 repeats of 'BlockLoop'

# get names of stimulus parameters
if BlockLoop.trialList in ([], [None], None):
    params = []
else:
    params = BlockLoop.trialList[0].keys()
# save data for this loop
BlockLoop.saveAsExcel(filename + '.xlsx', sheetName='BlockLoop',
    stimOut=params,
    dataOut=['n','all_mean','all_std', 'all_raw'])
BlockLoop.saveAsText(filename + 'BlockLoop.csv', delim=',',
    stimOut=params,
    dataOut=['n','all_mean','all_std', 'all_raw'])


"""============================================================="""
# --- Prepare to start Routine "Appreciation" ---
continueRoutine = True
routineForceEnded = False
# update component parameters for each repeat
# keep track of which components have finished
AppreciationComponents = [APPRECIATION]

for thisComponent in AppreciationComponents:
    thisComponent.tStart = None
    thisComponent.tStop = None
    thisComponent.tStartRefresh = None
    thisComponent.tStopRefresh = None
    if hasattr(thisComponent, 'status'):
        thisComponent.status = NOT_STARTED
# reset timers
t = 0
_timeToFirstFrame = win.getFutureFlipTime(clock="now")
frameN = -1


"""============================================================="""
# --- Run Routine "Appreciation" ---
while continueRoutine and routineTimer.getTime() < APPRECIATION_TIME:
    # get current time
    t = routineTimer.getTime()
    tThisFlip = win.getFutureFlipTime(clock=routineTimer)
    tThisFlipGlobal = win.getFutureFlipTime(clock=None)
    frameN = frameN + 1  # number of completed frames (so 0 is the first frame)
    # update/draw components on each frame
    
    """============================================================="""
    # *APPRECIATION* updates
    if APPRECIATION.status == NOT_STARTED and tThisFlip >= 0.0-frameTolerance:
        # keep track of start time/frame for later
        APPRECIATION.frameNStart = frameN  # exact frame index
        APPRECIATION.tStart = t  # local t and not account for scr refresh
        APPRECIATION.tStartRefresh = tThisFlipGlobal  # on global time
        win.timeOnFlip(APPRECIATION, 'tStartRefresh')  # time at next scr refresh
        # add timestamp to datafile
        thisExp.timestampOnFlip(win, 'APPRECIATION.started')
        APPRECIATION.setAutoDraw(True)
    if APPRECIATION.status == STARTED:
        # is it time to stop? (based on global clock, using actual start)
        if tThisFlipGlobal > APPRECIATION.tStartRefresh + APPRECIATION_TIME-frameTolerance:
            # keep track of stop time/frame for later
            APPRECIATION.tStop = t  # not accounting for scr refresh
            APPRECIATION.frameNStop = frameN  # exact frame index
            # add timestamp to datafile
            thisExp.timestampOnFlip(win, 'APPRECIATION.stopped')
            APPRECIATION.setAutoDraw(False)
    
    """============================================================="""
    # check for quit (typically the Esc key)
    if endExpNow or defaultKeyboard.getKeys(keyList=["escape"]):
    # if endExpNow:
        core.quit()
    
    # check if all components have finished
    if not continueRoutine:  # a component has requested a forced-end of Routine
        routineForceEnded = True
        break
    continueRoutine = False  # will revert to True if at least one component still running
    for thisComponent in AppreciationComponents:
        if hasattr(thisComponent, "status") and thisComponent.status != FINISHED:
            continueRoutine = True
            break  # at least one component has not yet finished
    
    # refresh the screen
    if continueRoutine:  # don't flip if this routine is over or we'll get a blank screen
        win.flip()


"""============================================================="""
# --- Ending Routine "Appreciation" ---
for thisComponent in AppreciationComponents:
    if hasattr(thisComponent, "setAutoDraw"):
        thisComponent.setAutoDraw(False)
# using non-slip timing so subtract the expected duration of this Routine (unless ended on request)
if routineForceEnded:
    routineTimer.reset()
else:
    routineTimer.addTime(-APPRECIATION_TIME)


"""============================================================="""
# --- End experiment ---
# Flip one final time so any remaining win.callOnFlip() 
# and win.timeOnFlip() tasks get executed before quitting
win.flip()

# these shouldn't be strictly necessary (should auto-save)
thisExp.saveAsWideText(filename+'.csv', delim='auto')
thisExp.saveAsPickle(filename)
# make sure everything is closed down
if eyetracker:
    eyetracker.setConnectionState(False)
thisExp.abort()  # or data files will save again on exit
win.close()
core.quit()
