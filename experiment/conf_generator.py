#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Import modules
import random 

import numpy as np
from openpyxl import load_workbook
import pandas as pd

import q_threshold_generator as qth_gen
import template_experiment as tp

# NOTE: success seeds are slf: '5, 8, 17, 20'
#                         obs: '29, 48'
SEED = 0
GAME_PAT = 5
BLOCK_NUM = 4
TRIAL_NUM = 9
MEAN = [30, 40, 50]
SD = 7

LEFT = 'left'
RIGHT = 'right'
LOW = 'low'
HIGH = 'high'

ALPHA = 0.216
BETA = 0.141
Q_THRESHOLD = 10.735

SIM_NUM = 2**10

# Fix seed
print(f'{SEED=}')
random.seed(SEED)
np.random.seed(seed=SEED)

class RLmodel:
    def __init__(self, pair_pats, order_pats, chosens, rewards):        
        # self.q_values = np.zeros((3,), dtype=int)
        self.q_values = np.full(3, 20)
        self.pair_pats = pair_pats
        self.order_pats = order_pats
        self.chosens = chosens
        self.rewards = rewards

    def softmax(self, x, beta):
        return np.exp(beta * x) / np.sum(np.exp(beta * x), axis=0)

    def choose_action(self, q_values, pair_pat):
        policy = np.zeros((3,))
        policy[pair_pat] = self.softmax(q_values[pair_pat], BETA)
        action = np.random.choice(np.arange(3), p=policy)
        return policy, action
    
    def reward_generator(self):
        rewards = np.array([round(np.random.normal(mean, SD)) for mean in MEAN])
        return rewards
    
    def update_q_value(self, q_value, action, reward, alpha):
        q_value[action] += alpha * (reward[action] - q_value[action])
        return q_value
    
    def simulate(self):
        self.confs = []
        
        for block in range(BLOCK_NUM):
            for trial in range(TRIAL_NUM):
                pair_pat = self.pair_pats[block, trial]
                order_pat = self.order_pats[block, trial]
                chosen_pos = self.chosens[block, trial]
                reward = self.rewards[block, trial]
                if pair_pat == 1:
                    if order_pat == 1:
                        if chosen_pos == LEFT:
                            idx_chosen = 0
                            idx_unchosen = 1
                        elif chosen_pos == RIGHT:
                            idx_chosen = 1
                            idx_unchosen = 0
                    elif order_pat == 2:
                        if chosen_pos == LEFT:
                            idx_chosen = 1
                            idx_unchosen = 0
                        elif chosen_pos == RIGHT:
                            idx_chosen = 0
                            idx_unchosen = 1
                elif pair_pat == 2:
                    if order_pat == 1:
                        if chosen_pos == LEFT:
                            idx_chosen = 1
                            idx_unchosen = 2
                        elif chosen_pos == RIGHT:
                            idx_chosen = 2
                            idx_unchosen = 1
                    elif order_pat == 2:
                        if chosen_pos == LEFT:
                            idx_chosen = 2
                            idx_unchosen = 1
                        elif chosen_pos == RIGHT:
                            idx_chosen = 1
                            idx_unchosen = 2
                elif pair_pat == 3:
                    if order_pat == 1:
                        if chosen_pos == LEFT:
                            idx_chosen = 2
                            idx_unchosen = 0
                        elif chosen_pos == RIGHT:
                            idx_chosen = 0
                            idx_unchosen = 2
                    elif order_pat == 2:
                        if chosen_pos == LEFT:
                            idx_chosen = 0
                            idx_unchosen = 2
                        elif chosen_pos == RIGHT:
                            idx_chosen = 2
                            idx_unchosen = 0
                
                if self.q_values[idx_chosen] - self.q_values[idx_unchosen] < Q_THRESHOLD:
                    conf = LOW
                else:
                    conf = HIGH
                print(self.q_values)
                self.q_values[idx_chosen] += ALPHA * (self.rewards[block, trial] - self.q_values[idx_chosen])
                self.confs.append(conf)


def load_data(game_pat):
    pair_pats = []
    order_pats = []
    chosens = []
    rewards = []
    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        df = (pd.read_excel(f'{condition_dir_path}game{game_pat}/obs{game_pat}{block+1}.xlsx', sheet_name='condition'))
        pair_pats.append(df['obs_pair_pat'].values)
        order_pats.append(df['obs_order_pat'].values)
        chosens.append(df['obs_pos_chosen'].values)
        rewards.append(df['obs_reward'].values)
    pair_pats = np.array(pair_pats)
    order_pats = np.array(order_pats)
    chosens = np.array(chosens)
    rewards = np.array(rewards)

    return pair_pats, order_pats, chosens, rewards 


def reward_generator(rept=BLOCK_NUM*TRIAL_NUM):
    rewards = []
    for _ in range(rept):
        rewards.append([round(np.random.normal(mean, SD)) for mean in MEAN])
    rewards = np.array(rewards)
    return rewards


def conf_pats_generator():
    conf_pats = [1]*(BLOCK_NUM*TRIAL_NUM//2) + [2]*(BLOCK_NUM*TRIAL_NUM//2)
    random.shuffle(conf_pats)
    conf_pats = np.array(conf_pats)
    return conf_pats


# def check_conf(conf_list):
#     for block in range(BLOCK_NUM):
#         if block == 0:
#             if acc_list[block] < 0.5:
#                 return False
#         else:
#             if acc_list[block] < acc_list[block-1]:
#                 return False
            
#     acc_mean = acc_list.mean()
#     if acc_mean < 0.78:
#         return False
    
#     return True

def write_excel(confs, game_pat):

    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        condition_file_path = f'{condition_dir_path}game{game_pat}/obs{game_pat}{block+1}.xlsx'
        wb = load_workbook(condition_file_path)
        ws = wb['condition']
        for trial in range(TRIAL_NUM):
            ws[f'H{trial+2}'] = confs[BLOCK_NUM*block+trial]
        wb.save(condition_file_path)


def main():
    game_pat = GAME_PAT
    pair_pats, order_pats, chosens, rewards = load_data(game_pat)
    print(pair_pats)
    print(order_pats)
    print(chosens)
    print(rewards)

    model = RLmodel(pair_pats, order_pats, chosens, rewards)
    model.simulate()
    confs = np.array(model.confs)
    write_excel(confs, game_pat)

if __name__ == '__main__':
    main()
