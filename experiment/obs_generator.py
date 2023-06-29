#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Import modules
import random 

import numpy as np
from openpyxl import load_workbook
import pandas as pd

import template_experiment as tp

# NOTE: success seeds are slf: '5, 8, 17, 20'
#                         obs: '29, 48, 77, 107'
#                         T: '5'
SEED = 5
GAME_PAT = 5
BLOCK_NUM = 4
TRIAL_NUM = 9
MEAN = [30, 40, 50]
SD = 7

LEFT = 'left'
RIGHT = 'right'

ALPHA = 0.216
BETA = 0.141

SIM_NUM = 2**10

# Fix seed
print(f'{SEED=}')
random.seed(SEED)
np.random.seed(seed=SEED)

class RLmodel:
    def __init__(self, pair_pats, rewards):        
        # self.q_values = np.zeros((3,), dtype=int)
        self.q_values = np.full(3, 20)
        self.pair_pats = pair_pats
        self.rewards = rewards

    def softmax(self, x, beta):
        return np.exp(beta * x) / np.sum(np.exp(beta * x), axis=0)

    def choose_action(self, q_values, pair_pat):
        policy = np.zeros((3,))
        policy[pair_pat] = self.softmax(q_values[pair_pat], BETA)
        action = np.random.choice(np.arange(3), p=policy)
        return policy, action
    
    def reward_generator(self):
        rewards = np.array([round(np.random.normal(mean, SD)) for mean in MEAN])
        return rewards
    
    def update_q_value(self, q_value, action, reward, alpha):
        q_value[action] += alpha * (reward[action] - q_value[action])
        return q_value
    
    def simulate(self):
        self.acc_list = []
        self.actions = []
        
        for block in range(BLOCK_NUM):
            acc = 0
            for trial in range(TRIAL_NUM):
                pair_pat = self.pair_pats[block, trial]
                if pair_pat == 1:
                    pair_pat_ = np.array([True, True, False])
                elif pair_pat == 2:
                    pair_pat_ = np.array([False, True, True])
                elif pair_pat == 3:
                    pair_pat_ = np.array([True, False, True])
                
                policy, action = self.choose_action(self.q_values, pair_pat_)
                # rewards = self.reward_generator()
                # print(self.rewards[TRIAL_NUM*block+trial])
                self.q_values = self.update_q_value(self.q_values, action, self.rewards[TRIAL_NUM*block+trial], ALPHA)
                self.actions.append(action)
                if (pair_pat==1 and action==1) or (pair_pat==2 and action==2) or (pair_pat==3 and action==2):
                    acc += 1
                
                # print(f'Policy: {policy}')
                # print(f'Action: {action}')
                # print(f'Q_value: {self.q_values}\n')
            
            self.acc_list.append(acc / TRIAL_NUM)


def load_data(game_pat):
    pair_pats = []
    order_pats = []
    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        df = (pd.read_excel(f'{condition_dir_path}game{game_pat}/obs{game_pat}{block+1}.xlsx', sheet_name='condition'))
        pair_pats.append(df['obs_pair_pat'].values)
        order_pats.append(df['obs_order_pat'].values)
    pair_pats = np.array(pair_pats)
    order_pats = np.array(order_pats)

    return pair_pats, order_pats


def reward_generator(rept=BLOCK_NUM*TRIAL_NUM):
    rewards = []
    for _ in range(rept):
        rewards.append([round(np.random.normal(mean, SD)) for mean in MEAN])
    rewards = np.array(rewards)
    return rewards


def conf_pats_generator():
    conf_pats = [1]*(BLOCK_NUM*TRIAL_NUM//2) + [2]*(BLOCK_NUM*TRIAL_NUM//2)
    random.shuffle(conf_pats)
    conf_pats = np.array(conf_pats)
    return conf_pats


def check_acc(acc_list):
    for block in range(BLOCK_NUM):
        if block == 0:
            if acc_list[block] < 0.5:
                return False
        else:
            if acc_list[block] < acc_list[block-1]:
                return False
            
    acc_mean = acc_list.mean()
    if acc_mean < 0.78:
        return False
    
    return True

def write_excel(pair_pats, order_pats, actions, rewards, game_pat):

    pair_pats_reshape = pair_pats.reshape(-1)
    order_pats_reshape = order_pats.reshape(-1)

    pos_chosen_ = []
    rewards_ = []
    rewards_chosen_ = []
    
    for pair_pat, order_pat, action, reward in zip(pair_pats_reshape, order_pats_reshape, actions, rewards):
        if pair_pat==1:
            if order_pat==1:
                reward_left = reward[0]
                reward_right = reward[1]
                if action==0:
                    pos_chosen = LEFT
                elif action==1:
                    pos_chosen = RIGHT
            elif order_pat==2:
                reward_left = reward[1]
                reward_right = reward[0]
                if action==1:
                    pos_chosen = LEFT
                elif action==0:
                    pos_chosen = RIGHT
        if pair_pat==2:
            if order_pat==1:
                reward_left = reward[1]
                reward_right = reward[2]
                if action==1:
                    pos_chosen = LEFT
                elif action==2:
                    pos_chosen = RIGHT
            elif order_pat==2:
                reward_left = reward[2]
                reward_right = reward[1]
                if action==2:
                    pos_chosen = LEFT
                elif action==1:
                    pos_chosen = RIGHT
        if pair_pat==3:
            if order_pat==1:
                reward_left = reward[2]
                reward_right = reward[0]
                if action==2:
                    pos_chosen = LEFT
                elif action==0:
                    pos_chosen = RIGHT
            elif order_pat==2:
                reward_left = reward[0]
                reward_right = reward[2]
                if action==0:
                    pos_chosen = LEFT
                elif action==2:
                    pos_chosen = RIGHT

        pos_chosen_.append(pos_chosen)
        rewards_.append([reward_left, reward_right])
        rewards_chosen_.append(reward[action])

    # print(pos_chosen_)
    # print(rewards_)
    # print(rewards_chosen_)

    condition_dir_path = tp.get_condition_dir_path()
    for block in range(BLOCK_NUM):
        condition_file_path = f'{condition_dir_path}game{game_pat}/obs{game_pat}{block+1}.xlsx'
        wb = load_workbook(condition_file_path)
        ws = wb['condition']
        for trial in range(TRIAL_NUM):
            ws[f'E{trial+2}'] = rewards_[TRIAL_NUM*block+trial][0]
            ws[f'F{trial+2}'] = rewards_[TRIAL_NUM*block+trial][1]
            ws[f'G{trial+2}'] = pos_chosen_[TRIAL_NUM*block+trial]
            ws[f'I{trial+2}'] = rewards_chosen_[TRIAL_NUM*block+trial]
        wb.save(condition_file_path)


def main():
    game_pat = GAME_PAT
    acc_lists = []
    pair_pats, order_pats = load_data(game_pat)
    rewards = reward_generator()
    # print(rewards)

    for _ in range(SIM_NUM):
        model = RLmodel(pair_pats, rewards)
        model.simulate()
        acc_lists.append(model.acc_list)
    
    acc_mean = np.mean(acc_lists, axis=0)
    print(f'Accuracy: {acc_mean}')
    print(f'Avg Accuracy: {acc_mean.mean()}')
    print(f'Q_values: {model.q_values}')

    chk = check_acc(acc_mean)
    if chk:
        print('Success!')

        while True:
            model = RLmodel(pair_pats, rewards)
            model.simulate()
            if check_acc(np.array(model.acc_list)):
                actions = np.array(model.actions)
                break
        
        write_excel(pair_pats, order_pats, actions, rewards, game_pat)
    else:
        print('Failed...')

if __name__ == '__main__':
    main()
