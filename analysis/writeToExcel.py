#!/usr/bin/env python
# -*- coding: utf-8 -*-

# NOTE: 被験者番号とパターンについては、1-indexedで記述している

# imports
# standard library
import glob
from collections import defaultdict
from tqdm import tqdm

# third party
import numpy as np
import polars as pl
from openpyxl import load_workbook

# local

# random seed
seed = 0
np.random.seed(seed)

# variables
subj_size = 15
pat_size = 2
block_size = 3
seq_size_slf = 12
seq_size_test_ss = 6
img_size = 3

rept_num = 10**4

result_file = '/Users/strix_uralensis/Documents/Experiment/analysis/result_ss_pl.xlsx'

def write_to_excel(p_softmax, log_like_softmax, p_matching, log_like_matching):
    out_book  = load_workbook(filename=result_file)
    out_sheet = out_book[out_book.sheetnames[0]]
    for subj in range(1, subj_size+1):
        for pat in range(1, pat_size+1):
            out_sheet.cell(pat_size*subj+pat+2, 1, value=subj)
            out_sheet.cell(pat_size*subj+pat+2, 2, value=pat)
            out_sheet.cell(pat_size*subj+pat+2, 3, value=p_softmax)
            out_sheet.cell(pat_size*subj+pat+2, 4, value=log_like_softmax)
            out_sheet.cell(pat_size*subj+pat+2, 5, value=p_matching)
            out_sheet.cell(pat_size*subj+pat+2, 6, value=log_like_matching)
    out_book.save(result_file)


def main():
    pass

if __name__ == '__main__':
    main()